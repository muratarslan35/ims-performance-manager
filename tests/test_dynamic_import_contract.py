import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from flask_migrate import upgrade
from openpyxl import Workbook

from app import create_app
from app.extensions import db
from app.models import IMSSummary, IMSUpload, Product, Representative, Target
from app.services.alias_service import AliasService
from app.services.competition_import_service import CompetitionImportService
from app.services.dynamic_import_contract import WorkbookSemanticLocator
from app.services.dynamic_import_refinement import FlexibleSemanticLocator
from app.services.ims_import_service import IMSImportService
from app.services.official_aggregate_service import OfficialAggregateService


class Config:
    TESTING = True
    SECRET_KEY = "test-secret"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    UPLOAD_FOLDER = Path(tempfile.gettempdir()) / "dynamic-import-test-uploads"
    REPORT_FOLDER = Path(tempfile.gettempdir()) / "dynamic-import-test-reports"
    BACKUP_FOLDER = Path(tempfile.gettempdir()) / "dynamic-import-test-backups"
    LOG_FOLDER = Path(tempfile.gettempdir()) / "dynamic-import-test-logs"


@pytest.fixture()
def dynamic_env():
    temp_dir = tempfile.TemporaryDirectory()
    db_path = Path(temp_dir.name) / "dynamic.db"
    config = type("RuntimeConfig", (Config,), {"SQLALCHEMY_DATABASE_URI": f"sqlite:///{db_path}"})
    app = create_app(config)
    ctx = app.app_context()
    ctx.push()
    migrations_dir = str(Path(__file__).resolve().parents[1] / "migrations")
    upgrade(directory=migrations_dir)

    product = Product.query.filter_by(product_code="TRAVAZOL").first()
    if product is None:
        product = Product(product_code="TRAVAZOL", product_name="Travazol", is_active=True)
        db.session.add(product)
    else:
        product.is_active = True

    rep = Representative(
        rep_code="DYNAMIC-REP",
        rep_name="DYNAMIC TEST REP",
        region="901",
        city="DIYARBAKIR",
        active=True,
    )
    db.session.add(rep)
    db.session.flush()

    upload = IMSUpload(
        file_name="renamed.xlsx",
        year=2038,
        month=2,
        week_number=7,
        quarter="Q1",
        status="COMPLETED",
        completed_at=datetime.utcnow(),
    )
    db.session.add(upload)
    db.session.flush()

    target = Target(
        year=2038,
        month=2,
        quarter="Q1",
        representative_id=rep.id,
        product_id=product.id,
        tl_target=1.0,
        unit_target=1.0,
    )
    summary = IMSSummary(
        upload_id=upload.id,
        year=2038,
        month=2,
        quarter="Q1",
        representative_id=rep.id,
        product_id=product.id,
        target_tl=1.0,
        target_unit=1.0,
    )
    db.session.add_all([target, summary])
    db.session.commit()
    AliasService.clear()
    AliasService.warmup()

    try:
        yield temp_dir, app, upload, product, rep
    finally:
        db.session.remove()
        ctx.pop()
        temp_dir.cleanup()


def _shifted_semantic_workbook():
    balance_noise = [["noise", None, None, None, None, None, None, None, None, None, None, None] for _ in range(16)]
    balance_header = [
        None, "ŞUBAT HEDEF TL", "TRAVAZOL",
        None, "ŞUBAT TL BAKİYE", "TRAVAZOL",
        "ŞUBAT KUTU BAKİYE", "TRAVAZOL",
        "TTS ISMI", "BOLGE",
        "ŞUBAT ÇIKIŞ TL", "TRAVAZOL",
    ]
    balance_rows = [
        [None, None, 1005.0, None, None, 600.0, None, 60.0, "NATIONAL", None, None, 250.0],
        [None, None, 1005.0, None, None, 600.0, None, 60.0, "901 DIYARBAKIR", None, None, 250.0],
        [None, None, 1005.0, None, None, 600.0, None, 60.0, "DYNAMIC TEST REP", "901 DIYARBAKIR", None, 250.0],
    ]
    balance = pd.DataFrame(balance_noise + [balance_header] + balance_rows)

    weekly_noise = [["metadata", None, None, None, None, None, None, None, None, None] for _ in range(19)]
    weekly_sections = [
        "TTS ISMI", "BOLGE",
        "1-20 ŞUBAT TL ÇIKIŞI", None,
        "1-20 ŞUBAT KUTU ÇIKIŞI", None,
        "7. HAFTA TL ÇIKIŞI", None,
        "7. HAFTA KUTU ÇIKIŞI", None,
    ]
    weekly_products = [
        None, None, None, "TRAVAZOL", None, "TRAVAZOL",
        None, "TRAVAZOL", None, "TRAVAZOL",
    ]
    weekly_rows = [
        ["NATIONAL", None, None, 250.0, None, 17.25, None, 99.0, None, 9.0],
        ["901 DIYARBAKIR", None, None, 250.0, None, 17.25, None, 99.0, None, 9.0],
        ["DYNAMIC TEST REP", "901 DIYARBAKIR", None, 250.0, None, 17.25, None, 99.0, None, 9.0],
    ]
    weekly = pd.DataFrame(weekly_noise + [weekly_sections, weekly_products] + weekly_rows)

    return {
        "TARGET SOURCE RENAMED AND MOVED": balance,
        "CURRENT PERIOD SALES RENAMED AND MOVED": weekly,
        "OPTIONAL EMPTY EXTRA": pd.DataFrame(),
    }


def test_renamed_shifted_reordered_sources_keep_same_business_result(dynamic_env):
    _temp, _app, upload, product, rep = dynamic_env
    service = IMSImportService("unused.xlsx")
    service.upload = upload
    service.workbook = _shifted_semantic_workbook()

    service.bootstrap_vacancy_representatives_from_balance()
    service.apply_balance_summary(2038, 2)
    service.apply_weekly_sales_summary(2038, 2)
    service.persist_national_dashboard_metrics(2038, 2)
    db.session.commit()

    target = Target.query.filter_by(
        year=2038, month=2, representative_id=rep.id, product_id=product.id
    ).one()
    summary = IMSSummary.query.filter_by(
        year=2038, month=2, representative_id=rep.id, product_id=product.id
    ).one()
    assert target.tl_target == pytest.approx(1005.0)
    assert target.unit_target == pytest.approx(100.0)
    assert target.tl_realization == pytest.approx(250.0)
    assert target.unit_realization == pytest.approx(17.25)
    assert summary.tl == pytest.approx(250.0)
    assert summary.unit == pytest.approx(17.25)

    national = OfficialAggregateService.product_totals(2038, 2, "NATIONAL")
    region = OfficialAggregateService.product_totals(2038, 2, "901")
    assert national and region
    assert national[0]["target_tl"] == pytest.approx(1005.0)
    assert national[0]["target_unit"] == pytest.approx(100.5)
    assert national[0]["actual_tl"] == pytest.approx(250.0)
    assert national[0]["actual_unit"] == pytest.approx(17.25)
    assert region[0]["target_tl"] == pytest.approx(1005.0)
    assert region[0]["actual_unit"] == pytest.approx(17.25)


def test_later_tl_only_ims_keeps_unit_target_seen_in_any_prior_week(dynamic_env):
    _temp, _app, upload, product, rep = dynamic_env
    service = IMSImportService("unused.xlsx")
    service.upload = upload
    service.workbook = {
        "MART HEDEF": pd.DataFrame([["MART HEDEF TL", "TRAVAZOL"]])
    }
    service._prior_month_unit_targets = {(rep.id, product.id): 123.0}
    target = Target.query.filter_by(
        year=2038, month=2, representative_id=rep.id, product_id=product.id
    ).one()
    target.unit_target = 0.0
    summary = IMSSummary.query.filter_by(
        upload_id=upload.id, representative_id=rep.id, product_id=product.id
    ).one()
    summary.target_unit = 0.0
    db.session.flush()

    assert service.restore_locked_monthly_unit_targets(2038, 2) == 1
    assert target.unit_target == 123.0
    assert summary.target_unit == 123.0


def test_explicit_box_source_keeps_numeric_zero_as_real_value(dynamic_env):
    _temp, _app, upload, product, rep = dynamic_env
    service = IMSImportService("unused.xlsx")
    service.upload = upload
    service.workbook = {
        "BAKİYE": pd.DataFrame([["MART KUTU BAKİYE", "TRAVAZOL"], [0, 0]])
    }
    service._prior_month_unit_targets = {(rep.id, product.id): 123.0}
    target = Target.query.filter_by(
        year=2038, month=2, representative_id=rep.id, product_id=product.id
    ).one()
    target.unit_target = 0.0
    summary = IMSSummary.query.filter_by(
        upload_id=upload.id, representative_id=rep.id, product_id=product.id
    ).one()
    summary.target_unit = 0.0
    db.session.flush()

    assert service.restore_locked_monthly_unit_targets(2038, 2) == 0
    assert target.unit_target == 0.0
    assert summary.target_unit == 0.0


def test_unit_target_lock_does_not_resurrect_rows_absent_from_current_roster(dynamic_env):
    _temp, _app, upload, product, rep = dynamic_env
    service = IMSImportService("unused.xlsx")
    service.upload = upload
    service.workbook = {
        "MART HEDEF": pd.DataFrame([["MART HEDEF TL", "TRAVAZOL"]])
    }
    service._prior_month_unit_targets = {(rep.id, product.id): 123.0}
    Target.query.filter_by(
        year=2038, month=2, representative_id=rep.id, product_id=product.id
    ).delete()
    db.session.flush()

    assert service.restore_locked_monthly_unit_targets(2038, 2) == 0
    assert Target.query.filter_by(
        year=2038, month=2, representative_id=rep.id, product_id=product.id
    ).count() == 0


def test_equal_authoritative_sources_fail_closed(dynamic_env):
    _temp, _app, upload, _product, _rep = dynamic_env
    workbook = _shifted_semantic_workbook()
    workbook["SECOND EQUALLY STRONG TARGET SOURCE"] = workbook["TARGET SOURCE RENAMED AND MOVED"].copy()

    service = IMSImportService("unused.xlsx")
    service.upload = upload
    service.workbook = workbook

    with pytest.raises(ValueError, match="belirsiz kaynak"):
        WorkbookSemanticLocator(service).locate("balance", required=True)


def test_weekly_source_prefers_resolved_representative_semantics_over_brick_matrix(dynamic_env):
    _temp, _app, upload, _product, _rep = dynamic_env
    weekly = _shifted_semantic_workbook()["CURRENT PERIOD SALES RENAMED AND MOVED"].copy()
    brick = weekly.copy()
    section_row = 19
    brick.iloc[section_row, 0] = "IAM BRICK"
    brick.iloc[section_row + 2, 0] = "DYNAMIC TEST REP"
    brick.iloc[section_row + 3, 0] = "DYNAMIC TEST REP"
    brick.iloc[section_row + 4, 0] = "DYNAMIC TEST REP"

    service = IMSImportService("unused.xlsx")
    service.upload = upload
    service.workbook = {
        "LOCATION MATRIX": brick,
        "REPRESENTATIVE CUMULATIVE SOURCE": weekly,
    }

    profile = FlexibleSemanticLocator(service).locate("weekly", required=True)

    assert profile.sheet_name == "REPRESENTATIVE CUMULATIVE SOURCE"
    assert profile.representative_column == 0


def test_competition_header_and_dimension_positions_can_move(dynamic_env):
    temp_dir, _app, upload, _product, _rep = dynamic_env
    file_path = Path(temp_dir.name) / "shifted-competition.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AYLIK REKABET KUTU"
    for row in range(1, 25):
        sheet.cell(row=row, column=1, value=f"noise {row}")
    sheet.cell(row=24, column=1, value="GRUP A")
    sheet.cell(row=24, column=3, value="GRUP A")
    sheet.cell(row=25, column=1, value="RAKIP A")
    sheet.cell(row=25, column=2, value="TERRITORY")
    sheet.cell(row=25, column=3, value="RAKIP B")
    sheet.cell(row=25, column=4, value="TTS ISMI")
    sheet.cell(row=25, column=5, value="IAM BRICK")
    sheet.cell(row=26, column=1, value=10)
    sheet.cell(row=26, column=2, value="901 DIYARBAKIR")
    sheet.cell(row=26, column=3, value=20)
    sheet.cell(row=26, column=4, value="DYNAMIC TEST REP")
    sheet.cell(row=26, column=5, value="DIYARBAKIR MERKEZ")
    sheet.cell(row=27, column=1, value=30)
    sheet.cell(row=27, column=2, value="901 DIYARBAKIR")
    sheet.cell(row=27, column=3, value=40)
    sheet.cell(row=27, column=4, value="OTHER REP")
    sheet.cell(row=27, column=5, value="DIYARBAKIR BATI")
    workbook.save(file_path)

    service = CompetitionImportService(
        file_path=str(file_path),
        upload_id=upload.id,
        year=2038,
        month=2,
        week_number=7,
    )
    service.load_workbook(str(file_path))
    try:
        structure = service._parse_sheet_structure("AYLIK REKABET KUTU")
    finally:
        service._workbook.close()
        service._workbook = None

    assert structure["header_row"] == 25
    assert structure["territory_column"] == 2
    assert structure["subterritory_column"] == 5
    assert set(structure["product_columns"]) == {1, 3}
    assert structure["data_start_row"] == 26

