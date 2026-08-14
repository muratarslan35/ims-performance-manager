import tempfile
from pathlib import Path

from flask_migrate import upgrade
from openpyxl import Workbook

from app import create_app
from app.extensions import db
from app.models import CompetitionData, IMSRawData, IMSSummary, IMSUpload, Product, Representative, RepresentativeBrickAssignment, Target
from app.services.representative_market_service import RepresentativeMarketService


def test_representative_market_analysis_is_brick_scoped_and_keeps_seven_products():
    temporary = tempfile.TemporaryDirectory()
    database_path = Path(temporary.name) / "representative-market.db"

    class Config:
        TESTING = True
        SECRET_KEY = "test-secret"
        SQLALCHEMY_DATABASE_URI = f"sqlite:///{database_path}"
        SQLALCHEMY_TRACK_MODIFICATIONS = False
        UPLOAD_FOLDER = Path(temporary.name) / "uploads"
        REPORT_FOLDER = Path(temporary.name) / "reports"
        BACKUP_FOLDER = Path(temporary.name) / "backups"
        LOG_FOLDER = Path(temporary.name) / "logs"

    application = create_app(Config)
    try:
        with application.app_context():
            upgrade(directory=str(Path(__file__).resolve().parents[1] / "migrations"))
            representative = Representative(rep_code="R1", rep_name="Temsilci Bir", active=True)
            other = Representative(rep_code="R2", rep_name="Temsilci İki", active=True)
            names = ("Travazol", "Monurol", "Acnemix", "Mixovul", "Stiderm", "Brimoder", "Fentivag")
            products = [
                Product(product_code=name.upper(), product_name=name, display_order=index, is_active=True)
                for index, name in enumerate(names)
            ]
            db.session.add_all([representative, other, *products])
            db.session.flush()
            upload = IMSUpload(file_name="august.xlsx", year=2026, month=8, quarter="Q3", status="COMPLETED")
            previous_upload = IMSUpload(file_name="july.xlsx", year=2026, month=7, quarter="Q3", status="COMPLETED")
            db.session.add_all([upload, previous_upload])
            db.session.flush()
            db.session.add_all(
                [
                    RepresentativeBrickAssignment(representative_id=representative.id, year=2026, month=8, brick="BRICK A"),
                    RepresentativeBrickAssignment(representative_id=representative.id, year=2026, month=8, brick="BRICK C"),
                    RepresentativeBrickAssignment(representative_id=other.id, year=2026, month=8, brick="BRICK B"),
                    IMSSummary(upload_id=upload.id, representative_id=representative.id, product_id=products[0].id, year=2026, month=8, quarter="Q3", tl=100, unit=10),
                    IMSSummary(upload_id=previous_upload.id, representative_id=representative.id, product_id=products[0].id, year=2026, month=7, quarter="Q3", tl=50, unit=5),
                ]
            )

            def competition(brick, product_name, metric_type, value):
                return CompetitionData(
                    upload_id=upload.id, year=2026, month=8, sheet_name=f"REKABET {metric_type}",
                    period_type="MONTHLY", territory="101", subterritory=brick,
                    product_group="TRAVAZOL GRUBU", product_name=product_name,
                    metric_type=metric_type, metric_value=value, source_row=1,
                )

            db.session.add_all(
                [
                    competition("BRICK A", "TRAVAZOL", "TL", 100),
                    competition("BRICK A", "RAKIP A", "TL", 300),
                    competition("BRICK A", "TRAVAZOL", "UNIT", 10),
                    competition("BRICK A", "RAKIP A", "UNIT", 30),
                    competition("BRICK A", "TRAVAZOL GRUP SUBTOTAL", "UNIT", 35),
                    competition("BRICK C", "TRAVAZOL", "UNIT", 5),
                    competition("BRICK C", "RAKIP C", "UNIT", 95),
                    competition("BRICK B", "RAKIP B", "UNIT", 900),
                ]
            )
            db.session.add(CompetitionData(
                upload_id=previous_upload.id, year=2026, month=7, sheet_name="REKABET UNIT",
                period_type="MONTHLY", territory="101", subterritory="BRICK A",
                product_group="TRAVAZOL GRUBU", product_name="RAKIP A",
                metric_type="UNIT", metric_value=20, source_row=1,
            ))
            db.session.commit()

            result = RepresentativeMarketService(representative, 2026, 8).build()

            assert result["scope"] == "brick"
            assert len(result["rows"]) == 7
            travazol = result["rows"][0]
            assert travazol["actual_unit"] == 10
            assert travazol["market_unit"] == 140
            assert travazol["competitor_unit"] == 130
            assert travazol["share_percent"] == 7.1
            assert travazol["has_previous"] is True
            assert travazol["previous_actual_unit"] == 5
            assert travazol["actual_change_unit"] == 5
            assert travazol["actual_change_percent"] == 100.0
            assert travazol["competitor_change_unit"] == 115
            assert travazol["rivals"] == [
                {"name": "RAKIP C", "unit": 95.0},
                {"name": "RAKIP A", "unit": 30.0},
            ]
            assert [row["brick"] for row in result["brick_rows"]] == ["BRICK C", "BRICK A"]
            assert result["brick_rows"][0]["attention"] == "critical"
            assert result["brick_rows"][0]["threats"][0]["product_name"] == "Travazol"
            assert len(result["brick_product_rows"]) == 2
            assert result["brick_product_rows"][0]["product_name"] == "Travazol"
            assert result["brick_product_rows"][0]["company_unit"] == 10
            assert result["brick_product_rows"][0]["market_products"] == [
                {"name": "TRAVAZOL", "unit": 10.0, "is_company": True, "share_percent": 25.0, "realization_percent": None},
                {"name": "RAKIP A", "unit": 30.0, "is_company": False, "share_percent": 75.0, "realization_percent": None},
            ]
            assert all("SUBTOTAL" not in item["name"] for item in result["brick_product_rows"][0]["market_products"])
            assert result["brick_product_rows"][0]["market_unit"] == 40
            assert result["brick_product_rows"][0]["subtotal_unit"] == 35
            assert result["brick_product_rows"][0]["group_total_unit"] == 35
    finally:
        temporary.cleanup()


def test_representative_grained_competition_and_raw_brick_market_are_combined():
    temporary = tempfile.TemporaryDirectory()
    database_path = Path(temporary.name) / "representative-market-grains.db"

    class Config:
        TESTING = True
        SECRET_KEY = "test-secret"
        SQLALCHEMY_DATABASE_URI = f"sqlite:///{database_path}"
        SQLALCHEMY_TRACK_MODIFICATIONS = False
        UPLOAD_FOLDER = Path(temporary.name) / "uploads"
        REPORT_FOLDER = Path(temporary.name) / "reports"
        BACKUP_FOLDER = Path(temporary.name) / "backups"
        LOG_FOLDER = Path(temporary.name) / "logs"

    application = create_app(Config)
    try:
        with application.app_context():
            upgrade(directory=str(Path(__file__).resolve().parents[1] / "migrations"))
            representative = Representative(rep_code="R1", rep_name="Murat Arslan", active=True)
            product = Product(product_code="TRAVAZOL", product_name="Travazol", display_order=1, is_active=True)
            db.session.add_all([representative, product])
            db.session.flush()
            upload = IMSUpload(file_name="january.xlsx", year=2026, month=1, quarter="Q1", status="COMPLETED")
            db.session.add(upload)
            db.session.flush()
            db.session.add_all([
                RepresentativeBrickAssignment(
                    representative_id=representative.id, year=2026, month=1, brick="MARDIN MERKEZ"
                ),
                IMSSummary(
                    upload_id=upload.id, representative_id=representative.id, product_id=product.id,
                    year=2026, month=1, quarter="Q1", unit=100,
                ),
                CompetitionData(
                    upload_id=upload.id, year=2026, month=1, sheet_name="TTS REKABET",
                    period_type="MONTHLY", territory="901 DIYARBAKIR", subterritory="MURAT ARSLAN",
                    product_group="TRAVAZOL GRUP", product_name="TRAVAZOL", metric_type="UNIT",
                    metric_value=100, source_row=1,
                ),
                CompetitionData(
                    upload_id=upload.id, year=2026, month=1, sheet_name="AYLIK REKABET KUTU",
                    period_type="MONTHLY", territory="901 DIYARBAKIR", subterritory="MARDIN MERKEZ",
                    product_group="TRAVAZOL GRUP", product_name="TRAVAZOL", metric_type="UNIT",
                    metric_value=40, source_row=5,
                ),
                CompetitionData(
                    upload_id=upload.id, year=2026, month=1, sheet_name="AYLIK REKABET KUTU",
                    period_type="MONTHLY", territory="901 DIYARBAKIR", subterritory="MARDIN MERKEZ",
                    product_group="TRAVAZOL GRUP", product_name="TRAVAZOL GRUBU Subtotal", metric_type="UNIT",
                    metric_value=160, source_row=5,
                ),
                CompetitionData(
                    upload_id=upload.id, year=2026, month=1, sheet_name="AYLIK REKABET KUTU",
                    period_type="MONTHLY", territory="901 DIYARBAKIR", subterritory="MARDIN MERKEZ",
                    product_group="TRAVAZOL GRUP", product_name="RAKIP A", metric_type="UNIT",
                    metric_value=80, source_row=5,
                ),
                CompetitionData(
                    upload_id=upload.id, year=2026, month=1, sheet_name="AYLIK REKABET KUTU",
                    period_type="MONTHLY", territory="901 DIYARBAKIR", subterritory="MARDIN MERKEZ",
                    product_group="TRAVAZOL GRUP", product_name="RAKIP B", metric_type="UNIT",
                    metric_value=40, source_row=5,
                ),
                CompetitionData(
                    upload_id=upload.id, year=2026, month=1, sheet_name="TTS REKABET",
                    period_type="MONTHLY", territory="901 DIYARBAKIR", subterritory="MURAT ARSLAN",
                    product_group="TRAVAZOL GRUP", product_name="RAKIP A", metric_type="UNIT",
                    metric_value=300, source_row=1,
                ),
                CompetitionData(
                    upload_id=upload.id, year=2026, month=1, sheet_name="TTS REKABET",
                    period_type="MONTHLY", territory="901 DIYARBAKIR", subterritory="MURAT ARSLAN",
                    product_group="TRAVAZOL GRUP", product_name="RAKIP B", metric_type="UNIT",
                    metric_value=100, source_row=1,
                ),
                IMSRawData(
                    upload_id=upload.id, year=2026, month=1, quarter="Q1", source_row=1,
                    sheet_name="BRICK SATIS", sheet_type="brick_sales",
                    representative_id=representative.id, product_id=product.id, representative="MURAT ARSLAN",
                    brick="MARDIN MERKEZ", product="TRAVAZOL", unit=40, raw_json="{}",
                ),
                IMSRawData(
                    upload_id=upload.id, year=2026, month=1, quarter="Q1", source_row=1,
                    sheet_name="REKABET KUTU", sheet_type="competition_box",
                    representative_id=representative.id, product_id=product.id, representative="MURAT ARSLAN",
                    brick="MARDIN MERKEZ", product="TRAVAZOL", unit=160, raw_json="{}",
                ),
            ])
            db.session.commit()

            result = RepresentativeMarketService(representative, 2026, 1).build()

            travazol = result["rows"][0]
            assert travazol["market_unit"] == 500
            assert travazol["competitor_unit"] == 400
            assert travazol["rivals"] == [
                {"name": "RAKIP A", "unit": 300.0},
                {"name": "RAKIP B", "unit": 100.0},
            ]
            assert result["brick_rows"][0]["brick"] == "MARDIN MERKEZ"
            assert result["brick_rows"][0]["company_unit"] == 40
            assert result["brick_rows"][0]["market_unit"] == 160
            market_products = result["brick_product_rows"][0]["market_products"]
            assert [(item["name"], item["unit"]) for item in market_products] == [
                ("TRAVAZOL", 40.0),
                ("RAKIP A", 80.0),
                ("RAKIP B", 40.0),
            ]
            assert result["brick_product_rows"][0]["group_total_unit"] == 160
            assert all("SUBTOTAL" not in item["name"].upper() for item in market_products)
            assert sum(item["unit"] for item in market_products if not item["is_company"]) == 120

            # A newer completed snapshot is authoritative. A competitor that
            # disappeared from it must not leak forward from the older upload.
            newer_upload = IMSUpload(
                file_name="january-revised.xlsx", year=2026, month=1, quarter="Q1", status="COMPLETED"
            )
            db.session.add(newer_upload)
            db.session.commit()
            revised = RepresentativeMarketService(representative, 2026, 1).build()
            assert revised["has_competition"] is False
            assert revised["brick_rows"] == []
            assert revised["brick_product_rows"] == []
    finally:
        temporary.cleanup()


def test_shared_brick_source_is_visible_to_both_reps_without_duplicate_facts():
    temporary = tempfile.TemporaryDirectory()
    database_path = Path(temporary.name) / "shared-brick.db"

    class Config:
        TESTING = True
        SECRET_KEY = "test-secret"
        SQLALCHEMY_DATABASE_URI = f"sqlite:///{database_path}"
        SQLALCHEMY_TRACK_MODIFICATIONS = False
        UPLOAD_FOLDER = Path(temporary.name) / "uploads"
        REPORT_FOLDER = Path(temporary.name) / "reports"
        BACKUP_FOLDER = Path(temporary.name) / "backups"
        LOG_FOLDER = Path(temporary.name) / "logs"

    application = create_app(Config)
    try:
        with application.app_context():
            upgrade(directory=str(Path(__file__).resolve().parents[1] / "migrations"))
            ahmet = Representative(rep_code="A1", rep_name="Ahmet", active=True)
            mehmet = Representative(rep_code="M1", rep_name="Mehmet", active=True)
            product = Product(product_code="TRAVAZOL", product_name="Travazol", display_order=1, is_active=True)
            db.session.add_all([ahmet, mehmet, product])
            db.session.flush()
            upload = IMSUpload(file_name="shared.xlsx", year=2026, month=1, quarter="Q1", status="COMPLETED")
            db.session.add(upload)
            db.session.flush()
            db.session.add_all([
                RepresentativeBrickAssignment(representative_id=ahmet.id, year=2026, month=1, brick="ORTAK BRICK"),
                RepresentativeBrickAssignment(representative_id=mehmet.id, year=2026, month=1, brick="ORTAK BRICK"),
                IMSRawData(
                    upload_id=upload.id, year=2026, month=1, quarter="Q1", source_row=10,
                    sheet_name="1001 BRICK SATIS", sheet_type="brick_sales",
                    representative_id=ahmet.id, product_id=product.id, representative="AHMET",
                    brick="ORTAK BRICK", product="TRAVAZOL", unit=125, tl=2500, raw_json="{}",
                ),
                IMSSummary(
                    upload_id=upload.id, representative_id=ahmet.id, product_id=product.id,
                    year=2026, month=1, quarter="Q1", unit=125, tl=2500,
                ),
                Target(
                    representative_id=ahmet.id, product_id=product.id, year=2026, month=1,
                    quarter="Q1", unit_target=200, tl_target=4000,
                ),
                CompetitionData(
                    upload_id=upload.id, year=2026, month=1, sheet_name="AYLIK REKABET KUTU",
                    period_type="MONTHLY", territory="101", subterritory="ORTAK BRICK",
                    product_group="TRAVAZOL GRUP", product_name="TRAVAZOL", metric_type="UNIT",
                    metric_value=125, source_row=10,
                ),
                CompetitionData(
                    upload_id=upload.id, year=2026, month=1, sheet_name="AYLIK REKABET KUTU",
                    period_type="MONTHLY", territory="101", subterritory="ORTAK BRICK",
                    product_group="TRAVAZOL GRUP", product_name="RAKIP X", metric_type="UNIT",
                    metric_value=75, source_row=10,
                ),
            ])
            db.session.commit()

            ahmet_result = RepresentativeMarketService(ahmet, 2026, 1).build()
            mehmet_result = RepresentativeMarketService(mehmet, 2026, 1).build()

            for result in (ahmet_result, mehmet_result):
                assert result["rows"][0]["actual_unit"] == 125
                assert result["rows"][0]["target_unit"] == 200
                assert result["brick_product_rows"][0]["market_products"] == [
                    {"name": "TRAVAZOL", "unit": 125.0, "is_company": True, "share_percent": 62.5, "realization_percent": 62.5},
                    {"name": "RAKIP X", "unit": 75.0, "is_company": False, "share_percent": 37.5, "realization_percent": None},
                ]

            assert IMSRawData.query.filter_by(sheet_type="brick_sales").count() == 1
            assert IMSSummary.query.count() == 1
            assert Target.query.count() == 1
    finally:
        temporary.cleanup()


def test_legacy_upload_reads_named_brick_rivals_directly_from_source_workbook():
    temporary = tempfile.TemporaryDirectory()
    root = Path(temporary.name)
    database_path = root / "legacy-brick-rivals.db"

    class Config:
        TESTING = True
        SECRET_KEY = "test-secret"
        SQLALCHEMY_DATABASE_URI = f"sqlite:///{database_path}"
        SQLALCHEMY_TRACK_MODIFICATIONS = False
        UPLOAD_FOLDER = root / "uploads"
        REPORT_FOLDER = root / "reports"
        BACKUP_FOLDER = root / "backups"
        LOG_FOLDER = root / "logs"

    application = create_app(Config)
    try:
        with application.app_context():
            upgrade(directory=str(Path(__file__).resolve().parents[1] / "migrations"))
            representative = Representative(rep_code="LEGACY", rep_name="Legacy Temsilci", active=True)
            product = Product(product_code="TRAVAZOL", product_name="Travazol", display_order=1, is_active=True)
            db.session.add_all([representative, product])
            db.session.flush()
            upload = IMSUpload(file_name="legacy-source.xlsx", year=2026, month=1, quarter="Q1", status="COMPLETED")
            db.session.add(upload)
            db.session.flush()
            db.session.add_all([
                RepresentativeBrickAssignment(representative_id=representative.id, year=2026, month=1, brick="MARDIN BATI"),
                IMSRawData(
                    upload_id=upload.id, year=2026, month=1, quarter="Q1", source_row=5,
                    sheet_name="1001 BRICK SATIS", sheet_type="brick_sales",
                    representative_id=representative.id, product_id=product.id,
                    representative="LEGACY TEMSILCI", brick="MARDIN BATI", product="TRAVAZOL",
                    unit=40, raw_json="{}",
                ),
                IMSRawData(
                    upload_id=upload.id, year=2026, month=1, quarter="Q1", source_row=5,
                    sheet_name="REKABET KUTU", sheet_type="competition_box",
                    representative_id=representative.id, product_id=product.id,
                    representative="LEGACY TEMSILCI", brick="MARDIN BATI", product="TRAVAZOL",
                    unit=120, raw_json="{}",
                ),
            ])
            db.session.commit()

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "AYLIK REKABET KUTU"
            sheet.cell(1, 1, "OCAK 2026")
            sheet.cell(2, 6, "TRAVAZOL GRUP")
            for column, value in enumerate(
                ("BOLGE", "NATIONAL", "IAM BRICK", "1. TTS ISMI", "2. TTS ISMI", "TRAVAZOL", "RAKIP X"),
                start=1,
            ):
                sheet.cell(3, column, value)
            for column, value in enumerate(
                ("901", "DIYARBAKIR", "MARDIN BATI", "LEGACY TEMSILCI", None, 40, 80),
                start=1,
            ):
                sheet.cell(5, column, value)
            workbook.save(Config.UPLOAD_FOLDER / upload.file_name)

            RepresentativeMarketService._workbook_competition_cache.clear()
            result = RepresentativeMarketService(representative, 2026, 1).build()
            market_products = result["brick_product_rows"][0]["market_products"]

            assert [(item["name"], item["unit"]) for item in market_products] == [
                ("TRAVAZOL", 40.0),
                ("RAKIP X", 80.0),
            ]
            assert all(item["name"] != "Rakip toplamı" for item in market_products)
    finally:
        RepresentativeMarketService._workbook_competition_cache.clear()
        temporary.cleanup()
