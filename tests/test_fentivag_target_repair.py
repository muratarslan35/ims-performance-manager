from pathlib import Path

from openpyxl import Workbook

from repair_fentivag_targets import _is_rep_row, extract_source_records


def _build_workbook(path: Path):
    wb = Workbook()
    bakiye = wb.active
    bakiye.title = "BAKİYE"
    bakiye.append([None, None, None, None, None, None])
    bakiye.append([None, "OCAK HEDEF TL", None, "FENTIVAG", "OCAK MF siz KUTU BAKİYE", "FENTIVAG"])
    bakiye.append([None, "NATIONAL", None, 300.0, None, 20.0])
    bakiye.append(["901 DIYARBAKIR", "901 DIYARBAKIR", None, 300.0, None, 20.0])
    bakiye.append(["901 DIYARBAKIR", "MURAT ARSLAN", None, 100.0, None, 6.0])
    bakiye.append(["901 DIYARBAKIR", "DIYARBAKIR BOS", None, 200.0, None, 14.0])

    weekly = wb.create_sheet("TTS HAFTALIK ÇIKIŞLARI")
    weekly.append([None, None, "1-18 OCAK TL ÇIKIŞI", None, "1-18 OCAK KUTU ÇIKIŞI", None])
    weekly.append([None, None, None, "FENTIVAG", None, "FENTIVAG"])
    weekly.append([None, "NATIONAL", None, 30.0, None, 2.0])
    weekly.append(["901 DIYARBAKIR", "901 DIYARBAKIR", None, 30.0, None, 2.0])
    weekly.append(["901 DIYARBAKIR", "MURAT ARSLAN", None, 10.0, None, 1.0])
    weekly.append(["901 DIYARBAKIR", "DIYARBAKIR BOS", None, 20.0, None, 1.0])

    wb.save(path)
    wb.close()


def test_fentivag_repair_uses_workbook_box_equation_not_price(tmp_path):
    source = tmp_path / "ocak.xlsx"
    _build_workbook(source)

    records = extract_source_records(source)
    by_rep = {row["representative"]: row for row in records}

    assert len(records) == 2
    assert by_rep["MURAT ARSLAN"]["target_tl"] == 100.0
    assert by_rep["MURAT ARSLAN"]["actual_unit"] == 1.0
    assert by_rep["MURAT ARSLAN"]["unit_target"] == 7.0
    assert by_rep["DIYARBAKIR BOS"]["target_tl"] == 200.0
    assert by_rep["DIYARBAKIR BOS"]["unit_target"] == 15.0


def test_representative_rows_remain_authoritative_when_national_differs(tmp_path):
    source = tmp_path / "ocak-national-differs.xlsx"
    _build_workbook(source)

    from openpyxl import load_workbook

    workbook = load_workbook(source)
    bakiye = workbook["BAKİYE"]
    weekly = workbook["TTS HAFTALIK ÇIKIŞLARI"]
    # Intentionally make NATIONAL disagree with the two representative rows.
    bakiye.cell(3, 4, 999999.0)
    bakiye.cell(3, 6, 999.0)
    weekly.cell(3, 4, 0.0)
    weekly.cell(3, 6, 0.0)
    workbook.save(source)
    workbook.close()

    records = extract_source_records(source)
    by_rep = {row["representative"]: row for row in records}

    assert len(records) == 2
    assert sum(row["target_tl"] for row in records) == 300.0
    assert sum(row["unit_target"] for row in records) == 22.0
    assert by_rep["MURAT ARSLAN"]["target_tl"] == 100.0
    assert by_rep["DIYARBAKIR BOS"]["target_tl"] == 200.0


def test_vacant_position_is_a_valid_target_representative_row():
    assert _is_rep_row("901 DIYARBAKIR", "DIYARBAKIR BOS") is True
    assert _is_rep_row("901 DIYARBAKIR", "DIYARBAKIR BOS KADRO") is True
    assert _is_rep_row("901 DIYARBAKIR", "901 DIYARBAKIR") is False
