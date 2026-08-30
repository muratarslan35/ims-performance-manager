from pathlib import Path

from openpyxl import Workbook

from app import create_app


def _app(tmp_path):
    class Config:
        TESTING = True
        SECRET_KEY = "region-balance-unit"
        SQLALCHEMY_DATABASE_URI = f"sqlite:///{tmp_path / 'region-balance-unit.db'}"
        SQLALCHEMY_TRACK_MODIFICATIONS = False
        WTF_CSRF_ENABLED = False
        LOGIN_DISABLED = True
        UPLOAD_FOLDER = Path(tmp_path) / "uploads"
        REPORT_FOLDER = Path(tmp_path) / "reports"
        BACKUP_FOLDER = Path(tmp_path) / "backups"
        LOG_FOLDER = Path(tmp_path) / "logs"
        TEMP_FOLDER = Path(tmp_path) / "temp"
    return create_app(Config)


def test_region_balance_units_reads_mf_siz_kutu_region_subtotal(tmp_path):
    app = _app(tmp_path)
    from app.extensions import db
    from app.models import Product
    from app.services.region_balance_unit_service import _CACHE, region_balance_units

    with app.app_context():
        db.create_all()
        product = Product(product_code="TRAVAZOL", product_name="Travazol", ims_name="TRAVAZOL", is_active=True)
        db.session.add(product)
        db.session.commit()

        archive_dir = Path(app.config["UPLOAD_FOLDER"]) / "ims_archive"
        archive_dir.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "BAKİYE"
        sheet.cell(1, 5).value = "ŞUBAT MF SİZ KUTU BAKİYE"
        sheet.cell(3, 7).value = "TRAVAZOL"
        sheet.cell(13, 1).value = "901 DIYARBAKIR"
        sheet.cell(13, 2).value = "901 DIYARBAKIR"
        sheet.cell(13, 7).value = 213.25983842942765
        workbook.save(archive_dir / "upload-20.xlsx")

        _CACHE.clear()
        values = region_balance_units(20, "901")
        assert abs(values[product.id] - 213.25983842942765) < 1e-9


def test_region_balance_units_keeps_numeric_zero(tmp_path):
    app = _app(tmp_path)
    from app.extensions import db
    from app.models import Product
    from app.services.region_balance_unit_service import _CACHE, region_balance_units

    with app.app_context():
        db.create_all()
        product = Product(product_code="TRAVAZOL", product_name="Travazol", ims_name="TRAVAZOL", is_active=True)
        db.session.add(product)
        db.session.commit()

        archive_dir = Path(app.config["UPLOAD_FOLDER"]) / "ims_archive"
        archive_dir.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "BAKİYE"
        sheet.cell(1, 5).value = "MF SİZ KUTU BAKİYE"
        sheet.cell(3, 7).value = "TRAVAZOL"
        sheet.cell(13, 1).value = "901 DIYARBAKIR"
        sheet.cell(13, 2).value = "901 DIYARBAKIR"
        sheet.cell(13, 7).value = 0
        workbook.save(archive_dir / "upload-21.xlsx")

        _CACHE.clear()
        values = region_balance_units(21, "901")
        assert values[product.id] == 0.0
