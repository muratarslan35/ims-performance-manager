from pathlib import Path

from openpyxl import Workbook

from app import create_app
from app.extensions import db
from app.models import Product, Representative, Target
from app.services.production_result_import_service import ProductionResultImportService


class AdaptiveProductionConfig:
    TESTING = True
    SECRET_KEY = "adaptive-production-parser"
    SQLALCHEMY_DATABASE_URI = "sqlite://"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    UPLOAD_FOLDER = Path("/tmp/adaptive-production-uploads")
    REPORT_FOLDER = Path("/tmp/adaptive-production-reports")
    BACKUP_FOLDER = Path("/tmp/adaptive-production-backups")
    LOG_FOLDER = Path("/tmp/adaptive-production-logs")


def _write_metric_sheet(workbook, title, metric, products, targets, actuals, representative):
    """Build a valid workbook with shuffled blocks/product order and formula-only percents."""
    sheet = workbook.create_sheet(title)
    target_order = list(products)
    actual_order = list(reversed(products))
    percent_order = list(reversed(products))

    header = ["SICIL", "BOLGE", "TEMSILCI"]
    header += actual_order + [f"{metric} CIKIS", "NOT"]
    header += target_order + [f"{metric} HEDEF", "ACIKLAMA"]
    header += percent_order + ["REALIZASYON"]
    sheet.append(header)

    def row(name, region):
        total_target = sum(targets.values())
        total_actual = sum(actuals.values())
        values = ["", region, name]
        values += [actuals[product] for product in actual_order] + [total_actual, ""]
        values += [targets[product] for product in target_order] + [total_target, ""]
        # Formula cells deliberately have no cached result when saved by openpyxl.
        # The importer must derive these deterministic percentages from target/actual.
        values += ["=1" for _ in percent_order] + ["=1"]
        return values

    sheet.append(row("NATIONAL", ""))
    sheet.append(row(representative.rep_name, representative.region))
    return sheet


def test_numeric_parser_keeps_zero_negatives_and_locale_formats():
    parse = ProductionResultImportService._number
    assert parse(0) == 0.0
    assert parse(-1484) == -1484.0
    assert parse("1.234,56") == 1234.56
    assert parse("1,234.56") == 1234.56
    assert parse("(1.234,56)") == -1234.56
    assert parse("₺ 0,00") == 0.0
    assert parse("-") is None


def test_semantic_parser_accepts_negative_actuals_and_discovers_layout(tmp_path):
    path = tmp_path / "subat-ikinci-uretim-farkli-sablon.xlsx"
    app = create_app(AdaptiveProductionConfig)
    with app.app_context():
        db.create_all()
        db.session.query(Target).delete()

        representative = Representative(
            rep_code="ADAPTIVE-REP",
            rep_name="ADAPTIVE TEMSILCI",
            region="601 SAMSUN",
            active=True,
        )
        db.session.add(representative)
        db.session.flush()

        product_codes = ["TRAVAZOL", "BRIMODER"]
        products = []
        for code in product_codes:
            product = Product.query.filter_by(product_code=code).first()
            if product is None:
                product = Product(product_code=code, product_name=code.title(), is_active=True)
                db.session.add(product)
            products.append(product)
        db.session.flush()

        tl_targets = {"TRAVAZOL": 100000.0, "BRIMODER": 20544.384147371456}
        tl_actuals = {"TRAVAZOL": 110000.0, "BRIMODER": -1484.0}
        unit_targets = {"TRAVAZOL": 1000.0, "BRIMODER": 200.0}
        unit_actuals = {"TRAVAZOL": 1100.0, "BRIMODER": -10.0}
        db.session.add_all([
            Target(
                year=2026,
                month=2,
                representative_id=representative.id,
                product_id=product.id,
                tl_target=tl_targets[product.product_code],
                unit_target=unit_targets[product.product_code],
            )
            for product in products
        ])
        db.session.commit()

        workbook = Workbook()
        workbook.remove(workbook.active)
        # Sheet names intentionally do not contain the old fixed TTS REALIZASYONLARI contract.
        _write_metric_sheet(
            workbook,
            "Para Sonuclari Yeni Format",
            "TL",
            product_codes,
            tl_targets,
            tl_actuals,
            representative,
        )
        _write_metric_sheet(
            workbook,
            "Adet Sonuclari Yeni Format",
            "KUTU",
            product_codes,
            unit_targets,
            unit_actuals,
            representative,
        )
        workbook.save(path)

        report = ProductionResultImportService(path, 2026, 2, production_stage=2).parse()
        assert report.matched_rows == 1
        assert report.matched_result_count == 2

        by_product = {row["product_id"]: row for row in report.product_results}
        brimoder = next(product for product in products if product.product_code == "BRIMODER")
        assert by_product[brimoder.id]["actual_tl"] == -1484.0
        assert by_product[brimoder.id]["actual_unit"] == -10.0
        assert by_product[brimoder.id]["realization_percent"] < 0
        assert by_product[brimoder.id]["unit_realization_percent"] < 0

        expected = -1484.0 * 100 / tl_targets["BRIMODER"]
        assert abs(by_product[brimoder.id]["realization_percent"] - expected) < 1e-9


def test_negative_target_still_fails_closed(tmp_path):
    path = tmp_path / "negative-target.xlsx"
    app = create_app(AdaptiveProductionConfig)
    with app.app_context():
        db.create_all()
        db.session.query(Target).delete()
        representative = Representative(
            rep_code="NEG-TARGET",
            rep_name="NEGATIVE TARGET REP",
            region="901 DIYARBAKIR",
            active=True,
        )
        db.session.add(representative)
        product = Product.query.filter_by(product_code="TRAVAZOL").first()
        if product is None:
            product = Product(product_code="TRAVAZOL", product_name="Travazol", is_active=True)
            db.session.add(product)
        db.session.flush()
        db.session.add(Target(
            year=2026,
            month=2,
            representative_id=representative.id,
            product_id=product.id,
            tl_target=100,
            unit_target=10,
        ))
        db.session.commit()

        workbook = Workbook()
        workbook.remove(workbook.active)
        _write_metric_sheet(
            workbook, "TL", "TL", ["TRAVAZOL"], {"TRAVAZOL": -100.0},
            {"TRAVAZOL": 10.0}, representative,
        )
        _write_metric_sheet(
            workbook, "KUTU", "KUTU", ["TRAVAZOL"], {"TRAVAZOL": 10.0},
            {"TRAVAZOL": 1.0}, representative,
        )
        workbook.save(path)

        try:
            ProductionResultImportService(path, 2026, 2).parse()
        except Exception as exc:
            assert "hedef" in str(exc).lower()
        else:
            raise AssertionError("Negative target must remain fail-closed")
