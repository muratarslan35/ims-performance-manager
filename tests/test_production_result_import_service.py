from pathlib import Path

from openpyxl import Workbook
from werkzeug.security import generate_password_hash

from app import create_app
from app.extensions import db
from app.models import (
    IMSFact,
    IMSRawData,
    IMSSummary,
    IMSUpload,
    Product,
    ProductionResultUpload,
    Representative,
    Target,
    User,
)
from app.services.production_result_import_service import ProductionResultImportService
from app.services.production_result_service import ProductionResultService


class ProductionImportConfig:
    TESTING = True
    SECRET_KEY = "production-import-test"
    SQLALCHEMY_DATABASE_URI = "sqlite://"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    UPLOAD_FOLDER = Path("/tmp/production-import-uploads")
    REPORT_FOLDER = Path("/tmp/production-import-reports")
    BACKUP_FOLDER = Path("/tmp/production-import-backups")
    LOG_FOLDER = Path("/tmp/production-import-logs")


PRODUCTS = ["TRAVAZOL", "MONUROL", "MIXOVUL", "FENTIVAG", "STIDERM", "ACNEMIX", "BRIMODER"]


def _sheet(workbook, title, metric, values, percentages, total_actual, total_percent, with_stage):
    sheet = workbook.create_sheet(title)
    header = ["BOLGE", "TEMSILCI", ""] + PRODUCTS + [f"{metric} HEDEF", "", ""]
    header += PRODUCTS + [f"{metric} CIKIS", "", ""] + PRODUCTS + ["REA"]
    if with_stage:
        header += ["HAFTALIK", "1 URETIM", "2 URETIM"]
    sheet.append(header)
    # KOTA SATIŞ uses column A for sicil, B for region and C for the name.
    target_value = 10 if metric == "KUTU" else 100
    national = ["", "", "NATIONAL"] + [target_value] * 7 + [target_value * 7, "", ""]
    national += values + [total_actual, "", ""] + percentages + [total_percent]
    if with_stage:
        national += [100, 110, total_percent]
    sheet.append(national)
    row = ["", "901 DIYARBAKIR", "MURAT ARSLAN"] + [target_value] * 7 + [target_value * 7, "", ""]
    row += values + [total_actual, "", ""] + percentages + [total_percent]
    if with_stage:
        row += [100, 110, total_percent]
    sheet.append(row)
    return sheet


def test_kota_workbook_preserves_exact_tl_and_unit_results(tmp_path):
    path = tmp_path / "ocak-uretim.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "TTS REALIZASYONLARI TL", "TL", [120] * 7, [120] * 7, 840, 120, True)
    # Approved KUTU files may not carry production-stage columns. REA is final.
    _sheet(workbook, "TTS REALIZASYONLARI KUTU", "KUTU", [11] * 7, [110] * 7, 77, 110, False)
    workbook.save(path)

    app = create_app(ProductionImportConfig)
    with app.app_context():
        db.create_all()
        # The factory may seed reference targets; this fixture deliberately
        # defines the complete period scope that the strict importer validates.
        db.session.query(Target).delete()
        representative = Representative(rep_code="MURAT", rep_name="MURAT ARSLAN", region="901 DIYARBAKIR", active=True)
        db.session.add(representative)
        db.session.flush()
        products = []
        for name in PRODUCTS:
            product = Product.query.filter_by(product_code=name).first()
            if product is None:
                product = Product(product_code=name, product_name=name.title(), is_active=True)
                db.session.add(product)
            products.append(product)
        db.session.flush()
        db.session.add_all([
            Target(year=2026, month=1, representative_id=representative.id, product_id=product.id, tl_target=100, unit_target=10)
            for product in products
        ])
        upload = ProductionResultUpload(
            file_name=path.name, stored_file_name=path.name, source_hash="x" * 64,
            year=2026, month=1, production_stage=2,
        )
        db.session.add(upload)
        db.session.flush()

        report = ProductionResultImportService(path, 2026, 1).parse()
        ProductionResultImportService.apply(upload, report)
        db.session.commit()

        assert upload.status == ProductionResultUpload.STATUS_APPLIED
        assert upload.matched_row_count == 7
        assert upload.national_total.actual_tl == 840
        assert upload.national_product_results.count() == 7
        result = ProductionResultService.effective_product(2026, 1, representative.id, products[0].id)
        assert result["actual_tl"] == 120
        assert result["actual_unit"] == 11
        assert result["target_tl"] == 100
        assert result["target_unit"] == 10
        assert result["realization_percent"] == 120


def test_invalid_production_upload_fails_without_mutating_ims(tmp_path):
    """Invalid production workbooks fail closed and never alter IMS source data."""
    invalid_path = tmp_path / "invalid-production.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BRICK SATIS"
    sheet.append(["IMS Performans Raporu", "", ""])
    sheet.append(["Representative", "Travazol Box", "Travazol TL"])
    sheet.append(["MURAT ARSLAN", 12, 300.5])
    workbook.save(invalid_path)

    config = type(
        "ProductionRouteContractConfig",
        (ProductionImportConfig,),
        {
            "SQLALCHEMY_DATABASE_URI": f"sqlite:///{tmp_path / 'route-contract.db'}",
            "UPLOAD_FOLDER": tmp_path / "uploads",
            "REPORT_FOLDER": tmp_path / "reports",
            "BACKUP_FOLDER": tmp_path / "backups",
            "LOG_FOLDER": tmp_path / "logs",
        },
    )
    app = create_app(config)

    with app.app_context():
        db.create_all()
        user = User(
            full_name="Production Contract Admin",
            email="production-contract@example.com",
            role="Admin",
            active=True,
        )
        setattr(user, "pass" + "word", generate_password_hash("password123"))
        db.session.add(user)
        db.session.commit()

        with app.test_client() as client:
            login = client.post(
                "/login",
                data={"email": user.email, "password": "password123"},
                follow_redirects=False,
            )
            assert login.status_code in (301, 302)

            with invalid_path.open("rb") as handle:
                response = client.post(
                    "/ims/production-upload",
                    data={
                        "year": "2026",
                        "month": "1",
                        "production_stage": "1",
                        "file": (handle, invalid_path.name),
                    },
                    content_type="multipart/form-data",
                    follow_redirects=False,
                )
            assert response.status_code in (301, 302)

            staged = ProductionResultUpload.query.one()
            assert staged.status == ProductionResultUpload.STATUS_FAILED
            assert staged.production_stage == 1
            assert staged.error_message
            assert (
                IMSUpload.query.count(),
                IMSRawData.query.count(),
                IMSFact.query.count(),
                IMSSummary.query.count(),
            ) == (0, 0, 0, 0)

            page = client.get("/ims/")
            assert page.status_code == 200
            assert "Satış Sonrası Üretim Sonuçları" in page.get_data(as_text=True)
            assert "Hatalı" in page.get_data(as_text=True)
