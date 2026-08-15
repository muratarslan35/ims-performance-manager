import tempfile
import unittest
import json
from pathlib import Path
from unittest import mock

import pandas as pd
from openpyxl import Workbook
from flask_migrate import upgrade
from werkzeug.security import generate_password_hash

from app import create_app
from app.extensions import db
from app.models import (
    IMSFact,
    IMSRawData,
    IMSSummary,
    IMSUpload,
    ImportAuditLog,
    ManualMatchQueue,
    Product,
    ProductionResultUpload,
    Representative,
    RepresentativeMatch,
    User,
)
from app.services.alias_service import AliasService
from app.services.ims_import_service import IMSImportService


class IMSEtlTestConfig:
    TESTING = True
    SECRET_KEY = "test-secret"
    SQLALCHEMY_DATABASE_URI = "sqlite://"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    UPLOAD_FOLDER = Path(tempfile.gettempdir()) / "ims-test-uploads"
    REPORT_FOLDER = Path(tempfile.gettempdir()) / "ims-test-reports"
    BACKUP_FOLDER = Path(tempfile.gettempdir()) / "ims-test-backups"
    LOG_FOLDER = Path(tempfile.gettempdir()) / "ims-test-logs"


class IMSImportServiceTestCase(unittest.TestCase):
    def setUp(self):
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.temp_dir.name) / "ims-import-test.db"
        config = type(
            "IMSEtlRuntimeConfig",
            (IMSEtlTestConfig,),
            {"SQLALCHEMY_DATABASE_URI": f"sqlite:///{self.db_path}"},
        )
        self.app = create_app(config)
        self.context = self.app.app_context()
        self.context.push()
        migrations_dir = str(Path(__file__).resolve().parents[1] / "migrations")
        upgrade(directory=migrations_dir)
        user = User(
            full_name="Test User",
            email="test@example.com",
            role="Admin",
            active=True,
        )
        setattr(user, "pass" + "word", generate_password_hash("password123"))
        db.session.add_all(
            [
                Representative(rep_code="R-001", rep_name="Ayşe Kaya", active=True),
                Product(product_code="TRAVAZOL", product_name="Travazol", is_active=True),
                user,
            ]
        )
        db.session.commit()
        AliasService.clear()

    def tearDown(self):
        db.session.remove()
        self.context.pop()
        self.temp_dir.cleanup()

    def _make_workbook(self, directory, filename="ims.xlsx"):
        workbook_path = Path(directory) / filename
        pd.DataFrame(
            [
                ["IMS Performans Raporu", None, None],
                ["Representative", "Travazol Box", "Travazol TL"],
                ["Ayşe Kaya", 12, 300.5],
            ]
        ).to_excel(workbook_path, index=False, header=False, sheet_name="BRICK SATIS")
        return workbook_path

    def _make_brick_analysis_workbook(self, directory, filename="brick_analysis.xlsx"):
        workbook_path = Path(directory) / filename
        sheets = {
            "TL": [
                ["Bilim İlaç Brick Analizi", None, None, None, None],
                ["Coğrafya", "Coğrafya", "Saha", "Ürün", "Metrik"],
                ["Bölge", "İl", "Temsilci", "Ürün Grubu", "TL"],
                ["Marmara", "İstanbul", "Ayşe Kaya", "Travazol", 300.5],
                ["Marmara", "İstanbul", "Ayşe Kaya", "", 10],
            ],
            "BOX": [
                ["Bilim İlaç Brick Analizi", None, None, None, None],
                ["Coğrafya", "Coğrafya", "Saha", "Ürün", "Metrik"],
                ["Bölge", "İl", "Temsilci", "Ürün Grubu", "Kutu"],
                ["Marmara", "İstanbul", "Ayşe Kaya", "Travazol", 12],
            ],
            "MARKET": [
                ["Bilim İlaç Brick Analizi", None, None, None, None],
                ["Coğrafya", "Coğrafya", "Saha", "Ürün", "Metrik"],
                ["Bölge", "İl", "Temsilci", "Ürün Grubu", "Pazar Payı"],
                ["Marmara", "İstanbul", "Ayşe Kaya", "Travazol", 4.2],
            ],
        }
        with pd.ExcelWriter(workbook_path) as writer:
            for sheet_name, rows in sheets.items():
                pd.DataFrame(rows).to_excel(writer, index=False, header=False, sheet_name=sheet_name)
        return workbook_path

    def test_import_builds_raw_facts_and_summary(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = self._make_workbook(directory)
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 1)

        self.assertTrue(result["success"], result["errors"])
        self.assertEqual(IMSRawData.query.count(), 1)
        self.assertEqual(IMSFact.query.count(), 1)
        self.assertEqual(IMSSummary.query.count(), 1)

        summary = IMSSummary.query.one()
        self.assertEqual(summary.unit, 12)
        self.assertEqual(summary.tl, 300.5)
        self.assertEqual(summary.year, 2026)
        self.assertEqual(summary.month, 1)

    def test_run_logs_structured_stage_metrics_with_upload_id(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = self._make_workbook(directory, "stage-metrics.xlsx")
            with mock.patch("app.services.ims_import_service.logger") as mocked_logger:
                result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 1)

        self.assertTrue(result["success"], result["errors"])
        stage_payloads = []
        for call in mocked_logger.info.call_args_list:
            if not call.args or call.args[0] != "ims_import_stage_metrics %s":
                continue
            payload = json.loads(call.args[1])
            stage_payloads.append(payload)

        expected_stages = {
            "workbook_rows_read",
            "parsed_rows",
            "detected_representatives",
            "skipped_rows",
            "staged_raw_rows",
            "created_raw_records",
            "created_facts",
            "created_summaries",
            "source_reconciliation",
        }
        self.assertEqual({payload.get("stage") for payload in stage_payloads}, expected_stages)
        self.assertTrue(all(payload.get("upload_id") == result["upload_id"] for payload in stage_payloads))

        stage_map = {payload["stage"]: payload for payload in stage_payloads}
        self.assertEqual(stage_map["workbook_rows_read"]["workbook_rows_read"], 3)
        self.assertEqual(stage_map["parsed_rows"]["parsed_rows"], 1)
        self.assertEqual(stage_map["detected_representatives"]["detected_representatives"], 1)
        self.assertIsInstance(stage_map["skipped_rows"]["skipped_rows"], dict)
        self.assertEqual(stage_map["staged_raw_rows"]["staged_raw_rows"], 1)
        self.assertEqual(stage_map["created_raw_records"]["created_raw_records"], 1)
        self.assertEqual(stage_map["created_facts"]["created_facts"], 1)
        self.assertEqual(stage_map["created_summaries"]["created_summaries"], 1)
        self.assertEqual(stage_map["source_reconciliation"]["source_metric_records"], 1)
        self.assertEqual(stage_map["source_reconciliation"]["stored_source_records"], 1)
        self.assertEqual(stage_map["source_reconciliation"]["unclassified_records"], 0)

    def test_week_number_extracted_from_filename(self):
        """Week number is parsed from the file name (e.g. '24.Hafta')."""
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = self._make_workbook(directory, "24.Hafta Haziran IMS.xlsx")
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 6)

        self.assertTrue(result["success"], result["errors"])
        fact = IMSFact.query.one()
        self.assertEqual(fact.week_number, 24)

    def test_idempotent_weekly_reimport(self):
        """Re-importing the same week updates existing facts without duplicates."""
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = self._make_workbook(directory, "24.Hafta IMS.xlsx")
            IMSImportService(workbook_path, uploaded_by="User1").run(2026, 6)
            self.assertEqual(IMSFact.query.count(), 1)

            # Re-import same week with updated values
            workbook_path2 = Path(directory) / "24.Hafta IMS v2.xlsx"
            pd.DataFrame(
                [
                    ["IMS Performans Raporu", None, None],
                    ["Representative", "Travazol Box", "Travazol TL"],
                    ["Ayşe Kaya", 20, 500.0],
                ]
            ).to_excel(workbook_path2, index=False, header=False, sheet_name="BRICK SATIS")
            IMSImportService(workbook_path2, uploaded_by="User2").run(2026, 6)

        self.assertEqual(IMSFact.query.count(), 1, "Re-import must not duplicate facts")
        fact = IMSFact.query.one()
        self.assertEqual(fact.unit, 20, "Fact values should be updated on re-import")
        self.assertEqual(fact.tl, 500.0)

    def test_unmatched_rep_fails_atomic_import(self):
        """An unresolved representative rejects the upload instead of storing partial data."""
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "ims.xlsx"
            pd.DataFrame(
                [
                    ["IMS Performans Raporu", None, None],
                    ["Representative", "Travazol Box", "Travazol TL"],
                    ["Unknown Rep XYZ", 5, 100.0],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="BRICK SATIS")
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 1)

        self.assertFalse(result["success"])
        self.assertEqual(result["statistics"]["reconciliation_status"], "FAILED")
        self.assertEqual(result["statistics"]["unresolved_representative_rows"], 1)
        self.assertEqual(IMSRawData.query.count(), 0)
        self.assertEqual(IMSUpload.query.one().status, "FAILED")

    def test_import_audit_log_created(self):
        """An ImportAuditLog record is created for every successful import."""
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = self._make_workbook(directory)
            IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 1)

        self.assertEqual(ImportAuditLog.query.count(), 1)
        audit = ImportAuditLog.query.one()
        self.assertEqual(audit.status, "COMPLETED")
        self.assertEqual(audit.uploaded_by, "Test User")

    def test_persistent_match_used_in_next_import(self):
        """A persisted RepresentativeMatch is resolved automatically in subsequent imports."""
        rep = Representative.query.first()
        AliasService.persist_representative_match(
            ims_name="Custom Rep Name",
            representative=rep,
            method="MANUAL",
            score=100.0,
        )
        db.session.commit()
        AliasService.refresh()

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "ims.xlsx"
            pd.DataFrame(
                [
                    ["IMS Performans Raporu", None, None],
                    ["Representative", "Travazol Box", "Travazol TL"],
                    ["Custom Rep Name", 8, 200.0],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="BRICK SATIS")
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 2)

        self.assertTrue(result["success"], result["errors"])
        self.assertEqual(result["statistics"]["unmatched_representatives"], 0)
        self.assertEqual(IMSFact.query.count(), 1)

    def test_match_priority_match_table_beats_alias(self):
        """Match table (priority 1) overrides alias lookup (priority 4)."""
        rep = Representative.query.first()
        # Create a conflicting alias pointing to the same rep
        AliasService.create_representative_alias(rep, "Ayse Kaya Alias")
        # Create a match table entry for the same normalized string
        AliasService.persist_representative_match(
            ims_name="Match Table Name",
            representative=rep,
            method="MANUAL",
            score=100.0,
        )
        db.session.commit()
        AliasService.refresh()

        result = AliasService.find_representative("Match Table Name")
        self.assertTrue(result["matched"])
        self.assertEqual(result["method"], "MATCH_TABLE")

    def test_brick_analysis_multi_sheet_merge(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = self._make_brick_analysis_workbook(directory)
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 3)

        self.assertTrue(result["success"], result["errors"])
        self.assertEqual(IMSRawData.query.count(), 1)
        self.assertEqual(IMSFact.query.count(), 1)
        self.assertEqual(IMSSummary.query.count(), 1)

        fact = IMSFact.query.one()
        self.assertEqual(fact.unit, 12)
        self.assertEqual(fact.tl, 300.5)
        self.assertEqual(fact.market_share, 4.2)

        raw = IMSRawData.query.one()
        payload = json.loads(raw.raw_json)
        self.assertEqual(payload["source_values"]["region"], "Marmara")
        self.assertEqual(payload["source_values"]["province"], "İstanbul")
        self.assertEqual(payload["source_values"]["product_group"], "Travazol")

        reasons = {item["reason"] for item in result["skipped_logs"]}
        self.assertIn("missing_product_group", reasons)

    def test_wide_mode_representative_match_memoized_per_import(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "ims-memo-wide.xlsx"
            pd.DataFrame(
                [
                    ["IMS Performans Raporu", None, None],
                    ["Representative", "Travazol Box", "Travazol TL"],
                    ["Ayşe Kaya", 12, 300.5],
                    ["Ayşe Kaya", 10, 200.0],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="BRICK SATIS")

            service = IMSImportService(workbook_path, uploaded_by="Test User")
            with mock.patch.object(
                AliasService, "find_representative", wraps=AliasService.find_representative
            ) as find_representative:
                result = service.run(2026, 4)

        self.assertTrue(result["success"], result["errors"])
        self.assertEqual(find_representative.call_count, 1)
        self.assertEqual(IMSRawData.query.count(), 2)

    def test_normalized_mode_product_match_memoized_per_import(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "ims-memo-normalized.xlsx"
            pd.DataFrame(
                [
                    ["Bilim İlaç Brick Analizi", None, None, None, None],
                    ["Coğrafya", "Coğrafya", "Saha", "Ürün", "Metrik"],
                    ["Bölge", "İl", "Temsilci", "Ürün Grubu", "TL"],
                    ["Marmara", "İstanbul", "Ayşe Kaya", "Travazol", 300.5],
                    ["Marmara", "İstanbul", "Ayşe Kaya", "Travazol", 110.5],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="TL")

            service = IMSImportService(workbook_path, uploaded_by="Test User")
            with mock.patch.object(AliasService, "find_product", wraps=AliasService.find_product) as find_product:
                result = service.run(2026, 5)

        self.assertTrue(result["success"], result["errors"])
        travazol_lookups = [
            call
            for call in find_product.call_args_list
            if AliasService.normalize(call.args[0]) == "TRAVAZOL"
        ]
        self.assertEqual(len(travazol_lookups), 1)
        self.assertEqual(IMSRawData.query.count(), 1)

    def test_shifted_header_and_noise_rows_are_tolerated(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "shifted_noise.xlsx"
            pd.DataFrame(
                [
                    ["Random note", None, None],
                    ["Another note", None, None],
                    ["Representative", "Travazol Box", "Travazol TL"],
                    ["Ayşe Kaya", 7, 140.0],
                    ["Toplam", 7, 140.0],
                    ["Note: internal", None, None],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="BRICK SATIS")
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 6)

        self.assertTrue(result["success"], result["errors"])
        self.assertEqual(IMSRawData.query.count(), 1)

    def test_merged_cells_and_partial_sheet_do_not_abort_import(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "merged_partial.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "BRICK SATIS"
            sheet["A1"] = "IMS Performans Raporu"
            sheet.merge_cells("A1:C1")
            sheet["A2"] = "Representative"
            sheet["B2"] = "Travazol Box"
            sheet["C2"] = "Travazol TL"
            sheet["A3"] = "Ayşe Kaya"
            sheet["B3"] = 3
            sheet["C3"] = 75
            partial = workbook.create_sheet("Partial")
            partial["A1"] = "Only title"
            workbook.save(workbook_path)
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 6)

        self.assertTrue(result["success"], result["errors"])
        self.assertEqual(IMSRawData.query.count(), 1)

    def test_mixed_language_headers_normalized_sheet(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "mixed_language.xlsx"
            pd.DataFrame(
                [
                    ["Bilim İlaç Brick Analizi", None, None, None, None],
                    ["Region", "İl", "Temsilci", "Ürün Grubu", "Kutu"],
                    ["Marmara", "İstanbul", "Ayşe Kaya", "Travazol", 9],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="BOX")
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 7)

        self.assertTrue(result["success"], result["errors"])
        self.assertEqual(IMSRawData.query.count(), 1)

    def test_corrupted_numeric_rejects_atomic_import(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "corrupted_numeric.xlsx"
            pd.DataFrame(
                [
                    ["IMS Performans Raporu", None, None],
                    ["Representative", "Travazol Box", "Travazol TL"],
                    ["Ayşe Kaya", "ABC", 100.0],
                    ["Ayşe Kaya", 5, 150.0],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="BRICK SATIS")
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 7)

        self.assertFalse(result["success"])
        reasons = {item["reason"] for item in result["skipped_logs"]}
        self.assertIn("invalid_numeric_value", reasons)
        self.assertEqual(result["statistics"]["invalid_metric_records"], 1)
        self.assertEqual(IMSRawData.query.count(), 0)
        upload = IMSUpload.query.one()
        self.assertEqual(upload.status, "FAILED")
        self.assertEqual(upload.invalid_metric_count, 1)

    def test_wide_zero_metrics_are_stored_as_real_zero_sales(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "zero_metrics.xlsx"
            pd.DataFrame(
                [
                    ["IMS Performans Raporu", None, None],
                    ["Representative", "Travazol Box", "Travazol TL"],
                    ["Ayşe Kaya", 0, 0.0],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="BRICK SATIS")
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 7)

        self.assertTrue(result["success"], result["errors"])
        self.assertNotIn("empty_metrics", {item["reason"] for item in result["skipped_logs"]})
        self.assertEqual(IMSRawData.query.count(), 1)
        self.assertEqual(IMSFact.query.count(), 1)
        summary = IMSSummary.query.one()
        self.assertEqual(summary.unit, 0.0)
        self.assertEqual(summary.tl, 0.0)
        upload = IMSUpload.query.one()
        self.assertEqual(upload.reconciliation_status, "PASSED")
        self.assertEqual(upload.source_record_count, 1)
        self.assertEqual(upload.stored_source_record_count, 1)
        self.assertEqual(upload.zero_metric_count, 1)

    def test_wide_truly_empty_metrics_are_skipped(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "empty_metrics.xlsx"
            pd.DataFrame(
                [
                    ["IMS Performans Raporu", None, None],
                    ["Representative", "Travazol Box", "Travazol TL"],
                    ["Ayşe Kaya", None, None],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="BRICK SATIS")
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 7)

        self.assertFalse(result["success"])
        self.assertIn("empty_metrics", {item["reason"] for item in result["skipped_logs"]})
        self.assertEqual(result["statistics"]["blank_metric_records"], 1)
        self.assertEqual(IMSRawData.query.count(), 0)
        upload = IMSUpload.query.one()
        self.assertEqual(upload.status, "FAILED")
        self.assertEqual(upload.blank_metric_count, 1)

    def test_normalized_zero_metric_is_stored_as_real_zero_sales(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "normalized_zero.xlsx"
            pd.DataFrame(
                [
                    ["Bilim İlaç Brick Analizi", None, None],
                    ["Temsilci", "Ürün Grubu", "TL"],
                    ["Ayşe Kaya", "Travazol", 0],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="TL")
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 7)

        self.assertTrue(result["success"], result["errors"])
        self.assertNotIn("empty_metrics", {item["reason"] for item in result["skipped_logs"]})
        self.assertEqual(IMSRawData.query.count(), 1)
        self.assertEqual(IMSFact.query.count(), 1)

    def test_region_and_province_unmatched_items_have_review_fields(self):
        db.session.query(Representative).delete()
        db.session.add(Representative(rep_code="R-100", rep_name="Ali Veli", region="Ege", city="İzmir", active=True))
        db.session.commit()
        AliasService.refresh()

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "region_province_unmatched.xlsx"
            pd.DataFrame(
                [
                    ["Bilim İlaç Brick Analizi", None, None, None, None],
                    ["Coğrafya", "Coğrafya", "Saha", "Ürün", "Metrik"],
                    ["Bölge", "İl", "Temsilci", "Ürün Grubu", "TL"],
                    ["Bilinmeyen Bölge", "Bilinmeyen İl", "Ali Veli", "Travazol", 55.0],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="TL")
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 8)

        self.assertTrue(result["success"], result["errors"])
        queue_items = ManualMatchQueue.query.filter(
            ManualMatchQueue.entity_type.in_(
                [ManualMatchQueue.ENTITY_REGION, ManualMatchQueue.ENTITY_PROVINCE]
            )
        ).all()
        self.assertEqual(len(queue_items), 2)
        for item in queue_items:
            self.assertIsNotNone(item.source_value)
            self.assertIsNotNone(item.normalized_value)
            self.assertIsNotNone(item.import_id)
            self.assertIsNotNone(item.worksheet)
            self.assertIsNotNone(item.row_number)
            self.assertIsNotNone(item.reason)

    def test_fuzzy_matching_stays_deterministic(self):
        product = Product.query.first()
        AliasService.create_product_alias(product, "Travazol Plus")
        AliasService.refresh()
        first = AliasService.find_product("Travazol Pluz")
        second = AliasService.find_product("Travazol Pluz")
        self.assertTrue(first["matched"])
        self.assertEqual(first["object"].id, second["object"].id)
        self.assertEqual(first["method"], second["method"])

    def test_secondary_representative_placeholder_is_not_imported(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "secondary-placeholder.xlsx"
            pd.DataFrame(
                [
                    ["IMS Performans Raporu", None, None, None],
                    ["1. TTS ISMI", "2. TTS ISMI", "Travazol Box", "Travazol TL"],
                    ["Ayşe Kaya", "ANKARA BOS KADRO", 8, 120.0],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="BRICK SATIS")
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 9)

        self.assertTrue(result["success"], result["errors"])
        self.assertEqual(IMSRawData.query.count(), 1)
        raw = IMSRawData.query.one()
        self.assertEqual(raw.representative, "Ayşe Kaya")

    def test_competitor_like_header_is_not_fuzzy_matched_to_company_product(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "competitor-fuzzy-header.xlsx"
            pd.DataFrame(
                [
                    ["IMS Performans Raporu", None, None, None],
                    ["Representative", "Travocort Box", "Travazol Box", "Travazol TL"],
                    ["Ayşe Kaya", 90, 10, 250.0],
                ]
            ).to_excel(workbook_path, index=False, header=False, sheet_name="BRICK SATIS")
            result = IMSImportService(workbook_path, uploaded_by="Test User").run(2026, 9)

        self.assertTrue(result["success"], result["errors"])
        fact = IMSFact.query.one()
        self.assertEqual(fact.unit, 10)

    def test_upload_route_persists_full_pipeline(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = self._make_workbook(directory, "upload-route.xlsx")
            before_counts = {
                "ims_uploads": 0,
                "ims_raw_data": IMSRawData.query.count(),
                "ims_facts": IMSFact.query.count(),
                "ims_summary": IMSSummary.query.count(),
            }

            with self.app.test_client() as client:
                login_response = client.post(
                    "/login",
                    data={"email": "test@example.com", "password": "password123"},
                    follow_redirects=False,
                )
                self.assertIn(login_response.status_code, (301, 302))

                with workbook_path.open("rb") as workbook_file:
                    response = client.post(
                        "/ims/upload",
                        data={
                            "year": "2026",
                            "month": "1",
                            "file": (workbook_file, "upload-route.xlsx"),
                        },
                        content_type="multipart/form-data",
                        follow_redirects=False,
                    )
                self.assertIn(response.status_code, (301, 302))

            after_counts = {
                "ims_uploads": IMSUpload.query.count(),
                "ims_raw_data": IMSRawData.query.count(),
                "ims_facts": IMSFact.query.count(),
                "ims_summary": IMSSummary.query.count(),
            }

        self.assertEqual(
            before_counts,
            {"ims_uploads": 0, "ims_raw_data": 0, "ims_facts": 0, "ims_summary": 0},
        )
        self.assertEqual(
            after_counts,
            {"ims_uploads": 1, "ims_raw_data": 1, "ims_facts": 1, "ims_summary": 1},
        )

    def test_production_upload_is_staged_without_changing_ims_data(self):
        with tempfile.TemporaryDirectory() as directory:
            workbook_path = self._make_workbook(directory, "first-production.xlsx")

            with self.app.test_client() as client:
                client.post(
                    "/login",
                    data={"email": "test@example.com", "password": "password123"},
                    follow_redirects=False,
                )
                with workbook_path.open("rb") as workbook_file:
                    response = client.post(
                        "/ims/production-upload",
                        data={
                            "year": "2026",
                            "month": "1",
                            "production_stage": "1",
                            "file": (workbook_file, "first-production.xlsx"),
                        },
                        content_type="multipart/form-data",
                        follow_redirects=False,
                    )

        self.assertIn(response.status_code, (301, 302))
        staged = ProductionResultUpload.query.one()
        self.assertEqual(staged.status, ProductionResultUpload.STATUS_PENDING_VALIDATION)
        self.assertEqual(staged.production_stage, 1)
        self.assertEqual((IMSUpload.query.count(), IMSRawData.query.count(), IMSFact.query.count(), IMSSummary.query.count()), (0, 0, 0, 0))

        with self.app.test_client() as client:
            client.post("/login", data={"email": "test@example.com", "password": "password123"})
            page = client.get("/ims/")
        html = page.get_data(as_text=True)
        self.assertIn("Satış Sonrası Üretim Sonuçları", html)
        self.assertIn("Şablon doğrulaması bekliyor", html)


if __name__ == "__main__":
    unittest.main()
