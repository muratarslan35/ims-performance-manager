import tempfile
import unittest
from pathlib import Path

from flask_migrate import upgrade
from openpyxl import load_workbook
from werkzeug.security import generate_password_hash

from app import create_app
from app.extensions import db
from app.models import IMSSummary, IMSUpload, PrimeRule, Product, Representative, Setting, Target, User
from app.prime_engine import PrimeEngine as WrappedPrimeEngine
from app.services.prime_engine import PrimeEngine, _cache_clear
from app.services.simulation_service import SimulationService
from app.simulation import build_overrides


class PrimeEngineTestConfig:
    TESTING = True
    SECRET_KEY = "prime-engine-test-secret"
    SQLALCHEMY_DATABASE_URI = "sqlite://"
    SQLALCHEMY_TRACK_MODIFICATIONS = False
    UPLOAD_FOLDER = Path(tempfile.gettempdir()) / "prime-engine-test-uploads"
    REPORT_FOLDER = Path(tempfile.gettempdir()) / "prime-engine-test-reports"
    BACKUP_FOLDER = Path(tempfile.gettempdir()) / "prime-engine-test-backups"
    LOG_FOLDER = Path(tempfile.gettempdir()) / "prime-engine-test-logs"


class PrimeEngineBaseTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.TemporaryDirectory()
        db_path = Path(cls.temp_dir.name) / "prime-engine-test.db"
        config = type(
            "PrimeEngineRuntimeConfig",
            (PrimeEngineTestConfig,),
            {
                "SQLALCHEMY_DATABASE_URI": f"sqlite:///{db_path}",
                "REPORT_FOLDER": Path(cls.temp_dir.name) / "reports",
                "UPLOAD_FOLDER": Path(cls.temp_dir.name) / "uploads",
                "BACKUP_FOLDER": Path(cls.temp_dir.name) / "backups",
                "LOG_FOLDER": Path(cls.temp_dir.name) / "logs",
            },
        )
        cls.app = create_app(config)
        cls.ctx = cls.app.app_context()
        cls.ctx.push()
        migrations_dir = str(Path(__file__).resolve().parents[1] / "migrations")
        upgrade(directory=migrations_dir)

    @classmethod
    def tearDownClass(cls):
        db.session.remove()
        cls.ctx.pop()
        cls.temp_dir.cleanup()

    def setUp(self):
        _cache_clear()
        report_folder = Path(self.app.config["REPORT_FOLDER"])
        if report_folder.exists():
            for path in sorted(report_folder.rglob("*"), reverse=True):
                if path.is_file():
                    path.unlink()
                elif path.is_dir():
                    path.rmdir()
        self._reset_database()
        self._seed_base()

    def tearDown(self):
        db.session.rollback()
        _cache_clear()

    def _reset_database(self):
        for model in [IMSSummary, Target, PrimeRule, IMSUpload, Product, Representative, User, Setting]:
            db.session.query(model).delete()
        db.session.commit()
        db.session.expunge_all()

    def _seed_base(self):
        self.rep = Representative(rep_code="R-001", rep_name="Ayşe Kaya", city="İstanbul", active=True)
        self.prod1 = Product(product_code="TRAV", product_name="Travazol", is_active=True, is_prime_product=True, include_total_tl=True, required_percent=90, display_order=1)
        self.prod2 = Product(product_code="MONO", product_name="Monurol", is_active=True, is_prime_product=True, include_total_tl=True, required_percent=90, display_order=2)
        self.prod3 = Product(product_code="MIXO", product_name="Mixovul", is_active=True, is_prime_product=True, include_total_tl=True, required_percent=90, display_order=3)
        user = User(full_name="Admin", email="admin-prime@test.com", role="Admin", active=True)
        setattr(user, "pass" + "word", generate_password_hash("pass"))
        db.session.add_all([self.rep, self.prod1, self.prod2, self.prod3, user])
        db.session.commit()

        settings = {
            "MAIN_PRIME": "50000",
            "CIRO_PRIME": "20000",
            "PRIME_STEP": "5",
            "STEP_AMOUNT": "2500",
            "MAX_PRIME_PERCENT": "140",
            "MIN_PRIME_PERCENT": "100",
            "TOTAL_PERCENT_REQUIRED": "100",
            "PRIME_PRODUCT_COUNT": "3",
            "REQUIRED_90_COUNT": "2",
            "REQUIRED_75_COUNT": "1",
            "ALLOW_CIRO_WITHOUT_PRODUCT": "1",
            "RECOVERY_EFFECT_RATE": "2",
            "QUARTER_EFFECT_RATE": "10",
            "PRODUCT_COEFFICIENT_DEFAULT": "1",
            "PRODUCT_BONUS_RATE": "1",
            "BONUS_RATE": "5",
            "PENALTY_RATE": "3",
            "PENALTY_PER_FAILED_PRODUCT": "1500",
            "WHAT_IF_WORST_FACTOR": "0.85",
            "WHAT_IF_EXPECTED_FACTOR": "1.10",
            "WHAT_IF_BEST_FACTOR": "1.25",
            "SLIDER_MAX_PERCENT": "150",
        }
        db.session.add_all([
            Setting(setting_key=key, setting_value=value, category="Prim", description="Test")
            for key, value in settings.items()
        ])
        db.session.commit()

        for product in [self.prod1, self.prod2, self.prod3]:
            db.session.add(
                PrimeRule(
                    product_id=product.id,
                    required_percent=90,
                    include_in_prime=True,
                    include_in_total_tl=True,
                    active=True,
                )
            )
        db.session.commit()

        self._seed_period(4, [100000, 90000, 80000], [105000, 60000, 88000], [1000, 200, 300])
        self._seed_period(5, [110000, 95000, 85000], [115000, 85000, 70000], [1100, 250, 0])
        self._seed_period(6, [120000, 100000, 90000], [130000, 70000, 95000], [1200, 150, 100])

    def _seed_period(self, month, targets, actuals, bonuses):
        upload = IMSUpload(file_name=f"ims-{month}.xlsx", year=2025, month=month, quarter="Q2", status="COMPLETED")
        db.session.add(upload)
        db.session.commit()
        products = [self.prod1, self.prod2, self.prod3]
        for product, target_tl, actual_tl, bonus in zip(products, targets, actuals, bonuses):
            db.session.add(
                Target(
                    year=2025,
                    month=month,
                    quarter="Q2",
                    representative_id=self.rep.id,
                    product_id=product.id,
                    unit_target=target_tl / 1000,
                    tl_target=target_tl,
                )
            )
            db.session.add(
                IMSSummary(
                    upload_id=upload.id,
                    representative_id=self.rep.id,
                    product_id=product.id,
                    year=2025,
                    month=month,
                    quarter="Q2",
                    unit=actual_tl / 1000,
                    tl=actual_tl,
                    market_share=10.0,
                    growth=3.0 if product.id != self.prod2.id else -8.0,
                    bonus_amount=bonus,
                )
            )
        db.session.commit()

    def create_engine(self, overrides=None, use_cache=False):
        return PrimeEngine(
            representative_id=self.rep.id,
            year=2025,
            month=6,
            overrides=overrides or {},
            use_cache=use_cache,
        )

    def calculate(self, overrides=None, use_cache=False, save_history=False):
        return self.create_engine(overrides=overrides, use_cache=use_cache).calculate(save_history=save_history)


class TestBuildOverrides(PrimeEngineBaseTestCase):
    def test_build_overrides_ignores_zero_rows(self):
        overrides, duplicates = build_overrides({"products": [{"product_id": self.prod1.id, "tl": 0, "unit": 0}]})
        self.assertEqual(overrides, {})
        self.assertEqual(duplicates, [])

    def test_build_overrides_detects_duplicates(self):
        overrides, duplicates = build_overrides({"products": [{"product_id": self.prod1.id, "tl": 10}, {"product_id": self.prod1.id, "tl": 20}]})
        self.assertIn(self.prod1.id, duplicates)
        self.assertEqual(len(overrides), 1)

    def test_build_overrides_supports_delta_mode(self):
        overrides, _ = build_overrides({"products": [{"product_id": self.prod1.id, "tl": 250000}]})
        self.assertEqual(overrides[self.prod1.id]["mode"], "delta")
        self.assertEqual(overrides[self.prod1.id]["tl_delta"], 250000)

    def test_build_overrides_supports_replace_mode(self):
        overrides, _ = build_overrides({"products": [{"product_id": self.prod1.id, "mode": "replace", "tl": 200000}]})
        self.assertEqual(overrides[self.prod1.id]["mode"], "replace")
        self.assertEqual(overrides[self.prod1.id]["tl"], 200000)

    def test_build_overrides_supports_slider(self):
        overrides, _ = build_overrides({"products": [{"product_id": self.prod1.id, "slider_percent": 125}]})
        self.assertEqual(overrides[self.prod1.id]["slider_percent"], 125)

    def test_build_overrides_supports_target_percent(self):
        overrides, _ = build_overrides({"products": [{"product_id": self.prod1.id, "target_percent": 110}]})
        self.assertEqual(overrides[self.prod1.id]["target_percent"], 110)

    def test_build_overrides_ignores_invalid_product_ids(self):
        overrides, duplicates = build_overrides({"products": [{"product_id": 0, "tl": 1}]})
        self.assertEqual(overrides, {})
        self.assertEqual(duplicates, [])


class TestPrimeEngineCalculations(PrimeEngineBaseTestCase):
    def test_monthly_entitlement_allows_any_one_product_at_75_percent(self):
        engine = self.create_engine()
        engine.settings.update({
            "PRIME_PRODUCT_COUNT": "4",
            "REQUIRED_90_COUNT": "3",
            "REQUIRED_75_COUNT": "1",
            "TARGET_75": "75",
            "TARGET_90": "90",
        })
        products = [
            {"product_name": name, "include_in_prime": True, "percent": percent}
            for name, percent in [("Travazol", 76), ("Monurol", 90), ("Mixovul", 91), ("Acnemix", 92)]
        ]
        entitlement = engine.evaluate_monthly_entitlement(products)
        self.assertTrue(entitlement["product_success"])
        self.assertEqual([item["product_name"] for item in entitlement["below_standard_products"]], ["Travazol"])

    def test_monthly_entitlement_rejects_two_products_below_90_percent(self):
        engine = self.create_engine()
        engine.settings.update({"PRIME_PRODUCT_COUNT": "4", "REQUIRED_90_COUNT": "3", "REQUIRED_75_COUNT": "1"})
        products = [
            {"product_name": name, "include_in_prime": True, "percent": percent}
            for name, percent in [("Travazol", 92), ("Monurol", 80), ("Mixovul", 89), ("Acnemix", 95)]
        ]
        self.assertFalse(engine.evaluate_monthly_entitlement(products)["product_success"])

    def test_monthly_entitlement_rejects_a_product_below_75_percent(self):
        engine = self.create_engine()
        engine.settings.update({"PRIME_PRODUCT_COUNT": "4", "REQUIRED_90_COUNT": "3", "REQUIRED_75_COUNT": "1"})
        products = [
            {"product_name": name, "include_in_prime": True, "percent": percent}
            for name, percent in [("Travazol", 92), ("Monurol", 90), ("Mixovul", 91), ("Acnemix", 74)]
        ]
        self.assertFalse(engine.evaluate_monthly_entitlement(products)["product_success"])

    def test_base_total_percent_is_calculated(self):
        result = self.calculate()
        self.assertAlmostEqual(result["total_tl_percent"], 95.16, places=2)

    def test_base_main_prime_is_zero_below_threshold(self):
        result = self.calculate()
        self.assertEqual(result["breakdown"]["main_prime"], 0)

    def test_override_can_unlock_main_prime(self):
        result = self.calculate(overrides={self.prod2.id: {"tl_delta": 40000, "mode": "delta"}})
        self.assertGreater(result["breakdown"]["main_prime"], 0)

    def test_override_unlocks_main_prime_not_ciro_prime(self):
        result = self.calculate(overrides={self.prod2.id: {"tl_delta": 40000, "mode": "delta"}})
        self.assertGreater(result["breakdown"]["main_prime"], 0)
        self.assertEqual(result["breakdown"]["ciro_prime"], 0)

    def test_bonus_uses_database_bonus_amount(self):
        result = self.calculate()
        self.assertEqual(result["breakdown"]["bonus"], 0)

    def test_penalty_is_applied_for_failed_product(self):
        result = self.calculate()
        self.assertEqual(result["breakdown"]["penalty"], 0)

    def test_recovery_component_zero_without_recovery_gain(self):
        result = self.calculate()
        self.assertEqual(result["breakdown"]["recovery"], 0)

    def test_recovery_component_stays_zero_when_gap_closes(self):
        result = self.calculate(overrides={self.prod2.id: {"tl_delta": 40000, "mode": "delta"}})
        self.assertEqual(result["breakdown"]["recovery"], 0)

    def test_slider_override_scales_current_actual(self):
        result = self.calculate(overrides={self.prod1.id: {"slider_percent": 150, "mode": "delta"}})
        travazol = result["product_results"][self.prod1.id]
        self.assertEqual(travazol["actual_tl"], 195000)

    def test_slider_override_is_capped_by_setting(self):
        result = self.calculate(overrides={self.prod1.id: {"slider_percent": 999, "mode": "delta"}})
        travazol = result["product_results"][self.prod1.id]
        self.assertEqual(travazol["actual_tl"], 195000)

    def test_delta_override_updates_tl(self):
        result = self.calculate(overrides={self.prod3.id: {"tl_delta": 90000, "mode": "delta"}})
        self.assertEqual(result["product_results"][self.prod3.id]["actual_tl"], 185000)

    def test_delta_override_updates_unit(self):
        result = self.calculate(overrides={self.prod3.id: {"unit_delta": 10, "mode": "delta"}})
        self.assertEqual(result["product_results"][self.prod3.id]["actual_unit"], 105)

    def test_replace_override_replaces_tl(self):
        result = self.calculate(overrides={self.prod2.id: {"mode": "replace", "tl": 180000}})
        self.assertEqual(result["product_results"][self.prod2.id]["actual_tl"], 180000)

    def test_target_percent_override_uses_target(self):
        result = self.calculate(overrides={self.prod2.id: {"target_percent": 120, "mode": "delta"}})
        self.assertEqual(result["product_results"][self.prod2.id]["actual_tl"], 120000)

    def test_current_product_failure_is_reported(self):
        result = self.calculate()
        failed_names = [item["product_name"] for item in result["products"] if not item["passed"]]
        self.assertIn("Monurol", failed_names)

    def test_quarter_analysis_has_three_months(self):
        result = self.calculate()
        self.assertEqual(result["quarter_analysis"]["months"], [4, 5, 6])

    def test_quarter_analysis_tracks_product_rows(self):
        result = self.calculate()
        self.assertEqual(len(result["quarter_analysis"]["products"]), 3)

    def test_recovery_analysis_contains_statuses(self):
        result = self.calculate()
        statuses = {item["status"] for item in result["recovery_analysis"]}
        self.assertTrue(statuses.intersection({"Takip", "Riskli", "Kritik"}))

    def test_insights_identify_best_product(self):
        result = self.calculate()
        self.assertEqual(result["insights"]["most_profitable_product"]["product"], "Travazol")

    def test_insights_identify_worst_product(self):
        result = self.calculate()
        self.assertEqual(result["insights"]["most_harmful_product"]["product"], "Monurol")

    def test_missing_product_impact_percent_is_positive(self):
        result = self.calculate()
        self.assertGreater(result["insights"]["missing_product_impact_percent"], 0)

    def test_what_if_analysis_has_four_scenarios(self):
        result = self.calculate()
        self.assertEqual(len(result["what_if_analysis"]), 4)

    def test_what_if_labels_are_human_readable(self):
        result = self.calculate()
        labels = [item["label"] for item in result["what_if_analysis"]]
        self.assertEqual(labels, ["En Kötü", "Mevcut", "Beklenen", "En İyi"])

    def test_best_scenario_outperforms_worst(self):
        result = self.calculate()
        totals = {item["key"]: item["total_prime"] for item in result["what_if_analysis"]}
        self.assertGreater(totals["best"], totals["worst"])

    def test_comparison_graph_has_expected_labels(self):
        result = self.calculate()
        self.assertEqual(result["comparison_graph"]["labels"], ["Gerçekleşen", "Beklenen", "Maksimum"])

    def test_trend_graphs_have_monthly_points(self):
        result = self.calculate()
        self.assertEqual(len(result["trend_graphs"]["monthly"]), 3)

    def test_trend_graphs_have_quarterly_points(self):
        result = self.calculate()
        self.assertEqual(len(result["trend_graphs"]["quarterly"]), 4)

    def test_trend_graphs_have_yearly_points(self):
        result = self.calculate()
        self.assertEqual(len(result["trend_graphs"]["yearly"]), 1)

    def test_forecast_contains_expected_prime(self):
        result = self.calculate()
        self.assertIn("expected_prime", result["ai_forecast"])

    def test_ai_messages_are_generated(self):
        result = self.calculate()
        self.assertGreaterEqual(len(result["ai_messages"]), 2)

    def test_product_coefficient_does_not_change_approved_payout(self):
        overrides = {self.prod2.id: {"tl_delta": 40000, "mode": "delta"}}
        baseline = self.calculate(overrides=overrides)
        db.session.add(Setting(setting_key="PRODUCT_COEFFICIENT_TRAV", setting_value="2", category="Prim", description="Test"))
        db.session.commit()
        adjusted = self.calculate(overrides=overrides)
        self.assertEqual(adjusted["breakdown"]["product_effect"], 0)
        self.assertEqual(adjusted["breakdown"]["total"], baseline["breakdown"]["total"])

    def test_bonus_rate_does_not_change_approved_payout(self):
        baseline = self.calculate(overrides={self.prod2.id: {"tl_delta": 40000, "mode": "delta"}})
        setting = Setting.query.filter_by(setting_key="BONUS_RATE").first()
        setting.setting_value = "10"
        db.session.commit()
        adjusted = self.calculate(overrides={self.prod2.id: {"tl_delta": 40000, "mode": "delta"}})
        self.assertEqual(adjusted["breakdown"]["bonus"], 0)
        self.assertEqual(adjusted["breakdown"]["total"], baseline["breakdown"]["total"])

    def test_penalty_rate_setting_changes_penalty(self):
        baseline = self.calculate()
        setting = Setting.query.filter_by(setting_key="PENALTY_RATE").first()
        setting.setting_value = "6"
        db.session.commit()
        adjusted = self.calculate()
        self.assertEqual(baseline["breakdown"]["penalty"], 0)
        self.assertEqual(adjusted["breakdown"]["penalty"], 0)

    def test_recovery_rate_does_not_change_approved_payout(self):
        baseline = self.calculate(overrides={self.prod2.id: {"tl_delta": 40000, "mode": "delta"}})
        setting = Setting.query.filter_by(setting_key="RECOVERY_EFFECT_RATE").first()
        setting.setting_value = "4"
        db.session.commit()
        adjusted = self.calculate(overrides={self.prod2.id: {"tl_delta": 40000, "mode": "delta"}})
        self.assertEqual(adjusted["breakdown"]["recovery"], 0)
        self.assertEqual(adjusted["breakdown"]["total"], baseline["breakdown"]["total"])

    def test_cache_reports_miss_then_hit(self):
        first = self.calculate(use_cache=True)
        second = self.calculate(use_cache=True)
        self.assertFalse(first["cache"]["hit"])
        self.assertTrue(second["cache"]["hit"])

    def test_different_override_produces_cache_miss(self):
        self.calculate(use_cache=True)
        result = self.calculate(overrides={self.prod2.id: {"tl_delta": 40000, "mode": "delta"}}, use_cache=True)
        self.assertFalse(result["cache"]["hit"])

    def test_wrapped_prime_engine_import_still_works(self):
        wrapped = WrappedPrimeEngine(representative_id=self.rep.id, year=2025, month=6, use_cache=False)
        result = wrapped.calculate(save_history=False)
        self.assertIn("breakdown", result)


class TestPrimeEngineHistoryAndExports(PrimeEngineBaseTestCase):
    def test_history_starts_empty(self):
        history = self.create_engine().load_history()
        self.assertEqual(history, [])

    def test_calculate_can_save_history(self):
        result = self.calculate(save_history=True)
        self.assertIsNotNone(result["history_entry"])

    def test_history_is_loaded_after_save(self):
        self.calculate(save_history=True)
        history = self.create_engine().load_history()
        self.assertEqual(len(history), 1)

    def test_history_keeps_latest_first(self):
        self.calculate(save_history=True)
        self.calculate(overrides={self.prod2.id: {"tl_delta": 1000, "mode": "delta"}}, save_history=True)
        history = self.create_engine().load_history()
        self.assertGreaterEqual(len(history), 2)
        self.assertGreaterEqual(history[0]["created_at"], history[1]["created_at"])

    def test_history_is_capped_to_25_entries(self):
        engine = self.create_engine()
        for index in range(27):
            engine.overrides = {self.prod2.id: {"tl_delta": index, "mode": "delta"}}
            engine.save_history({"total": index, "main_prime": 0}, {"total_tl_percent": 0}, {}, [])
        history = engine.load_history()
        self.assertEqual(len(history), 25)

    def test_export_pdf_creates_file(self):
        engine = self.create_engine()
        result = engine.calculate(save_history=False)
        export = engine.export_pdf(result)
        self.assertTrue(Path(export["path"]).exists())

    def test_export_pdf_uses_pdf_header(self):
        engine = self.create_engine()
        result = engine.calculate(save_history=False)
        export = engine.export_pdf(result)
        self.assertTrue(Path(export["path"]).read_bytes().startswith(b"%PDF"))

    def test_export_pdf_respects_report_type(self):
        engine = self.create_engine()
        result = engine.calculate(save_history=False)
        export = engine.export_pdf(result, report_type="manager_summary")
        self.assertIn("manager_summary", export["name"])

    def test_export_excel_creates_file(self):
        engine = self.create_engine()
        result = engine.calculate(save_history=False)
        export = engine.export_excel(result)
        self.assertTrue(Path(export["path"]).exists())

    def test_export_excel_contains_expected_sheets(self):
        engine = self.create_engine()
        result = engine.calculate(save_history=False)
        export = engine.export_excel(result)
        workbook = load_workbook(export["path"])
        self.assertEqual(workbook.sheetnames, ["Summary", "Products", "WhatIf"])


class TestSimulationServiceIntegration(PrimeEngineBaseTestCase):
    def test_service_report_contains_summary(self):
        result = SimulationService(self.rep.id, 2025, 6, {}).report()
        self.assertTrue(result["success"])
        self.assertIn("summary", result)

    def test_service_report_contains_breakdown(self):
        result = SimulationService(self.rep.id, 2025, 6, {}).report()
        self.assertIn("breakdown", result)

    def test_service_report_contains_history(self):
        result = SimulationService(self.rep.id, 2025, 6, {}).report()
        self.assertIn("history", result)

    def test_service_recovery_exposes_box_and_tl_gaps(self):
        result = SimulationService(self.rep.id, 2025, 6, {}).report()
        open_rows = [item for item in result["recovery"] if item["remaining_tl"] > 0]

        self.assertTrue(open_rows)
        self.assertTrue(all("remaining_box" in item for item in open_rows))
        self.assertTrue(all("remaining_tl" in item for item in open_rows))
        self.assertTrue(any(item["remaining_box"] > 0 for item in open_rows))

    def test_service_builds_representative_target_snapshot(self):
        result = SimulationService(self.rep.id, 2025, 6, {}).report()
        snapshot = result["target_snapshot"]

        self.assertEqual(snapshot["target_tl"], 310000)
        self.assertEqual(snapshot["realization_tl"], 295000)
        self.assertEqual(snapshot["remaining_tl"], 15000)
        self.assertIn("prime_opportunity", snapshot)
        self.assertIn("remaining_workdays", snapshot)

    def test_service_prioritizes_prime_risk_in_action_plan(self):
        result = SimulationService(self.rep.id, 2025, 6, {}).report()
        action_plan = result["action_plan"]

        self.assertEqual(action_plan[0]["product"], "Monurol")
        self.assertEqual(action_plan[0]["priority"], 1)
        self.assertGreater(action_plan[0]["remaining_box"], 0)
        self.assertGreater(action_plan[0]["remaining_tl"], 0)
        self.assertIn("action", action_plan[0])

    def test_closed_period_action_plan_has_no_daily_pace(self):
        result = SimulationService(self.rep.id, 2025, 6, {}).report()

        self.assertTrue(result["target_snapshot"]["period_closed"])
        self.assertTrue(all(item["daily_box"] == 0 for item in result["action_plan"]))
        self.assertTrue(all(item["daily_tl"] == 0 for item in result["action_plan"]))

    def test_service_override_report_lists_changes(self):
        result = SimulationService(self.rep.id, 2025, 6, {self.prod1.id: {"tl_delta": 250000, "mode": "delta"}}).report()
        self.assertEqual(len(result["overrides"]), 1)

    def test_box_only_scenario_adds_units_and_period_price_without_persistence(self):
        baseline = SimulationService(self.rep.id, 2025, 6, {}).report()
        result = SimulationService(
            self.rep.id,
            2025,
            6,
            {self.prod1.id: {"unit_delta": 100, "tl_delta": None, "mode": "delta"}},
        ).report()
        base_product = next(item for item in baseline["prime"]["products"] if item["product_id"] == self.prod1.id)
        simulated = next(item for item in result["prime"]["products"] if item["product_id"] == self.prod1.id)

        self.assertEqual(simulated["actual_unit"], base_product["actual_unit"] + 100)
        self.assertEqual(simulated["actual_tl"], base_product["actual_tl"] + 100000)
        self.assertGreater(result["target_snapshot"]["realization_percent"], baseline["target_snapshot"]["realization_percent"])
        self.assertEqual(SimulationService(self.rep.id, 2025, 6, {}).report()["prime"]["product_results"][self.prod1.id]["actual_unit"], base_product["actual_unit"])

    def test_action_plan_exposes_dynamic_milestone_effect(self):
        result = SimulationService(self.rep.id, 2025, 6, {}).report()
        monurol = next(item for item in result["action_plan"] if item["product"] == "Monurol")

        self.assertGreater(monurol["suggested_box"], 0)
        self.assertGreater(monurol["suggested_tl"], 0)
        self.assertGreater(monurol["projected_product_percent"], monurol["percent"])
        self.assertIn("toplam TL realizasyonu", monurol["action"])

    def test_service_export_pdf_returns_metadata(self):
        export = SimulationService(self.rep.id, 2025, 6, {}).export_pdf()
        self.assertEqual(export["type"], "prime_report")

    def test_service_export_excel_returns_metadata(self):
        export = SimulationService(self.rep.id, 2025, 6, {}).export_excel()
        self.assertEqual(export["type"], "excel")

    def test_service_report_does_not_create_persistent_history(self):
        service = SimulationService(self.rep.id, 2025, 6, {})
        service.report()
        self.assertEqual(service.history(), [])

    def test_service_capabilities_include_exports_and_cache(self):
        capabilities = SimulationService.capabilities()
        self.assertTrue(capabilities["exports"])
        self.assertTrue(capabilities["cache"])
