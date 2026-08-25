from types import SimpleNamespace
from unittest import mock

import pandas as pd
from openpyxl import Workbook

from app.services.alias_service import AliasService
from app.services.competition_import_service import CompetitionImportService, SheetType
from app.services.dynamic_import_refinement import refined_competition_structure
from app.services.semantic_import_discovery import (
    _competition_signature_from_frame,
    install_semantic_import_discovery,
)
from app.services.target_import_service import TargetImportService
from app.services.official_brick_spread_atomic import _discover_spread_sheet
from app.services.official_brick_spread_service import OfficialBrickSpreadService


def test_renamed_competition_sheet_is_discovered_from_content():
    products = [f"RAKIP URUN {index}" for index in range(1, 14)]
    frame = pd.DataFrame([
        ["Ocak Çıkış TL", None, *products],
        ["BÖLGE", "TTS ISMI", *(["TL"] * len(products))],
        ["101", "AYSE KAYA", *range(1, len(products) + 1)],
    ])
    assert _competition_signature_from_frame(frame) == "competition_tl"


def test_broad_brick_market_without_representative_header_is_competition():
    products = [f"RAKIP URUN {index}" for index in range(1, 21)]
    frame = pd.DataFrame([
        ["Aylık Kutu Raporu", None, *products],
        ["BÖLGE", "IAM BRICK", *(["KUTU"] * len(products))],
        ["101", "KADIKOY MERKEZ", *range(1, len(products) + 1)],
    ])
    assert _competition_signature_from_frame(frame) == "competition_box"


def test_small_company_sales_sheet_is_not_guessed_as_competition():
    frame = pd.DataFrame([
        ["BÖLGE", "TTS ISMI", "TRAVAZOL TL", "MONUROL TL"],
        ["101", "AYSE KAYA", 100, 200],
    ])
    assert _competition_signature_from_frame(frame) is None


def test_renamed_dedicated_target_sheet_is_discovered_from_content():
    install_semantic_import_discovery()
    frame = pd.DataFrame([
        ["Aylık plan", None, None, None],
        ["Temsilci", "Travazol Hedef TL", "Monurol Hedef TL", "Kutu"],
        ["Ayşe Kaya", 100, 200, 1],
    ])

    def fake_product(value):
        normalized = AliasService.normalize(value)
        if "TRAVAZOL" in normalized or "MONUROL" in normalized:
            return {"matched": True, "method": "EXACT", "object": object()}
        return {"matched": False, "method": "NONE", "object": None}

    service = TargetImportService("unused.xlsx", upload_id=1)
    with mock.patch.object(AliasService, "find_product", side_effect=fake_product):
        assert service._is_target_sheet("tamamen-yeni-bir-ad", frame) is True


def test_renamed_official_brick_spread_is_discovered_from_content():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2028 yeni dağılım raporu"
    headers = ["BÖLGE", "TTS ISMI", "Brick Sayısı", "Travazol", "Monurol", "Acnemix", "Mixovul", "Stiderm", "Brimoder", "TOPLAM"]
    sheet.append(headers)
    sheet.append(["101", "AYSE KAYA", 10, 1, 2, 3, 4, 5, 6, 21])
    assert _discover_spread_sheet(OfficialBrickSpreadService, workbook) == sheet.title


def test_competition_placeholders_are_missing_but_numeric_zero_is_preserved():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "renamed market matrix"
    sheet.append(["BÖLGE", "IAM BRICK", "PRODUCT A", "PRODUCT B"])
    sheet.append(["101", "KADIKOY", "-", 0])

    service = CompetitionImportService(upload_id=77, year=2026, month=2)
    service._workbook = workbook
    records = service._parse_sheet_records({
        "sheet_name": sheet.title,
        "sheet_type": "monthly_competition_units",
        "period_type": "monthly",
        "year": 2026,
        "month": 2,
        "data_start_row": 2,
        "data_end_row": 2,
        "territory_column": 1,
        "subterritory_column": 2,
        "product_groups": {
            "MARKET": [("PRODUCT A", 3), ("PRODUCT B", 4)],
        },
    })

    assert len(records) == 1
    assert records[0]["product_name"] == "PRODUCT B"
    assert records[0]["metric_value"] == 0.0
    assert service.parse_statistics == {
        "numeric_cells": 1,
        "blank_cells": 1,
        "invalid_cells": 0,
    }
    assert service.invalid_cells == []


def test_competition_structure_prefers_brick_over_parent_subterritory():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "renamed units matrix"
    products = [f"RIVAL {index}" for index in range(1, 21)]
    sheet.append(["BÖLGE", "SUBTERRITORIES", "IAM BRICK", *products])
    sheet.append(["101 ISTANBUL", "308", "308001 KADIKOY", *([0] * len(products))])
    sheet.append(["101 ISTANBUL", "308", "308002 USKUDAR", *([40] * len(products))])

    service = CompetitionImportService(upload_id=1, year=2026, month=2)
    service._workbook = workbook
    groups = {"MARKET": [(name, column) for column, name in enumerate(products, start=4)]}
    with (
        mock.patch.object(service, "_discover_metadata", return_value=("monthly", 2026, 2)),
        mock.patch.object(service, "_extract_product_groups", return_value=groups),
        mock.patch.object(service, "get_sheet_type", return_value="monthly_competition_units"),
    ):
        structure = refined_competition_structure(service, sheet.title)

    assert structure["subterritory_column"] == 3


def test_turkish_representative_dimension_beats_parent_subterritory():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "localized market matrix"
    products = [f"RIVAL {index}" for index in range(1, 21)]
    sheet.append(["BÖLGE", "SUBTERRITORIES", "TTS İSMİ", *products])
    sheet.append(["101 ISTANBUL", "308", "REP A", *([0] * len(products))])
    sheet.append(["101 ISTANBUL", "308", "REP B", *([40] * len(products))])

    service = CompetitionImportService(upload_id=1, year=2026, month=2)
    service._workbook = workbook
    groups = {"MARKET": [(name, column) for column, name in enumerate(products, start=4)]}
    with (
        mock.patch.object(service, "_discover_metadata", return_value=("monthly", 2026, 2)),
        mock.patch.object(service, "_extract_product_groups", return_value=groups),
        mock.patch.object(service, "get_sheet_type", return_value="monthly_competition_units"),
    ):
        structure = refined_competition_structure(service, sheet.title)

    assert structure["territory_column"] == 1
    assert structure["subterritory_column"] == 3


def test_unlabeled_finer_dimension_is_inferred_from_content_grain():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "content grain matrix"
    products = [f"RIVAL {index}" for index in range(1, 21)]
    sheet.append(["BÖLGE", "SUBTERRITORIES", "WORK AREA", *products])
    sheet.append(["101 ISTANBUL", "308", "0001", *([0] * len(products))])
    sheet.append(["101 ISTANBUL", "308", "0002", *([40] * len(products))])

    service = CompetitionImportService(upload_id=1, year=2026, month=2)
    service._workbook = workbook
    groups = {"MARKET": [(name, column) for column, name in enumerate(products, start=4)]}
    with (
        mock.patch.object(service, "_discover_metadata", return_value=("monthly", 2026, 2)),
        mock.patch.object(service, "_extract_product_groups", return_value=groups),
        mock.patch.object(service, "get_sheet_type", return_value="monthly_competition_units"),
    ):
        structure = refined_competition_structure(service, sheet.title)

    assert structure["subterritory_column"] == 3


def test_fine_dimension_is_not_lost_after_many_coarse_candidates():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "wide dimension matrix"
    products = [f"RIVAL {index}" for index in range(1, 21)]
    sheet.append(["BÖLGE", "SUBTERRITORIES", "LEVEL A", "LEVEL B", "LEVEL C", "LEVEL D", "FINE CODE", *products])
    sheet.append(["101", "308", "P", "P", "P", "P", "0001", *([0] * len(products))])
    sheet.append(["101", "308", "P", "P", "P", "P", "0002", *([40] * len(products))])
    service = CompetitionImportService(upload_id=1, year=2026, month=2)
    service._workbook = workbook
    groups = {"MARKET": [(name, column) for column, name in enumerate(products, start=8)]}
    with (
        mock.patch.object(service, "_discover_metadata", return_value=("monthly", 2026, 2)),
        mock.patch.object(service, "_extract_product_groups", return_value=groups),
        mock.patch.object(service, "get_sheet_type", return_value="monthly_competition_units"),
    ):
        structure = refined_competition_structure(service, sheet.title)
    assert structure["subterritory_column"] == 7


def test_competition_acceptance_canonicalizes_physical_grain_not_business_values():
    from verify_ims_acceptance import _competition_semantic_rows, _fingerprint

    class Query:
        def __init__(self, rows):
            self.rows = rows
        def all(self):
            return self.rows

    def record(subterritory, value):
        return SimpleNamespace(
            period_type="monthly", year=2026, month=2, week_number=None,
            territory="101", subterritory=subterritory,
            product_group="MARKET", product_name="RIVAL", metric_type="UNIT",
            metric_value=value, is_subtotal=False, is_grand_total=False,
        )

    legacy = Query([record("308", 40), record("308", 0), record("308", 60)])
    semantic = Query([record("0001", 40), record("0002", 60)])
    changed = Query([record("0001", 40), record("0002", 61)])

    baseline = _fingerprint(_competition_semantic_rows(legacy))
    assert _fingerprint(_competition_semantic_rows(semantic)) == baseline
    assert _fingerprint(_competition_semantic_rows(changed)) != baseline


def test_competition_product_text_cannot_override_upload_month():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["AUGUST MARKET PRODUCT", "WEEKLY BRAND NAME"])
    service = CompetitionImportService(upload_id=1, year=2026, month=2, week_number=7)
    period_type, year, month = service._discover_metadata(sheet)
    assert (year, month) == (2026, 2)


def test_competition_product_text_cannot_override_semantic_period_type():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "renamed monthly matrix"
    products = [f"RIVAL {index}" for index in range(1, 19)]
    sheet.append(["BÖLGE", "BRICK", *products])
    sheet.append(["101", "0001", *([10] * len(products))])
    sheet.append(["101", "0002", *([20] * len(products))])
    service = CompetitionImportService(upload_id=1, year=2026, month=2, week_number=7)
    service._workbook = workbook
    groups = {"MARKET": [(name, column) for column, name in enumerate(products, start=3)]}
    with (
        mock.patch.object(service, "_discover_metadata", return_value=("WEEKLY", 2026, 2)),
        mock.patch.object(service, "_extract_product_groups", return_value=groups),
        mock.patch.object(
            service,
            "get_sheet_type",
            return_value=SheetType.MONTHLY_COMPETITION_UNITS.value,
        ),
    ):
        structure = refined_competition_structure(service, sheet.title)

    assert structure["period_type"] == "MONTHLY"


def test_content_classification_overrides_misleading_legacy_competition_name():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "AYLIK REKABET"
    products = [f"RIVAL {index}" for index in range(1, 19)]
    sheet.append(["AYLIK"])
    sheet.append(["BÖLGE", "IAM BRICK", "1 TTS ISMI", "2 TTS ISMI", *products])
    sheet.append(["101", "0001", "AYSE", "FATMA", *([10] * len(products))])
    service = CompetitionImportService(upload_id=1, year=2026, month=2)
    service._workbook = workbook
    install_semantic_import_discovery()

    assert service.get_sheet_type(sheet.title) == SheetType.MONTHLY_COMPETITION_UNITS.value


def test_competition_metric_uses_semantic_type_not_misleading_sheet_name():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MARKET REPORT"
    sheet.append(["BÖLGE", "BRICK", "RIVAL"])
    sheet.append(["101", "0001", 125.5])
    service = CompetitionImportService(upload_id=1, year=2026, month=2)
    service._workbook = workbook
    records = service._parse_sheet_records({
        "sheet_name": sheet.title,
        "sheet_type": SheetType.MONTHLY_COMPETITION_VALUE.value,
        "period_type": "MONTHLY",
        "year": 2026,
        "month": 2,
        "data_start_row": 2,
        "data_end_row": 2,
        "territory_column": 1,
        "subterritory_column": 2,
        "product_groups": {"MARKET": [("RIVAL", 3)]},
    })

    assert len(records) == 1
    assert records[0]["metric_type"] == "TL"
    assert records[0]["metric_value"] == 125.5
