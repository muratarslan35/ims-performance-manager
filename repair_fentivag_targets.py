#!/usr/bin/env python
"""Idempotently repair missing Fentivag targets from the authoritative workbook.

The repair is deliberately narrow:
- only product_code=FENTIVAG is written;
- target TL comes from BAKİYE -> OCAK HEDEF TL;
- target unit comes from canonical cumulative TTS unit actual + BAKİYE MF-siz
  remaining unit, i.e. the workbook's own box equation, never TL / unit price;
- actual TL/unit come from the canonical TTS HAFTALIK ÇIKIŞLARI cumulative block;
- all non-Fentivag Target rows are fingerprinted before/after and must be byte-for-byte
  equivalent at the value level before commit.
"""

from __future__ import annotations

import hashlib
import json
import math
import re
from pathlib import Path

from openpyxl import load_workbook

from app import create_app
from app.extensions import db
from app.models import IMSUpload, IMSSummary, Product, Representative, Target
from app.services.alias_service import AliasService

YEAR = 2026
MONTH = 1
QUARTER = "Q1"
PRODUCT_CODE = "FENTIVAG"
TOTAL_LABELS = {"NATIONAL", "TOPLAM", "TOTAL", "GRAND TOTAL", "GENEL TOPLAM"}


def safe_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return 0.0
        return float(value)
    text = str(value).strip().replace("\u00a0", "")
    if not text or text.upper() in {"NAN", "NONE", "-", "#DIV/0!"}:
        return 0.0
    text = re.sub(r"[^0-9,.-]", "", text)
    if text.count(",") and text.count("."):
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif text.count(","):
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def norm(value) -> str:
    return AliasService.normalize(value)


def _is_rep_row(location, representative) -> bool:
    rep = norm(representative)
    loc = norm(location)
    if not rep or rep in TOTAL_LABELS:
        return False
    if rep == loc:
        return False
    if rep in {"TTS ISMI", "1 TTS ISMI", "2 TTS ISMI"}:
        return False
    return True


def _find_sheet(workbook, *tokens):
    for sheet_name in workbook.sheetnames:
        normalized = norm(sheet_name)
        if all(token in normalized for token in tokens):
            return workbook[sheet_name]
    raise RuntimeError(f"Required worksheet not found: {tokens}")


def _find_source_file() -> Path:
    upload = (
        IMSUpload.query.filter_by(year=YEAR, month=MONTH, status="COMPLETED")
        .order_by(IMSUpload.completed_at.desc(), IMSUpload.id.desc())
        .first()
    )
    names = []
    if upload and upload.file_name:
        names.append(upload.file_name)
    names.append("Tayfun-1_3.Hafta_Ocak_Brick_Analizi_.xlsx")

    for name in names:
        for root in (Path("uploads"), Path("instance/uploads"), Path("instance")):
            candidate = root / name
            if candidate.is_file():
                return candidate

    candidates = sorted(Path("uploads").glob("*Ocak*.xlsx"))
    if len(candidates) == 1:
        return candidates[0]
    raise RuntimeError(f"January source workbook could not be resolved safely: {candidates}")


def _find_bakiye_columns(header):
    columns = {}
    section = ""
    for index, value in enumerate(header):
        label = norm(value)
        if "HEDEF" in label and "TL" in label:
            section = "target_tl"
            continue
        if "CIKIS" in label and "TL" in label:
            section = "actual_tl_balance"
            continue
        if "BAKIYE" in label and "KUTU" in label:
            section = "balance_unit"
            continue
        if "BAKIYE" in label and "TL" in label:
            section = "balance_tl"
            continue
        if label == PRODUCT_CODE and section:
            columns[section] = index
    required = {"target_tl", "balance_unit"}
    missing = required - set(columns)
    if missing:
        raise RuntimeError(f"BAKİYE Fentivag columns missing: {sorted(missing)}; found={columns}")
    return columns


def _extract_bakiye(workbook):
    sheet = _find_sheet(workbook, "BAKIYE")
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next(
        (
            index
            for index, row in enumerate(rows[:12])
            if "HEDEF" in norm(" ".join(str(v or "") for v in row))
            and PRODUCT_CODE in norm(" ".join(str(v or "") for v in row))
            and "BAKIYE" in norm(" ".join(str(v or "") for v in row))
        ),
        None,
    )
    if header_index is None:
        raise RuntimeError("BAKİYE target header not found")
    columns = _find_bakiye_columns(rows[header_index])

    records = {}
    national = {"target_tl": 0.0, "balance_unit": 0.0}
    for row in rows[header_index + 1 :]:
        location = row[0] if len(row) > 0 else None
        representative = row[1] if len(row) > 1 else None
        rep_norm = norm(representative)
        if rep_norm == "NATIONAL":
            national["target_tl"] = safe_float(row[columns["target_tl"]])
            national["balance_unit"] = safe_float(row[columns["balance_unit"]])
            continue
        if not _is_rep_row(location, representative):
            continue
        records[rep_norm] = {
            "location": str(location or "").strip(),
            "representative": str(representative or "").strip(),
            "target_tl": safe_float(row[columns["target_tl"]]),
            "balance_unit": safe_float(row[columns["balance_unit"]]),
        }
    return records, national


def _first_cumulative_columns(section_row, product_row):
    sections = {}
    selected = {"tl": False, "unit": False}
    current = ""
    for index, value in enumerate(section_row):
        label = norm(value)
        if label:
            if "CIKIS" in label and "TL" in label:
                if not selected["tl"]:
                    current = "tl"
                    selected["tl"] = True
                else:
                    current = ""
            elif "CIKIS" in label and ("KUTU" in label or "UNIT" in label):
                if not selected["unit"]:
                    current = "unit"
                    selected["unit"] = True
                else:
                    current = ""
            elif "HAFTA" in label:
                current = ""
        sections[index] = current

    columns = {}
    for index, product in enumerate(product_row):
        if norm(product) != PRODUCT_CODE:
            continue
        metric = sections.get(index)
        if metric in {"tl", "unit"} and metric not in columns:
            columns[metric] = index
    if set(columns) != {"tl", "unit"}:
        raise RuntimeError(f"Canonical TTS Fentivag columns missing: {columns}")
    return columns


def _extract_weekly_actuals(workbook):
    sheet = _find_sheet(workbook, "HAFTALIK", "CIKIS")
    rows = list(sheet.iter_rows(values_only=True))
    product_header_index = next(
        (
            index
            for index, row in enumerate(rows[:12])
            if PRODUCT_CODE in norm(" ".join(str(v or "") for v in row))
        ),
        None,
    )
    if product_header_index is None or product_header_index == 0:
        raise RuntimeError("TTS weekly product header not found")
    columns = _first_cumulative_columns(rows[product_header_index - 1], rows[product_header_index])

    actuals = {}
    national = {"tl": 0.0, "unit": 0.0}
    for row in rows[product_header_index + 1 :]:
        location = row[0] if len(row) > 0 else None
        representative = row[1] if len(row) > 1 else None
        rep_norm = norm(representative)
        if rep_norm == "NATIONAL":
            national["tl"] = safe_float(row[columns["tl"]])
            national["unit"] = safe_float(row[columns["unit"]])
            continue
        if not _is_rep_row(location, representative):
            continue
        actuals[rep_norm] = {
            "tl": safe_float(row[columns["tl"]]),
            "unit": safe_float(row[columns["unit"]]),
        }
    return actuals, national


def extract_source_records(source_file: Path):
    workbook = load_workbook(source_file, data_only=True, read_only=True)
    try:
        targets, target_national = _extract_bakiye(workbook)
        actuals, actual_national = _extract_weekly_actuals(workbook)
    finally:
        workbook.close()

    records = []
    for rep_norm, target in targets.items():
        actual = actuals.get(rep_norm, {"tl": 0.0, "unit": 0.0})
        records.append(
            {
                **target,
                "source_key": rep_norm,
                "actual_tl": actual["tl"],
                "actual_unit": actual["unit"],
                "unit_target": target["balance_unit"] + actual["unit"],
            }
        )

    target_tl_sum = sum(item["target_tl"] for item in records)
    unit_target_sum = sum(item["unit_target"] for item in records)
    expected_tl = target_national["target_tl"]
    expected_unit = target_national["balance_unit"] + actual_national["unit"]
    if not math.isclose(target_tl_sum, expected_tl, rel_tol=0, abs_tol=0.05):
        raise RuntimeError(f"Fentivag TL source reconciliation failed: reps={target_tl_sum}, national={expected_tl}")
    if not math.isclose(unit_target_sum, expected_unit, rel_tol=0, abs_tol=0.05):
        raise RuntimeError(f"Fentivag unit source reconciliation failed: reps={unit_target_sum}, national={expected_unit}")
    if not math.isclose(sum(item["actual_tl"] for item in records), actual_national["tl"], rel_tol=0, abs_tol=0.05):
        raise RuntimeError("Fentivag canonical TL actual reconciliation failed")
    if not math.isclose(sum(item["actual_unit"] for item in records), actual_national["unit"], rel_tol=0, abs_tol=0.05):
        raise RuntimeError("Fentivag canonical unit actual reconciliation failed")
    return records


def _target_fingerprint(product_id: int) -> str:
    rows = (
        Target.query.filter(Target.product_id != product_id)
        .order_by(Target.id)
        .all()
    )
    payload = [
        [
            row.id,
            row.year,
            row.month,
            row.quarter,
            row.representative_id,
            row.product_id,
            row.unit_target,
            row.tl_target,
            row.unit_realization,
            row.tl_realization,
            row.realization_percent,
            row.prime_percent,
            row.bonus_amount,
        ]
        for row in rows
    ]
    return hashlib.sha256(json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")).hexdigest()


def _period_representatives():
    rep_ids = {
        rep_id
        for (rep_id,) in db.session.query(Target.representative_id)
        .filter(Target.year == YEAR, Target.month == MONTH)
        .distinct()
        .all()
    }
    representatives = Representative.query.filter(Representative.id.in_(rep_ids)).all() if rep_ids else []
    exact = {norm(rep.rep_name): rep for rep in representatives}
    return rep_ids, representatives, exact


def _resolve_rep(record, representatives, exact):
    source_key = record["source_key"]
    if source_key in exact:
        return exact[source_key]

    location_match = re.match(r"^(\d{3})\b", norm(record["location"]))
    location_code = location_match.group(1) if location_match else None
    candidates = []
    for rep in representatives:
        rep_norm = norm(rep.rep_name)
        if location_code and str(rep.region or "") != location_code:
            continue
        if source_key in rep_norm or rep_norm in source_key:
            candidates.append(rep)
    if len(candidates) == 1:
        return candidates[0]

    match = AliasService.find_representative(record["representative"])
    if match["matched"] and match["object"] in representatives:
        return match["object"]
    return None


def run_repair():
    product = Product.query.filter(db.func.upper(Product.product_code) == PRODUCT_CODE).first()
    if product is None:
        raise RuntimeError("Fentivag product master is missing; apply migrations first")
    if not product.is_active:
        raise RuntimeError("Fentivag product is still inactive; apply migrations first")

    source_file = _find_source_file()
    source_records = extract_source_records(source_file)
    period_rep_ids, representatives, exact = _period_representatives()
    if len(source_records) != len(period_rep_ids):
        raise RuntimeError(
            f"Representative count mismatch; source={len(source_records)} existing_period={len(period_rep_ids)}"
        )

    before_fingerprint = _target_fingerprint(product.id)
    upload = (
        IMSUpload.query.filter_by(year=YEAR, month=MONTH, status="COMPLETED")
        .order_by(IMSUpload.completed_at.desc(), IMSUpload.id.desc())
        .first()
    )

    unmatched = []
    touched = 0
    for record in source_records:
        representative = _resolve_rep(record, representatives, exact)
        if representative is None:
            unmatched.append(record["representative"])
            continue

        target = Target.query.filter_by(
            year=YEAR,
            month=MONTH,
            representative_id=representative.id,
            product_id=product.id,
        ).first()
        if target is None:
            target = Target(
                year=YEAR,
                month=MONTH,
                quarter=QUARTER,
                representative_id=representative.id,
                product_id=product.id,
            )
            db.session.add(target)

        target.tl_target = record["target_tl"]
        target.unit_target = record["unit_target"]
        target.tl_realization = record["actual_tl"]
        target.unit_realization = record["actual_unit"]
        target.realization_percent = (
            record["actual_tl"] * 100.0 / record["target_tl"] if record["target_tl"] else 0.0
        )

        summary = IMSSummary.query.filter_by(
            year=YEAR,
            month=MONTH,
            representative_id=representative.id,
            product_id=product.id,
        ).first()
        if summary is None:
            summary = IMSSummary(
                upload_id=upload.id if upload else None,
                representative_id=representative.id,
                product_id=product.id,
                year=YEAR,
                month=MONTH,
                quarter=QUARTER,
            )
            db.session.add(summary)
        summary.unit = record["actual_unit"]
        summary.tl = record["actual_tl"]
        summary.target_unit = record["unit_target"]
        summary.target_tl = record["target_tl"]
        summary.realization_percent = target.realization_percent
        touched += 1

    if unmatched:
        db.session.rollback()
        raise RuntimeError(f"Unmatched Fentivag representatives: {sorted(unmatched)}")

    db.session.flush()
    after_fingerprint = _target_fingerprint(product.id)
    if before_fingerprint != after_fingerprint:
        db.session.rollback()
        raise RuntimeError("Non-Fentivag target fingerprint changed; repair aborted")

    fentivag_rows = Target.query.filter_by(year=YEAR, month=MONTH, product_id=product.id).count()
    if fentivag_rows != len(source_records):
        db.session.rollback()
        raise RuntimeError(f"Fentivag row count mismatch after repair: db={fentivag_rows}, source={len(source_records)}")

    db.session.commit()
    print("=== FENTIVAG_TARGET_REPAIR ===")
    print(f"SOURCE_FILE|{source_file}")
    print(f"SOURCE_ROWS|{len(source_records)}")
    print(f"TOUCHED_ROWS|{touched}")
    print(f"DB_FENTIVAG_ROWS|{fentivag_rows}")
    print(f"FENTIVAG_TL_SUM|{sum(item['target_tl'] for item in source_records):.6f}")
    print(f"FENTIVAG_UNIT_SUM|{sum(item['unit_target'] for item in source_records):.6f}")
    print(f"NON_FENTIVAG_FINGERPRINT|{after_fingerprint}")
    print("STATUS|SUCCESS")


if __name__ == "__main__":
    app = create_app()
    with app.app_context():
        run_repair()
