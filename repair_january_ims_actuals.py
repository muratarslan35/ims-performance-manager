#!/usr/bin/env python
"""Guardedly reconcile January IMS actuals to the workbook's canonical TTS source.

This repair is intentionally narrow:
- person TL/unit actuals come only from the first cumulative TL/KUTU blocks in
  ``TTS HAFTALIK ÇIKIŞLARI``;
- target TL/unit allocations are never rewritten;
- official NATIONAL/region subtotal KPI rows are refreshed through the same
  import helper used by future uploads;
- CompetitionData must remain byte-for-byte equivalent at the value level;
- all 113 period cadres and all 7 managed products must map exactly once.
"""

from __future__ import annotations

import hashlib
import json
import math
import re
from pathlib import Path

from openpyxl import load_workbook

from app import create_app
from app.extensions import db
from app.models import CompetitionData, IMSUpload, IMSSummary, Product, Representative, Target
from app.services.alias_service import AliasService
from app.services.ims_import_service import IMSImportService

YEAR = 2026
MONTH = 1
PRODUCT_CODES = ("TRAVAZOL", "MONUROL", "MIXOVUL", "FENTIVAG", "STIDERM", "ACNEMIX", "BRIMODER")
TOTAL_LABELS = {"NATIONAL", "TOPLAM", "TOTAL", "GRAND TOTAL", "GENEL TOPLAM"}


def norm(value) -> str:
    return AliasService.normalize(value)


def safe_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return 0.0
        return float(value)
    text = str(value).strip().replace("\u00a0", "")
    if not text or text.upper() in {"NAN", "NONE", "-", "#DIV/0!"}:
        return 0.0
    text = re.sub(r"[^0-9,.-]", "", text)
    if text.count(",") and text.count("."):
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif text.count(","):
        text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def region_code(value) -> str:
    match = re.search(r"\b(\d{3})\b", norm(value))
    return match.group(1) if match else ""


def is_person_row(location, representative) -> bool:
    rep = norm(representative)
    loc = norm(location)
    if not rep or rep in TOTAL_LABELS or rep == loc:
        return False
    if rep in {"TTS ISMI", "1 TTS ISMI", "2 TTS ISMI"}:
        return False
    return True


def _find_weekly_sheet(workbook):
    for sheet_name in workbook.sheetnames:
        label = norm(sheet_name)
        if "HAFTALIK" in label and "CIKIS" in label:
            return workbook[sheet_name]
    raise RuntimeError("TTS HAFTALIK ÇIKIŞLARI source sheet not found")


def _canonical_columns(section_row, product_row):
    sections = {}
    selected = {"tl": False, "unit": False}
    current = ""
    for index, value in enumerate(section_row):
        label = norm(value)
        if label:
            if "CIKIS" in label and "TL" in label:
                if not selected["tl"]:
                    current = "tl"
                    selected["tl"] = True
                else:
                    current = ""
            elif "CIKIS" in label and ("KUTU" in label or "UNIT" in label):
                if not selected["unit"]:
                    current = "unit"
                    selected["unit"] = True
                else:
                    current = ""
            elif "HAFTA" in label:
                current = ""
        sections[index] = current

    columns = {code: {} for code in PRODUCT_CODES}
    for index, value in enumerate(product_row):
        product_code = norm(value)
        metric = sections.get(index)
        if product_code in columns and metric in {"tl", "unit"} and metric not in columns[product_code]:
            columns[product_code][metric] = index

    missing = {code: sorted({"tl", "unit"} - set(metrics)) for code, metrics in columns.items() if set(metrics) != {"tl", "unit"}}
    if missing:
        raise RuntimeError(f"Canonical weekly TL/KUTU product columns missing: {missing}")
    return columns


def extract_weekly_actuals(source_file: Path):
    workbook = load_workbook(source_file, data_only=True, read_only=True)
    try:
        sheet = _find_weekly_sheet(workbook)
        rows = list(sheet.iter_rows(values_only=True))
        header_index = next(
            (
                index
                for index, row in enumerate(rows[:12])
                if sum(1 for value in row if norm(value) in PRODUCT_CODES) >= 5
            ),
            None,
        )
        if header_index is None or header_index == 0:
            raise RuntimeError("Canonical weekly product header not found")
        columns = _canonical_columns(rows[header_index - 1], rows[header_index])

        records = []
        for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            location = row[0] if len(row) > 0 else None
            representative = row[1] if len(row) > 1 else None
            if not is_person_row(location, representative):
                continue
            values = {
                code: {
                    "tl": safe_float(row[metrics["tl"]]),
                    "unit": safe_float(row[metrics["unit"]]),
                }
                for code, metrics in columns.items()
            }
            records.append(
                {
                    "source_row": row_index,
                    "location": str(location or "").strip(),
                    "representative": str(representative or "").strip(),
                    "source_key": norm(representative),
                    "region_code": region_code(location),
                    "values": values,
                }
            )
        return records
    finally:
        workbook.close()


def _period_representatives(year, month):
    rep_ids = {
        rep_id
        for (rep_id,) in db.session.query(Target.representative_id)
        .filter(Target.year == year, Target.month == month)
        .distinct()
        .all()
    }
    representatives = Representative.query.filter(Representative.id.in_(rep_ids)).all() if rep_ids else []
    exact = {norm(rep.rep_name): rep for rep in representatives}
    return rep_ids, representatives, exact


def resolve_representative(record, representatives, exact):
    source_key = record["source_key"]
    if source_key in exact:
        return exact[source_key]

    source_region = record["region_code"]
    candidates = []
    for rep in representatives:
        rep_norm = norm(rep.rep_name)
        rep_region = region_code(rep.region) or region_code(rep.territory)
        if source_region and rep_region and source_region != rep_region:
            continue
        if source_key in rep_norm or rep_norm in source_key:
            candidates.append(rep)
    if len(candidates) == 1:
        return candidates[0]

    # Vacancy names vary between workbook generations (BOS / BOS KADRO / BOS BRICK).
    # Resolve them only when the source region has exactly one vacant period cadre.
    if any(token in source_key.split() for token in ("BOS", "KADRO", "BRICK")):
        vacancy_candidates = []
        for rep in representatives:
            rep_norm = norm(rep.rep_name)
            rep_region = region_code(rep.region) or region_code(rep.territory)
            if source_region and rep_region and source_region != rep_region:
                continue
            if "BOS" in rep_norm.split() or "KADRO" in rep_norm.split():
                vacancy_candidates.append(rep)
        if len(vacancy_candidates) == 1:
            return vacancy_candidates[0]

    match = AliasService.find_representative(record["representative"])
    if match["matched"] and match["object"] in representatives:
        return match["object"]
    return None


def _target_business_fingerprint(year, month):
    digest = hashlib.sha256()
    rows = Target.query.filter_by(year=year, month=month).order_by(Target.id).all()
    for row in rows:
        payload = [
            row.id, row.year, row.month, row.quarter, row.representative_id, row.product_id,
            row.unit_target, row.tl_target, row.prime_percent, row.bonus_amount,
        ]
        digest.update(json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest(), len(rows)


def _competition_fingerprint(upload_id):
    digest = hashlib.sha256()
    count = 0
    query = CompetitionData.query.filter_by(upload_id=upload_id).order_by(CompetitionData.id)
    for row in query.yield_per(1000):
        payload = [
            row.id, row.upload_id, row.sheet_name, row.period_type, row.year, row.month,
            row.week_number, row.territory, row.subterritory, row.product_group,
            row.product_name, row.is_company_product, row.is_competitor,
            row.metric_type, row.metric_value, row.is_subtotal, row.is_grand_total,
            row.source_row,
        ]
        digest.update(json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8"))
        digest.update(b"\n")
        count += 1
    return digest.hexdigest(), count


def _resolve_source_file(upload):
    candidates = []
    if upload and upload.file_name:
        candidates.extend([Path("uploads") / upload.file_name, Path("instance/uploads") / upload.file_name])
    candidates.append(Path("uploads/Tayfun-1_3.Hafta_Ocak_Brick_Analizi_.xlsx"))
    for candidate in candidates:
        if candidate.is_file():
            return candidate
    raise RuntimeError(f"January source workbook not found: {candidates}")


def repair_period(year=YEAR, month=MONTH, source_file=None):
    upload = (
        IMSUpload.query.filter_by(year=year, month=month, status="COMPLETED")
        .order_by(IMSUpload.completed_at.desc(), IMSUpload.id.desc())
        .first()
    )
    if upload is None:
        raise RuntimeError(f"Completed IMS upload not found for {year}/{month}")
    source_file = Path(source_file) if source_file else _resolve_source_file(upload)
    source_records = extract_weekly_actuals(source_file)

    rep_ids, representatives, exact = _period_representatives(year, month)
    if len(source_records) != len(rep_ids):
        raise RuntimeError(f"Representative coverage mismatch: source={len(source_records)} period={len(rep_ids)}")

    products = {}
    for code in PRODUCT_CODES:
        product = Product.query.filter(db.func.upper(Product.product_code) == code).first()
        if product is None:
            raise RuntimeError(f"Managed product missing: {code}")
        products[code] = product

    before_target_fp, before_target_count = _target_business_fingerprint(year, month)
    before_comp_fp, before_comp_count = _competition_fingerprint(upload.id)

    mapped_ids = set()
    mapped = []
    for record in source_records:
        representative = resolve_representative(record, representatives, exact)
        if representative is None:
            raise RuntimeError(f"Unmatched weekly representative: {record['representative']} ({record['location']})")
        if representative.id in mapped_ids:
            raise RuntimeError(f"Duplicate weekly representative mapping: {record['representative']} -> {representative.rep_name}")
        mapped_ids.add(representative.id)
        mapped.append((record, representative))
    if mapped_ids != rep_ids:
        missing_ids = sorted(rep_ids - mapped_ids)
        raise RuntimeError(f"Period representatives missing from weekly source mapping: {missing_ids}")

    touched = 0
    for record, representative in mapped:
        for code, source in record["values"].items():
            product = products[code]
            target = Target.query.filter_by(
                year=year, month=month, representative_id=representative.id, product_id=product.id
            ).first()
            summary = IMSSummary.query.filter_by(
                year=year, month=month, representative_id=representative.id, product_id=product.id
            ).first()
            if target is None or summary is None:
                raise RuntimeError(
                    f"Missing target/summary for {representative.rep_name} / {code}: "
                    f"target={bool(target)} summary={bool(summary)}"
                )

            actual_tl = source["tl"]
            actual_unit = source["unit"]
            summary.tl = actual_tl
            summary.unit = actual_unit
            summary.target_tl = target.tl_target or 0
            summary.target_unit = target.unit_target or 0
            summary.realization_percent = actual_tl * 100.0 / target.tl_target if target.tl_target else 0.0
            target.tl_realization = actual_tl
            target.unit_realization = actual_unit
            target.realization_percent = summary.realization_percent
            touched += 1

    # Persist official NATIONAL and 11 region subtotal rows using the production
    # import implementation that future uploads will use as well.
    source_service = IMSImportService(str(source_file))
    source_service.upload = upload
    source_service.load_workbook(str(source_file))
    try:
        source_service.persist_national_dashboard_metrics(year, month)
    finally:
        source_service.workbook = None

    db.session.flush()

    after_target_fp, after_target_count = _target_business_fingerprint(year, month)
    after_comp_fp, after_comp_count = _competition_fingerprint(upload.id)
    summary_count = IMSSummary.query.filter_by(year=year, month=month).count()
    if before_target_fp != after_target_fp or before_target_count != after_target_count:
        raise RuntimeError("Target allocation fingerprint changed during actual repair")
    if before_comp_fp != after_comp_fp or before_comp_count != after_comp_count:
        raise RuntimeError("CompetitionData changed during IMS actual repair")
    if touched != len(rep_ids) * len(PRODUCT_CODES):
        raise RuntimeError(f"Unexpected repaired cell count: {touched}")
    if summary_count != touched:
        raise RuntimeError(f"Summary coverage mismatch after repair: summaries={summary_count} touched={touched}")

    db.session.commit()
    return {
        "source_file": str(source_file),
        "representatives": len(mapped_ids),
        "products": len(PRODUCT_CODES),
        "touched": touched,
        "target_rows": after_target_count,
        "summary_rows": summary_count,
        "target_fingerprint": after_target_fp,
        "competition_rows": after_comp_count,
        "competition_fingerprint": after_comp_fp,
    }


def run_repair():
    app = create_app()
    with app.app_context():
        try:
            result = repair_period()
            print("JANUARY_IMS_ACTUAL_REPAIR|" + json.dumps(result, ensure_ascii=False, sort_keys=True))
            print("STATUS|SUCCESS")
        except Exception:
            db.session.rollback()
            raise


if __name__ == "__main__":
    run_repair()
