#!/usr/bin/env python3
"""Read-only Jan-Mar source-cell audit for Diyarbakir 901 Monurol."""

import json
from pathlib import Path

from openpyxl import load_workbook

from app import create_app
from app.models import Product
from app.services.alias_service import AliasService
from app.services.production_result_import_service import ProductionResultImportService
from app.services.production_result_service import ProductionResultService


def source_cell(sheet, row, column):
    cell = sheet.cell(row, column)
    return {"address": cell.coordinate, "value_or_formula": cell.value}


def main():
    app = create_app()
    with app.app_context():
        product = Product.query.filter(Product.product_name.ilike("Monurol")).one()
        upload_root = Path(app.config["UPLOAD_FOLDER"]) / "production_results"
        for month in (1, 2, 3):
            upload = ProductionResultService.final_upload(2026, month)
            path = upload_root / upload.stored_file_name
            values_book = load_workbook(path, read_only=False, data_only=True)
            formulas_book = load_workbook(path, read_only=False, data_only=False)
            reader = ProductionResultImportService(path, 2026, month, upload.production_stage)
            reader._load_master_maps()
            result = {"month": month, "file": upload.file_name, "stage": upload.production_stage}
            for metric in ("TL", "KUTU"):
                sheet = reader._find_sheet(values_book, metric)
                layout = reader._layout(sheet, metric)
                row = next(
                    number for number in range(layout["header_row"] + 1, sheet.max_row + 1)
                    if AliasService.normalize(sheet.cell(number, layout["name_column"]).value).startswith("901 ")
                )
                index = layout["product_ids"].index(product.id)
                target_column = layout["target_columns"][index]
                actual_column = layout["actual_columns"][index]
                percent_column = layout["percent_columns"][index]
                formula_sheet = formulas_book[sheet.title]
                target = sheet.cell(row, target_column).value
                actual = sheet.cell(row, actual_column).value
                result[metric] = {
                    "sheet": sheet.title,
                    "row": row,
                    "target": target,
                    "actual": actual,
                    "calculated_percent": actual * 100 / target,
                    "target_source": source_cell(formula_sheet, row, target_column),
                    "actual_source": source_cell(formula_sheet, row, actual_column),
                    "percent_source": source_cell(formula_sheet, row, percent_column),
                }
            result["cross_metric"] = {
                "target_tl_per_box": result["TL"]["target"] / result["KUTU"]["target"],
                "actual_tl_per_box": result["TL"]["actual"] / result["KUTU"]["actual"],
            }
            print("MONUROL_901_SOURCE|" + json.dumps(result, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
