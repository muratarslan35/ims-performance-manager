"""Workbook import service implementing IMSRawData -> IMSFact -> IMSSummary."""

import json
import logging
logger = logging.getLogger(__name__)
import logging
logger = logging.getLogger(__name__)
import logging
logger = logging.getLogger(__name__)
import math
import logging
logger = logging.getLogger(__name__)
import os
import logging
logger = logging.getLogger(__name__)
import re
import logging
logger = logging.getLogger(__name__)
import time
import logging
logger = logging.getLogger(__name__)
from collections import Counter
from datetime import datetime

import pandas as pd
import logging
logger = logging.getLogger(__name__)
from sqlalchemy import func
from sqlalchemy.exc import SQLAlchemyError
from openpyxl import load_workbook as openpyxl_load_workbook

from app.extensions import db
from app.models import (
    IMSFact,
    IMSRawData,
    IMSSummary,
    IMSUpload,
    ImportAuditLog,
    ManualMatchQueue,
    Product,
    Representative,
    RepresentativeBrickAssignment,
    Target,
)
from app.services.alias_service import AliasService
from app.services.competition_import_service import CompetitionImportService
from app.services.target_import_service import TargetImportService
from app.services.target_box_calculation_service import TargetBoxCalculationService


# Regex to extract week number from typical IMS file names.
# Examples: "Tayfun-1 24.Hafta Haziran Brick Analizi_.xlsx"
#           "25.Hafta Mayıs IMS.xlsx"
_WEEK_REGEX = re.compile(r"(\d{1,2})\s*\.?\s*hafta", re.IGNORECASE)
logger = logging.getLogger(__name__)


class IMSImportService:
    """Import a workbook in three explicit, auditable ETL stages.

    Raw data is never used by reporting logic.  It is first staged exactly as
    read from the workbook, then transformed into matched facts, and finally
    aggregated into period summaries.
    """

    REPORT_SHEETS = {
        "BRICK SATIS": "brick_sales",
        "BRICK REA": "brick_realization",
        "BAKIYE": "balance",
        "HAFTALIK": "weekly_sales",
        "REKABET TL": "competition_tl",
        "REKABET KUTU": "competition_box",
        "REKABET PP": "competition_pp",
    }
    MONTH_TOKENS = {
        "OCAK": 1, "JANUARY": 1, "JAN": 1, "ŞUBAT": 2, "FEBRUARY": 2, "FEB": 2,
        "MART": 3, "MARCH": 3, "MAR": 3, "NİSAN": 4, "APRIL": 4, "APR": 4,
        "MAYIS": 5, "MAY": 5, "HAZİRAN": 6, "JUNE": 6, "JUN": 6,
        "TEMMUZ": 7, "JULY": 7, "JUL": 7, "AĞUSTOS": 8, "AUGUST": 8, "AUG": 8,
        "EYLÜL": 9, "SEPTEMBER": 9, "SEP": 9, "EKİM": 10, "OCTOBER": 10, "OCT": 10,
        "KASIM": 11, "NOVEMBER": 11, "NOV": 11, "ARALIK": 12, "DECEMBER": 12, "DEC": 12,
    }
    REPRESENTATIVE_HEADERS = {
        "TEMSILCI",
        "REPRESENTATIVE",
        "MUMESSIL",
        "REP",
        "ADI SOYADI",
        "SATIS TEMSILCISI",
        "TTS ISMI",
    }
    REGION_HEADERS = {"BOLGE", "REGION"}
    PROVINCE_HEADERS = {"IL", "SEHIR", "CITY", "PROVINCE"}
    BRICK_HEADERS = {"IAM BRICK", "BRICK", "SUBTERRITORIES", "SUBTERRITORY", "TERRITORY"}
    MANAGER_HEADERS = {"MANAGER", "MUDUR", "MÜDÜR", "BOLGE MUDURU", "BÖLGE MÜDÜRÜ", "2 TTS ISMI"}
    PRODUCT_GROUP_HEADERS = {"URUN GRUBU", "PRODUCT GROUP", "MARKA", "URUN", "PRODUCT"}
    PRODUCT_HEADER_NOISE_TOKENS = {
        "SUBTOTAL",
        "TOPLAM",
        "TOTAL",
        "PAZAR",
        "MARKET",
        "GRUP",
        "GROUP",
        "HEDEF",
        "CIKIS",
        "ÇIKIŞ",
        "NATIONAL",
        "GRAND",
    }
    REPRESENTATIVE_NOISE_TOKENS = {
        "SUBTOTAL",
        "TOPLAM",
        "TOTAL",
        "NATIONAL",
        "GRAND",
        "BOS",
        "KADRO",
        "BRICK",
    }
    STRICT_PRODUCT_MATCH_METHODS = {
        "MATCH_TABLE",
        "EXACT",
        "ALIAS_EXACT",
        "NORMALIZED",
        "ALIAS_NORMALIZED",
        "CONTAINS",
        "ALIAS_CONTAINS",
    }
    NORMALIZED_SHEET_TYPES = {
        "TL": "competition_tl",
        "CIRO": "competition_tl",
        "TUTAR": "competition_tl",
        "KUTU": "competition_box",
        "BOX": "competition_box",
        "ADET": "competition_box",
        "MARKET": "competition_pp",
        "PAZAR": "competition_pp",
        "PAY": "competition_pp",
        "PP": "competition_pp",
    }
    TOTAL_LABELS = {"NATIONAL", "TOPLAM", "GRAND TOTAL", "TOTAL", "GENEL TOPLAM"}
    WRITE_BATCH_SIZE = 1000
    MAX_SKIPPED_LOGS_PER_SHEET = 100
    NOISE_ROW_TOKENS = {
        "NOT",
        "NOTE",
        "ACIKLAMA",
        "AÇIKLAMA",
        "COMMENTS",
        "YORUM",
        "BILGI",
        "BİLGİ",
    }

    def __init__(self, file_path, uploaded_by=None):
        self.file_path = str(file_path)
        self.uploaded_by = uploaded_by
        self.started = time.monotonic()
        self.upload = None
        self.workbook = None
        self.errors = []
        self.warnings = []
        self.unknown_products = []
        self.unknown_columns = []
        self.statistics = {
            "sheet_count": 0,
            "processed_sheets": 0,
            "processed_rows": 0,
            "raw_records": 0,
            "fact_records": 0,
            "facts_inserted": 0,
            "facts_updated": 0,
            "summary_records": 0,
            "matched_products": 0,
            "matched_representatives": 0,
            "vacancy_records": 0,
            "unmatched_representatives": 0,
            "unmatched_products": 0,
            "unmatched_regions": 0,
            "unmatched_provinces": 0,
            "queued_for_manual": 0,
            "skipped_records": 0,
            "rows_error": 0,
            "source_metric_records": 0,
            "stored_source_records": 0,
            "zero_metric_records": 0,
            "blank_metric_records": 0,
            "invalid_metric_records": 0,
            "unmatched_product_records": 0,
            "unresolved_representative_rows": 0,
            "aggregate_rows_excluded": 0,
            "reconciliation_difference": 0,
            "reconciliation_status": "PENDING",
        }
        self.skipped_logs = []
        self.excluded_logs = []
        self._skipped_log_counts = {}
        self.parser_decisions = []
        self._representative_match_cache = {}
        self._product_match_cache = {}
        self._pending_raw_records = 0
        self._raw_batch = []

    @staticmethod
    def extract_week_number(file_name):
        """Extract week number from an IMS file name (e.g. '24.Hafta' -> 24)."""
        match = _WEEK_REGEX.search(os.path.basename(file_name))
        if match:
            week = int(match.group(1))
            if 1 <= week <= 53:
                return week
        return None

    @staticmethod
    def quarter_for(month):
        if month < 1 or month > 12:
            raise ValueError("Ay değeri 1 ile 12 arasında olmalıdır.")
        return f"Q{((month - 1) // 3) + 1}"

    @staticmethod
    def safe_float(value):
        if value is None or (isinstance(value, float) and (math.isnan(value) or math.isinf(value))):
            return 0.0
        if isinstance(value, bool):
            return float(value)
        if isinstance(value, (int, float)):
            return float(value)

        text = str(value).strip().replace("\u00a0", "")
        if not text or text.upper() in {"NAN", "NONE", "-"}:
            return 0.0
        text = re.sub(r"[^0-9,.-]", "", text)
        if text.count(",") and text.count("."):
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif text.count(","):
            text = text.replace(".", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return 0.0

    @staticmethod
    def _json_dump(value):
        return json.dumps(value, ensure_ascii=False, default=str, sort_keys=True)

    @staticmethod
    def _value_for_json(value):
        if value is None or (isinstance(value, float) and (math.isnan(value) or math.isinf(value))):
            return None
        if isinstance(value, (str, int, float, bool)):
            return value
        return str(value)

    @staticmethod
    def clean_text(value):
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return ""
        text = str(value).strip()
        return "" if AliasService.normalize(text) in {"", "NAN", "NONE"} else text

    def create_upload(self, year, month, week_number=None):
        self.upload = IMSUpload(
            file_name=os.path.basename(self.file_path),
            year=year,
            month=month,
            week_number=week_number,
            quarter=self.quarter_for(month),
            uploaded_by=self.uploaded_by,
            status="PROCESSING",
        )
        db.session.add(self.upload)
        db.session.flush()
        return self.upload

    def load_workbook(self):
        workbook = {}
        hidden_rows_by_sheet = self._hidden_rows_by_sheet()
        raw_workbook = pd.read_excel(self.file_path, sheet_name=None, header=None)
        for sheet_name, dataframe in raw_workbook.items():
            hidden_rows = hidden_rows_by_sheet.get(str(sheet_name), set())
            workbook[str(sheet_name)] = self._sanitize_sheet_dataframe(dataframe, hidden_rows=hidden_rows)
        self.workbook = workbook
        self.statistics["sheet_count"] = len(self.workbook)
        return self.workbook

    def detect_workbook_month(self):
        """Read the reporting month from workbook labels, not a sheet name."""
        for dataframe in (self.workbook or {}).values():
            preview = dataframe.iloc[:5, :10].fillna("").astype(str).to_numpy().ravel()
            text = " ".join(preview).upper()
            for token, detected_month in self.MONTH_TOKENS.items():
                if re.search(r"(?<![A-ZÇĞİÖŞÜ])" + re.escape(token) + r"(?![A-ZÇĞİÖŞÜ])", text):
                    return detected_month
        return None

    def _hidden_rows_by_sheet(self):
        hidden_by_sheet = {}
        workbook = None
        try:
            workbook = openpyxl_load_workbook(self.file_path, data_only=True, read_only=False)
            for worksheet in workbook.worksheets:
                hidden_rows = {
                    row_number - 1
                    for row_number, dimensions in worksheet.row_dimensions.items()
                    if getattr(dimensions, "hidden", False)
                }
                hidden_by_sheet[str(worksheet.title)] = hidden_rows
        except Exception as exc:  # pragma: no cover
            logger.warning(f"ims_import_service: Could not read hidden rows dimensions. {exc}")
            return {}
        finally:
            if workbook is not None:
                workbook.close()
        return hidden_by_sheet

    def _sanitize_sheet_dataframe(self, dataframe, hidden_rows=None):
        prepared = dataframe.copy()
        if hidden_rows:
            prepared = prepared.drop(index=[idx for idx in hidden_rows if idx in prepared.index], errors="ignore")
        prepared = prepared.dropna(axis=0, how="all").dropna(axis=1, how="all")
        prepared.reset_index(drop=True, inplace=True)
        return prepared

    def find_header_row(self, dataframe):
        max_rows = min(80, len(dataframe))
        best_index = None
        best_score = -1
        for index in range(max_rows):
            values = [AliasService.normalize(value) for value in dataframe.iloc[index].tolist()]
            row_text = " ".join(values)
            score = 0
            if any(candidate in row_text for candidate in self.REPRESENTATIVE_HEADERS):
                score += 3
            if any(candidate in row_text for candidate in self.PRODUCT_GROUP_HEADERS):
                score += 2
            if any(candidate in row_text for candidate in self.REGION_HEADERS | self.PROVINCE_HEADERS):
                score += 1
            if any(candidate in row_text for candidate in self.NORMALIZED_SHEET_TYPES):
                score += 1
            if score > best_score:
                best_index = index
                best_score = score
        if best_score >= 3:
            return best_index
        return None

    def normalize_header(self, value):
        return AliasService.normalize(value)

    def build_headers(self, dataframe, header_row):
        headers = []
        used_headers = {}
        header_rows = [row_index for row_index in range(max(0, header_row - 2), header_row + 1)]

        # IMS exports use merged group headers (for example ``Haziran TL``
        # above several product columns).  Pandas leaves the continuation
        # cells empty, so preserve the group label horizontally before
        # building each product header.  This keeps the metric semantic
        # (TL/KUTU/PP) without hard-coding a month or column number.
        contextual_headers = {}
        for row_index in header_rows[:-1]:
            # A single-cell report title is not a column-group header. Carrying
            # it across the row makes words such as "BRICK" classify the
            # actual metric column as a dimension and drops every data row.
            populated_cells = sum(
                1
                for value in dataframe.iloc[row_index].tolist()
                if self.clean_text(value)
            )
            if populated_cells <= 1:
                for column_index in range(dataframe.shape[1]):
                    contextual_headers[(row_index, column_index)] = ""
                continue
            current_label = ""
            for column_index in range(dataframe.shape[1]):
                value = self.clean_text(dataframe.iloc[row_index, column_index])
                if value:
                    current_label = self.normalize_header(value)
                contextual_headers[(row_index, column_index)] = current_label

        for column_index in range(dataframe.shape[1]):
            parts = []
            for row_index in header_rows:
                if row_index == header_row:
                    token = self.normalize_header(self.clean_text(dataframe.iloc[row_index, column_index]))
                else:
                    token = contextual_headers[(row_index, column_index)]
                if token and token not in parts:
                    parts.append(token)

            header = " ".join(parts) or f"COLUMN_{column_index + 1}"
            duplicate_count = used_headers.get(header, 0)
            used_headers[header] = duplicate_count + 1
            headers.append(header if duplicate_count == 0 else f"{header}_{duplicate_count + 1}")

        result = dataframe.iloc[header_row + 1 :].copy()
        result.columns = headers
        result.reset_index(drop=True, inplace=True)
        return result

    def detect_sheet_type(self, sheet_name):
        normalized_name = AliasService.normalize(sheet_name)
        for sheet_label, sheet_type in self.REPORT_SHEETS.items():
            if AliasService.normalize(sheet_label) in normalized_name:
                return sheet_type
        for token, sheet_type in self.NORMALIZED_SHEET_TYPES.items():
            if token in normalized_name:
                return sheet_type
        return "unknown"

    def detect_dimension_columns(self, dataframe):
        dimensions = {
            "representative": None,
            "region": None,
            "province": None,
            "product_group": None,
            "brick": None,
            "manager": None,
        }
        for column in dataframe.columns:
            normalized = AliasService.normalize(column)
            if dimensions["representative"] is None and (
                normalized in self.REPRESENTATIVE_HEADERS
                or any(candidate in normalized for candidate in self.REPRESENTATIVE_HEADERS if len(candidate) > 3)
                or ("TEMSILCI" in normalized and "KOD" not in normalized)
            ):
                dimensions["representative"] = column
                continue
            if dimensions["region"] is None and any(token in normalized for token in self.REGION_HEADERS):
                dimensions["region"] = column
                continue
            if dimensions["province"] is None and any(token in normalized for token in self.PROVINCE_HEADERS):
                dimensions["province"] = column
                continue
            if dimensions["product_group"] is None and any(
                token in normalized for token in self.PRODUCT_GROUP_HEADERS
            ) and not any(
                token in normalized for token in ("TL", "CIRO", "VALUE", "KUTU", "BOX", "UNIT", "ADET")
            ):
                dimensions["product_group"] = column
                continue
            if dimensions["brick"] is None and any(token in normalized for token in self.BRICK_HEADERS):
                dimensions["brick"] = column
                continue
            if dimensions["manager"] is None and any(token in normalized for token in self.MANAGER_HEADERS):
                dimensions["manager"] = column
                continue
        return dimensions

    def detect_metric_kind(self, sheet_name, dataframe):
        normalized_name = AliasService.normalize(sheet_name)
        for token, sheet_type in self.NORMALIZED_SHEET_TYPES.items():
            if token in normalized_name:
                if sheet_type == "competition_tl":
                    return "tl"
                if sheet_type == "competition_box":
                    return "unit"
                if sheet_type == "competition_pp":
                    return "market_share"

        headers = " ".join(AliasService.normalize(column) for column in dataframe.columns)
        if any(token in headers for token in ("TL", "CIRO", "TUTAR", "VALUE")):
            return "tl"
        if any(token in headers for token in ("KUTU", "BOX", "ADET", "UNIT")):
            return "unit"
        return "market_share"

    def detect_metric_column(self, dataframe, dimensions, metric_kind):
        dimension_columns = {column for column in dimensions.values() if column}
        metric_tokens = {
            "tl": ("TL", "CIRO", "TUTAR", "VALUE"),
            "unit": ("KUTU", "BOX", "ADET", "UNIT"),
            "market_share": ("PAZAR", "PAY", "MARKET", "SHARE", "PP"),
        }
        candidates = [column for column in dataframe.columns if column not in dimension_columns]
        for column in candidates:
            normalized = AliasService.normalize(column)
            if any(token in normalized for token in metric_tokens[metric_kind]):
                return column

        scored = []
        for column in candidates:
            sample = dataframe[column].head(20)
            numeric_count = sum(1 for value in sample if self.safe_float(value) != 0.0)
            if numeric_count:
                scored.append((numeric_count, column))
        if scored:
            scored.sort(reverse=True)
            return scored[0][1]
        return None

    def _log_skipped_row(self, reason, sheet_name, source_row, **context):
        count = self._skipped_log_counts.get(sheet_name, 0)
        if count >= self.MAX_SKIPPED_LOGS_PER_SHEET:
            self._skipped_log_counts[sheet_name] = count + 1
            return
        self._skipped_log_counts[sheet_name] = count + 1
        payload = {"reason": reason, "sheet_name": sheet_name, "source_row": source_row, **context}
        self.skipped_logs.append(payload)
        logger.warning("ims_import_skipped_row %s", self._json_dump(payload))

    def _log_warning(self, reason, sheet_name, source_row, **context):
        payload = {"reason": reason, "sheet_name": sheet_name, "source_row": source_row, **context}
        self.warnings.append(self._json_dump(payload))

    def _log_excluded_row(self, reason, sheet_name, source_row, **context):
        payload = {"reason": reason, "sheet_name": sheet_name, "source_row": source_row, **context}
        self.excluded_logs.append(payload)
        logger.info("ims_import_excluded_row %s", self._json_dump(payload))

    def _finalize_source_reconciliation(self):
        source = self.statistics["source_metric_records"]
        classified = (
            self.statistics["stored_source_records"]
            + self.statistics["blank_metric_records"]
            + self.statistics["invalid_metric_records"]
            + self.statistics["unmatched_product_records"]
        )
        difference = source - classified
        self.statistics["reconciliation_difference"] = difference
        blocking = {
            "unclassified_records": difference,
            "blank_metric_records": self.statistics["blank_metric_records"],
            "invalid_metric_records": self.statistics["invalid_metric_records"],
            "unmatched_product_records": self.statistics["unmatched_product_records"],
            "unresolved_representative_rows": self.statistics["unresolved_representative_rows"],
            "row_errors": self.statistics["rows_error"],
        }
        failed = any(value != 0 for value in blocking.values())
        self.statistics["reconciliation_status"] = "FAILED" if failed else "PASSED"
        self._log_stage_metrics(
            "source_reconciliation",
            source_metric_records=source,
            stored_source_records=self.statistics["stored_source_records"],
            zero_metric_records=self.statistics["zero_metric_records"],
            aggregate_rows_excluded=self.statistics["aggregate_rows_excluded"],
            competition_source_records=self.statistics.get("competition_source_records", 0),
            competition_records=self.statistics.get("competition_records", 0),
            competition_duplicates=self.statistics.get("competition_duplicates", 0),
            competition_invalid=self.statistics.get("competition_invalid", 0),
            **blocking,
        )
        if failed:
            raise ValueError(f"IMS veri bütünlüğü doğrulaması başarısız: {self._json_dump(blocking)}")

    def _log_stage_metrics(self, stage, **metrics):
        payload = {"stage": stage, "upload_id": self.upload.id if self.upload else None, **metrics}
        logger.info("ims_import_stage_metrics %s", self._json_dump(payload))

    def parse_metric_value(self, value):
        if value is None or (isinstance(value, float) and (math.isnan(value) or math.isinf(value))):
            return 0.0, "empty"
        if isinstance(value, (int, float, bool)):
            return self.safe_float(value), "valid"
        text = str(value).strip()
        if not text or AliasService.normalize(text) in {"", "NAN", "NONE", "-"}:
            return 0.0, "empty"
        parsed = self.safe_float(text)
        has_numeric_marker = bool(re.search(r"\d", text))
        if has_numeric_marker or parsed != 0.0:
            return parsed, "valid"
        return parsed, "invalid"

    def resolve_representative_match(self, representative_name):
        normalized = AliasService.normalize(representative_name)
        if normalized not in self._representative_match_cache:
            self._representative_match_cache[normalized] = AliasService.find_representative(representative_name)
        return self._representative_match_cache[normalized]

    def resolve_product_match(self, product_group_name):
        normalized = AliasService.normalize(product_group_name)
        if normalized not in self._product_match_cache:
            self._product_match_cache[normalized] = AliasService.find_product(product_group_name)
        return self._product_match_cache[normalized]

    def _unique_product_code(self, product_name):
        base = re.sub(r"[^A-Z0-9]+", "", AliasService.normalize(product_name))[:24] or "PRODUCT"
        candidate = base
        suffix = 1
        while Product.query.filter_by(product_code=candidate).first() is not None:
            suffix += 1
            suffix_text = f"-{suffix}"
            candidate = f"{base[:30-len(suffix_text)]}{suffix_text}"
        return candidate

    def _ensure_product(self, product_group_name):
        """Resolve or create a product from an explicit IMS product-group value."""
        name = self.clean_text(product_group_name)
        normalized = AliasService.normalize(name)
        if not normalized:
            return None, False
        tokens = set(normalized.split())
        noise_tokens = {AliasService.normalize(token) for token in self.PRODUCT_HEADER_NOISE_TOKENS}
        if any(
            token == noise or (len(noise) >= 4 and token.startswith(noise))
            for token in tokens for noise in noise_tokens
        ):
            return None, False
        match = self.resolve_product_match(product_group_name)
        if match["matched"]:
            return match["object"], True
        product = Product(
            product_code=self._unique_product_code(name), product_name=name, ims_name=name,
            category="IMS AUTO", competitor_group=name, is_active=True,
            is_prime_product=False, include_total_tl=True,
        )
        db.session.add(product)
        db.session.flush()
        AliasService.refresh()
        self._product_match_cache[normalized] = AliasService.find_product(name)
        self.warnings.append(f"Yeni ürün grubu otomatik oluşturuldu ({name}).")
        return product, False

    def analyze_sheet(self, sheet_name, dataframe):
        header_row = self.find_header_row(dataframe)
        if header_row is None:
            self.warnings.append(f"{sheet_name}: temsilci başlığı bulunamadı; sayfa atlandı.")
            return None

        return {
            "sheet_name": str(sheet_name),
            "sheet_type": self.detect_sheet_type(sheet_name),
            "header_row": header_row,
            "dataframe": self.build_headers(dataframe, header_row),
        }

    def analyze_workbook(self):
        return [
            analysis
            for sheet_name, dataframe in self.workbook.items()
            if (analysis := self.analyze_sheet(sheet_name, dataframe)) is not None
        ]

    def detect_representative_column(self, dataframe):
        candidates = self.detect_representative_columns(dataframe)
        return candidates[0] if candidates else None

    def detect_representative_columns(self, dataframe):
        candidates = []
        for column in dataframe.columns:
            normalized = AliasService.normalize(column)
            if normalized in self.REPRESENTATIVE_HEADERS:
                candidates.append(column)
                continue
            if any(candidate in normalized for candidate in self.REPRESENTATIVE_HEADERS if len(candidate) > 3):
                candidates.append(column)

        if not candidates:
            return []

        def representative_score(column_name):
            values = dataframe[column_name].fillna("").astype(str).head(250)
            score = 0
            for value in values:
                text = self.clean_text(value)
                if not text:
                    continue
                normalized_value = AliasService.normalize(text)
                if normalized_value in self.TOTAL_LABELS:
                    continue
                has_letters = bool(re.search(r"[A-ZÇĞİÖŞÜ]", normalized_value))
                has_digits = bool(re.search(r"\d", normalized_value))
                if has_letters:
                    score += 2
                if " " in text:
                    score += 1
                if has_digits:
                    score -= 2
            return score

        ranked = sorted(candidates, key=representative_score, reverse=True)
        seen = set()
        unique_ranked = []
        for column in ranked:
            if str(column) in seen:
                continue
            seen.add(str(column))
            unique_ranked.append(column)
        return unique_ranked

    def _is_aggregate_label(self, text):
        normalized = AliasService.normalize(text)
        if not normalized:
            return True
        if normalized in self.TOTAL_LABELS:
            return True
        if normalized == "NATIONAL":
            return True
        if re.match(r"^\d{3}\s+[A-ZÇĞİÖŞÜ]+$", normalized):
            return True
        return False

    def _is_probable_representative_name(self, text):
        normalized = AliasService.normalize(text)
        if not normalized or self._is_aggregate_label(text):
            return False
        if bool(re.search(r"\d", normalized)):
            return False
        tokens = normalized.split()
        if any(token in self.REPRESENTATIVE_NOISE_TOKENS for token in tokens):
            return False
        if len(tokens) < 2:
            return False
        return bool(re.search(r"[A-ZÇĞİÖŞÜ]", normalized))

    def _is_vacancy_representative(self, text):
        """Identify explicit empty-headcount rows without matching place names such as Bostancı."""
        return "BOS" in set(AliasService.normalize(text).split())

    def _region_context(self, value, fallback_city=None):
        location = self.clean_text(value)
        match = re.match(r"^(\d{3})\s+(.+)$", location)
        if match:
            return match.group(1), match.group(2).strip()
        return None, self.clean_text(fallback_city) or None

    def _find_vacancy_placeholder(self, vacancy_name):
        """Match a source vacancy label (e.g. IZM BOS BRICK) to its BAKIYE region."""
        ignored = {"BOS", "KADRO", "BRICK"}
        source = " ".join(token for token in AliasService.normalize(vacancy_name).split() if token not in ignored)
        if not source:
            return None
        for representative in Representative.query.filter(Representative.rep_code.like("UNASSIGNED%")).all():
            city = AliasService.normalize(representative.city or "")
            if city and (city.startswith(source) or source.startswith(city)):
                return representative
        return None

    @staticmethod
    def _vacancy_code(region, vacancy_name):
        """Return a stable ID for each distinct unassigned headcount row."""
        identity = re.sub(r"[^A-Z0-9]+", "", AliasService.normalize(vacancy_name))[:48]
        return f"UNASSIGNED{region or 'GENERAL'}{identity or 'VACANCY'}"

    @staticmethod
    def _vacancy_label(region, city, vacancy_name):
        context = " ".join(part for part in (region, city) if part)
        return f"ATANMAMIŞ · {context} · {vacancy_name}".strip(" ·")

    def _migrate_legacy_vacancy_placeholders(self):
        """Split legacy one-per-region placeholders before importing a workbook.

        Earlier imports used ``UNASSIGNED<region>`` for every vacancy in the
        same region, causing the last BAKİYE row to overwrite the preceding
        empty-headcount rows.  Preserve its primary key by assigning it to the
        last source row (the value it currently owns); other rows are created
        with their own stable IDs during normal processing.
        """
        sheet_name = next((name for name in self.workbook if "BAKIYE" in AliasService.normalize(name)), None)
        if not sheet_name:
            return
        vacancies_by_region = {}
        for _, row in self.workbook[sheet_name].iterrows():
            vacancy_name = self.clean_text(row.iloc[1]) if len(row) > 1 else ""
            if not self._is_vacancy_representative(vacancy_name):
                continue
            region, city = self._region_context(self.clean_text(row.iloc[0]))
            if region:
                vacancies_by_region.setdefault(region, []).append((city, vacancy_name))
        for region, vacancies in vacancies_by_region.items():
            legacy = Representative.query.filter_by(rep_code=f"UNASSIGNED{region}").first()
            if legacy is None:
                continue
            city, vacancy_name = vacancies[-1]
            code = self._vacancy_code(region, vacancy_name)
            if Representative.query.filter_by(rep_code=code).first() is None:
                legacy.rep_code = code
                legacy.rep_name = self._vacancy_label(region, city, vacancy_name)
                legacy.city = city or legacy.city

    def _ensure_vacancy_representative(self, region_value=None, city=None, vacancy_name=None):
        """Store vacant brick activity under an active, region-owned cadre slot."""
        region, location_city = self._region_context(region_value, city)
        if not region and vacancy_name:
            matched = self._find_vacancy_placeholder(vacancy_name)
            if matched is not None:
                self.statistics["vacancy_records"] += 1
                return matched.id
        code = self._vacancy_code(region, vacancy_name)
        representative = Representative.query.filter_by(rep_code=code).first()
        if representative is None:
            representative = Representative(
                rep_code=code,
                rep_name=self._vacancy_label(region or "GENEL", location_city, vacancy_name or "BOŞ KADRO"),
                region=region,
                city=location_city,
                territory=location_city,
                team="TAYFUN-1",
                active=True,
            )
            db.session.add(representative)
            db.session.flush()
        self.statistics["vacancy_records"] += 1
        return representative.id

    def bootstrap_vacancy_representatives_from_balance(self):
        """Create region placeholders before brick rows are parsed, using BAKIYE's hierarchy."""
        sheet_name = next((name for name in self.workbook if "BAKIYE" in AliasService.normalize(name)), None)
        if not sheet_name:
            return
        self._migrate_legacy_vacancy_placeholders()
        frame = self.workbook[sheet_name]
        for _, row in frame.iterrows():
            if len(row) < 2:
                continue
            vacancy_name = self.clean_text(row.iloc[1])
            if self._is_vacancy_representative(vacancy_name):
                self._ensure_vacancy_representative(self.clean_text(row.iloc[0]), vacancy_name=vacancy_name)

    @staticmethod
    def _remove_legacy_general_vacancy_facts(year, month):
        """Facts are derived data; rebuild old generic vacancy facts from the current upload."""
        generic = Representative.query.filter_by(rep_code="UNASSIGNEDGENERAL").first()
        if generic is not None:
            IMSFact.query.filter_by(year=year, month=month, representative_id=generic.id).delete(synchronize_session=False)

    def _ensure_representative(self, name, *, territory=None, manager=None, region=None, city=None):
        match = self.resolve_representative_match(name)
        if match["matched"]:
            self.statistics["matched_representatives"] += 1
            representative = match["object"]
            changed = False
            if territory and not representative.territory:
                representative.territory = territory
                changed = True
            if manager and not representative.manager:
                representative.manager = manager
                changed = True
            if region and not representative.region:
                representative.region = region
                changed = True
            if city and not representative.city:
                representative.city = city
                changed = True
            if changed:
                db.session.flush()
            return representative.id, True

        self.statistics["unmatched_representatives"] += 1
        normalized = AliasService.normalize(name)
        # Master representative codes are stable, readable identifiers derived
        # from the Excel name; legacy AUTO- prefixes must never be introduced.
        base_code = re.sub(r'[^A-Z0-9]+', '', normalized)[:18] or "REP"
        candidate_code = base_code
        suffix = 1
        while Representative.query.filter_by(rep_code=candidate_code).first() is not None:
            suffix += 1
            candidate_code = f"{base_code[:18]}-{suffix}"

        representative = Representative(
            rep_code=candidate_code,
            rep_name=name,
            territory=territory,
            manager=manager,
            region=region,
            city=city,
            team=self._team_for_context(region, city),
            active=True,
        )
        db.session.add(representative)
        db.session.flush()
        AliasService.refresh()
        self._representative_match_cache[normalized] = AliasService.find_representative(name)
        self.statistics["matched_representatives"] += 1
        return representative.id, False

    @staticmethod
    def _team_for_context(region=None, city=None):
        """Reuse the area's team only when the workbook context is unambiguous."""
        query = Representative.query.filter(Representative.team.isnot(None), Representative.team != "")
        if region:
            query = query.filter_by(region=region)
        elif city:
            query = query.filter_by(city=city)
        else:
            return None
        teams = {str(row.team).strip() for row in query.all() if str(row.team or "").strip()}
        return next(iter(teams)) if len(teams) == 1 else None

    def _resolve_representative(
        self,
        *,
        representative_name,
        allow_auto_create,
        sheet_name,
        source_row,
        territory=None,
        manager=None,
        region=None,
        city=None,
    ):
        if allow_auto_create:
            return self._ensure_representative(
                representative_name,
                territory=territory,
                manager=manager,
                region=region,
                city=city,
            )

        representative_match = self.resolve_representative_match(representative_name)
        if representative_match["matched"]:
            self.statistics["matched_representatives"] += 1
            return representative_match["object"].id, True

        self.statistics["unmatched_representatives"] += 1
        self.warnings.append(
            f"{sheet_name} satır {source_row}: temsilci eşleşmedi ({representative_name})."
        )
        self._log_skipped_row(
            reason="unmatched_representative",
            sheet_name=sheet_name,
            source_row=source_row,
            representative=representative_name,
        )
        best = representative_match.get("object")
        AliasService.enqueue_unmatched_representative(
            ims_name=representative_name,
            upload_id=self.upload.id,
            best_candidate=best.rep_name if best else None,
            best_score=representative_match.get("score", 0.0),
            worksheet=sheet_name,
            row_number=source_row,
            reason="unmatched_representative",
        )
        self.statistics["queued_for_manual"] += 1
        return None, False

    @staticmethod
    def metric_for_column(header):
        normalized = AliasService.normalize(header)
        if "VALUE SHARE" in normalized or "DEGER PAY" in normalized:
            return "value_share"
        if "TL" in normalized or "CIRO" in normalized or "VALUE" in normalized:
            return "tl"
        if "KUTU" in normalized or "BOX" in normalized or "UNIT" in normalized or "ADET" in normalized:
            return "unit"
        if "PAY" in normalized or "SHARE" in normalized or normalized.endswith(" PP"):
            return "market_share"
        if "GROWTH" in normalized or "BUYUME" in normalized:
            return "growth"
        return "unit"

    def detect_product_columns(self, dataframe, representative_column, *, allow_auto_create=False):
        products = {}
        seen_metric_pairs = set()
        new_label_counts = Counter(
            label
            for column in dataframe.columns
            if column != representative_column
            if (label := self._product_label_from_metric_header(column))
        )
        for column_index, header in enumerate(dataframe.columns):
            if header == representative_column:
                continue
            normalized_header = AliasService.normalize(header)
            match = AliasService.find_product(header)
            product = (
                match.get("object")
                if match.get("matched") and match.get("method") in self.STRICT_PRODUCT_MATCH_METHODS
                else None
            )
            if product is None:
                product_label = self._product_label_from_metric_header(header)
                # Wide layouts do not have an explicit product-group cell.
                # Require two corroborating metric columns (normally box + TL)
                # so a lone competitor column cannot become a managed product.
                if not allow_auto_create or not product_label or new_label_counts[product_label] < 2:
                    continue
                product, _ = self._ensure_product(product_label)
                if product is None:
                    continue

            product_id = getattr(product, "id", None)
            product_code = getattr(product, "product_code", "") or ""
            product_name = getattr(product, "product_name", "") or ""
            ims_name = getattr(product, "ims_name", "") or ""

            canonical_labels = {
                AliasService.normalize(product_code),
                AliasService.normalize(product_name),
                AliasService.normalize(ims_name),
            }
            canonical_labels = {
                label for label in canonical_labels if label
            }

            if canonical_labels and not any(
                label in normalized_header
                for label in canonical_labels
            ):
                continue

            product_info = products.setdefault(
                product_id,
                {
                    "product_id": product_id,
                    "product_code": product_code,
                    "product_name": product_name,
                    "ims_name": ims_name,
                    "columns": [],
                },
            )

            product_info["columns"].append(
                {
                    "index": column_index,
                    "header": str(header),
                    "metric": self.metric_for_column(header),
                }
            )
            metric_pair = (product_id, self.metric_for_column(header))
            if metric_pair in seen_metric_pairs:
                self._log_warning(
                    reason="duplicate_product_metric_column",
                    sheet_name="unknown",
                    source_row=0,
                    product=product_name,
                    header=str(header),
                )
            seen_metric_pairs.add(metric_pair)
            self.statistics["matched_products"] += 1
        return products

    def _product_label_from_metric_header(self, header):
        """Extract a new product only from a column that explicitly carries a metric."""
        normalized = AliasService.normalize(self.clean_text(header))
        dimension_tokens = (
            self.REPRESENTATIVE_HEADERS | self.REGION_HEADERS
            | self.PROVINCE_HEADERS | self.PRODUCT_GROUP_HEADERS
        )
        if not normalized or normalized in dimension_tokens:
            return None
        metric_pattern = r"(?:\bTL\b|\bCIRO\b|\bVALUE\b|\bKUTU\b|\bBOX\b|\bUNIT\b|\bADET\b)"
        if not re.search(metric_pattern, normalized):
            return None
        label = re.sub(metric_pattern, " ", normalized)
        label = re.sub(r"\b(?:SATIS|CIKIS|TOPLAM)\b", " ", label)
        label = re.sub(r"\s+", " ", label).strip(" -_/|:.")
        return label or None

    def clean_dataframe(self, dataframe, representative_column):
        result = dataframe.copy()
        representative_values = result[representative_column].fillna("").astype(str).str.strip()
        normalized_values = representative_values.map(AliasService.normalize)
        valid_mask = representative_values != ""
        valid_mask &= ~normalized_values.isin(self.TOTAL_LABELS)
        valid_mask &= ~normalized_values.isin(self.REPRESENTATIVE_HEADERS)
        valid_mask &= ~normalized_values.str.startswith(tuple(self.NOISE_ROW_TOKENS))
        result = result[valid_mask]
        duplicate_mask = normalized_values.duplicated(keep=False) & ~normalized_values.isin(self.TOTAL_LABELS)
        duplicate_count = int(duplicate_mask.sum())
        if duplicate_count:
            self._log_warning(
                reason="duplicate_representative_rows",
                sheet_name="unknown",
                source_row=0,
                count=duplicate_count,
            )
        result.reset_index(drop=True, inplace=True)
        return result

    def prepare_sheet(self, sheet):
        dataframe = sheet["dataframe"]
        dimensions = self.detect_dimension_columns(dataframe)
        representative_columns = self.detect_representative_columns(dataframe)
        representative_column = dimensions["representative"] or (
            representative_columns[0] if representative_columns else None
        )
        if representative_column is None:
            self.warnings.append(f"{sheet['sheet_name']}: temsilci kolonu bulunamadı; sayfa atlandı.")
            return None
        if representative_columns and representative_column not in representative_columns:
            representative_columns.insert(0, representative_column)
        auto_create_representatives = bool(
            dimensions["brick"]
            and any("TTS ISMI" in AliasService.normalize(column) for column in representative_columns)
        )
        self.parser_decisions.append(
            {
                "sheet_name": sheet["sheet_name"],
                "sheet_type": sheet["sheet_type"],
                "header_row": sheet["header_row"],
                "representative_column": str(representative_column),
                "representative_columns": [str(column) for column in representative_columns],
                "auto_create_representatives": auto_create_representatives,
                "mode": "normalized" if dimensions["product_group"] is not None else "wide",
            }
        )

        if dimensions["product_group"] is not None:
            metric_kind = self.detect_metric_kind(sheet["sheet_name"], dataframe)
            metric_column = self.detect_metric_column(dataframe, dimensions, metric_kind)
            if metric_column is None:
                self.warnings.append(f"{sheet['sheet_name']}: metrik kolonu bulunamadı; sayfa atlandı.")
                return None
            return {
                **sheet,
                "mode": "normalized",
                "dataframe": self.clean_dataframe(dataframe, representative_column),
                "representative_column": representative_column,
                "region_column": dimensions["region"],
                "province_column": dimensions["province"],
                "product_group_column": dimensions["product_group"],
                "metric_kind": metric_kind,
                "metric_column": metric_column,
            }

        products = self.detect_product_columns(
            dataframe,
            representative_column,
            allow_auto_create=sheet.get("sheet_type") == "brick_sales",
        )
        if not products:
            self.warnings.append(f"{sheet['sheet_name']}: eşleşen ürün kolonu bulunamadı; sayfa atlandı.")
            return None

        return {
            **sheet,
            "mode": "wide",
            "dataframe": self.clean_dataframe(dataframe, representative_column),
            "representative_column": representative_column,
            "representative_columns": representative_columns or [representative_column],
            "brick_column": dimensions["brick"],
            "region_column": dimensions["region"],
            "province_column": dimensions["province"],
            "manager_column": dimensions["manager"],
            "auto_create_representatives": auto_create_representatives,
            "products": products,
        }

    def merge_normalized_sheets(self, normalized_sheets):
        merged = {}
        for sheet in normalized_sheets:
            dataframe = sheet["dataframe"]
            representative_column = sheet["representative_column"]
            product_group_column = sheet["product_group_column"]
            for dataframe_index, (_, row) in enumerate(dataframe.iterrows()):
                representative_name = self.clean_text(row[representative_column])
                product_group_name = self.clean_text(row[product_group_column])
                source_row = dataframe_index + sheet["header_row"] + 2
                if not representative_name:
                    self.statistics["skipped_records"] += 1
                    self._log_skipped_row(
                        reason="missing_representative",
                        sheet_name=sheet["sheet_name"],
                        source_row=source_row,
                    )
                    continue
                if not product_group_name:
                    self.statistics["skipped_records"] += 1
                    self._log_skipped_row(
                        reason="missing_product_group",
                        sheet_name=sheet["sheet_name"],
                        source_row=source_row,
                        representative=representative_name,
                    )
                    continue

                key = (
                    AliasService.normalize(representative_name),
                    AliasService.normalize(product_group_name),
                )
                merged_row = merged.setdefault(
                    key,
                    {
                        "representative_name": representative_name,
                        "product_group": product_group_name,
                        "region": self.clean_text(row[sheet["region_column"]]) if sheet["region_column"] else None,
                        "province": self.clean_text(row[sheet["province_column"]]) if sheet["province_column"] else None,
                        "source_rows": [],
                        "sheet_names": set(),
                        "metrics": {"unit": 0.0, "tl": 0.0, "market_share": 0.0, "value_share": 0.0, "growth": 0.0},
                        "source_values": {},
                        "has_metric_value": False,
                        "invalid_metrics": [],
                    },
                )
                merged_row["source_rows"].append(source_row)
                merged_row["sheet_names"].add(sheet["sheet_name"])
                metric_value = row[sheet["metric_column"]]
                parsed_value, metric_state = self.parse_metric_value(metric_value)
                if metric_state == "valid":
                    merged_row["has_metric_value"] = True
                    merged_row["metrics"][sheet["metric_kind"]] += parsed_value
                elif metric_state == "invalid":
                    merged_row["invalid_metrics"].append(
                        {
                            "field": sheet["metric_column"],
                            "sheet_name": sheet["sheet_name"],
                            "source_row": source_row,
                            "value": self._value_for_json(metric_value),
                        }
                    )
                merged_row["source_values"][f"{sheet['sheet_name']}::{sheet['metric_column']}"] = self._value_for_json(
                    metric_value
                )
        return list(merged.values())

    def stage_normalized_raw_data(self, normalized_rows, year, month, week_number=None):
        for item in normalized_rows:
            self.statistics["source_metric_records"] += 1
            representative_name = item["representative_name"]
            try:
                region_value = self.clean_text(item.get("region"))
                province_value = self.clean_text(item.get("province"))
                parsed_region, parsed_city = self._region_context(region_value, province_value)
                representative_match = self.resolve_representative_match(representative_name)
                if representative_match["matched"]:
                    representative_id = representative_match["object"].id
                    self.statistics["matched_representatives"] += 1
                elif parsed_region or province_value:
                    representative_id, _ = self._ensure_representative(
                        representative_name,
                        region=parsed_region or region_value or None,
                        city=province_value or parsed_city,
                    )
                    representative_match = self.resolve_representative_match(representative_name)
                    self.warnings.append(
                        f"Yeni temsilci bölge bağlamıyla otomatik oluşturuldu ({representative_name})."
                    )
                else:
                    self.statistics["unresolved_representative_rows"] += 1
                    self.statistics["unmatched_representatives"] += 1
                    self.statistics["queued_for_manual"] += 1
                    self._log_skipped_row(
                        reason="unmatched_representative_without_region_context",
                        sheet_name=" | ".join(sorted(item["sheet_names"])),
                        source_row=min(item["source_rows"]),
                        representative=representative_name,
                    )
                    representative_id = None

                product_group_name = item["product_group"]
                product, _ = self._ensure_product(product_group_name)
                if product is None:
                    self.statistics["skipped_records"] += 1
                    self.statistics["unmatched_product_records"] += 1
                    continue
                product_match = {"matched": True, "object": product}

                if region_value:
                    region_suggestion = AliasService.suggest_region(region_value)
                    if not region_suggestion["matched"]:
                        self.statistics["unmatched_regions"] += 1
                        self.statistics["queued_for_manual"] += 1
                        AliasService.enqueue_unmatched_region(
                            source_value=region_value,
                            import_id=self.upload.id,
                            worksheet=" | ".join(sorted(item["sheet_names"])),
                            row_number=min(item["source_rows"]),
                            suggested_match=region_suggestion.get("value"),
                            confidence_score=region_suggestion.get("score", 0.0),
                            reason="unmatched_region",
                        )
                    elif representative_match["matched"]:
                        canonical_region = self.clean_text(representative_match["object"].region)
                        if canonical_region and AliasService.normalize(canonical_region) != AliasService.normalize(
                            region_value
                        ):
                            self._log_warning(
                                reason="inconsistent_region",
                                sheet_name=" | ".join(sorted(item["sheet_names"])),
                                source_row=min(item["source_rows"]),
                                representative=representative_name,
                                region=region_value,
                                expected_region=canonical_region,
                            )

                province_value = self.clean_text(item.get("province"))
                if province_value:
                    province_suggestion = AliasService.suggest_province(province_value)
                    if not province_suggestion["matched"]:
                        self.statistics["unmatched_provinces"] += 1
                        self.statistics["queued_for_manual"] += 1
                        AliasService.enqueue_unmatched_province(
                            source_value=province_value,
                            import_id=self.upload.id,
                            worksheet=" | ".join(sorted(item["sheet_names"])),
                            row_number=min(item["source_rows"]),
                            suggested_match=province_suggestion.get("value"),
                            confidence_score=province_suggestion.get("score", 0.0),
                            reason="unmatched_province",
                        )
                    elif representative_match["matched"]:
                        canonical_province = self.clean_text(representative_match["object"].city)
                        if canonical_province and AliasService.normalize(
                            canonical_province
                        ) != AliasService.normalize(province_value):
                            self._log_warning(
                                reason="inconsistent_province",
                                sheet_name=" | ".join(sorted(item["sheet_names"])),
                                source_row=min(item["source_rows"]),
                                representative=representative_name,
                                province=province_value,
                                expected_province=canonical_province,
                            )

                self.statistics["matched_products"] += 1
                metrics = item["metrics"]
                if item["invalid_metrics"]:
                    self.statistics["skipped_records"] += 1
                    self.statistics["invalid_metric_records"] += 1
                    for invalid_metric in item["invalid_metrics"]:
                        self._log_skipped_row(
                            reason="invalid_numeric_value",
                            representative=representative_name,
                            product=product_group_name,
                            **invalid_metric,
                        )
                    continue
                if not item["has_metric_value"]:
                    self.statistics["skipped_records"] += 1
                    self.statistics["blank_metric_records"] += 1
                    self._log_skipped_row(
                        reason="empty_metrics",
                        sheet_name=" | ".join(sorted(item["sheet_names"])),
                        source_row=min(item["source_rows"]),
                        representative=representative_name,
                        product_group=product_group_name,
                    )
                    continue

                source_values = {
                    **item["source_values"],
                    "region": item["region"],
                    "province": item["province"],
                    "product_group": product_group_name,
                    "sheet_names": sorted(item["sheet_names"]),
                }

                self.create_raw_record(
                    year=year,
                    month=month,
                    week_number=week_number,
                    sheet_name=" | ".join(sorted(item["sheet_names"])),
                    sheet_type="brick_normalized",
                    source_row=min(item["source_rows"]),
                    representative_name=representative_name,
                    representative_id=representative_id,
                    product=product_match["object"],
                    metrics=metrics,
                    source_values=source_values,
                )
                self.statistics["stored_source_records"] += 1
                if not any(metrics.values()):
                    self.statistics["zero_metric_records"] += 1
                self.statistics["processed_rows"] += 1
            except Exception as exc:
                self.statistics["rows_error"] += 1
                self._log_skipped_row(
                    reason="row_processing_error",
                    sheet_name=" | ".join(sorted(item.get("sheet_names", []))),
                    source_row=min(item.get("source_rows", [0])),
                    error=str(exc),
                )
                continue

    def create_raw_record(
        self,
        *,
        year,
        month,
        week_number=None,
        sheet_name,
        sheet_type,
        source_row,
        representative_name,
        representative_id,
        product,
        product_id=None,
        metrics,
        source_values,
        manager=None,
        territory=None,
        brick=None,
        province=None,
        market=None,
        competitor=None,
    ):
        resolved_product_id = product_id if product_id is not None else getattr(product, "id", None)
        if resolved_product_id is None:
            raise ValueError("Raw IMS record requires a resolved product_id.")
        product_name = getattr(product, "product_name", product)
        self._raw_batch.append({
            "upload_id": self.upload.id,
            "year": year,
            "month": month,
            "week_number": week_number,
            "quarter": self.quarter_for(month),
            "sheet_name": sheet_name,
            "sheet_type": sheet_type,
            "source_row": source_row,
            "representative_id": representative_id,
            "product_id": resolved_product_id,
            "representative": representative_name,
            "manager": manager,
            "territory": territory,
            "brick": brick,
            "province": province,
            "product": product_name,
            "competitor": competitor,
            "market": market,
            "unit": metrics["unit"],
            "tl": metrics["tl"],
            "market_share": metrics["market_share"],
            "value_share": metrics["value_share"],
            "growth": metrics["growth"],
            "raw_json": self._json_dump({
                "representative": representative_name,
                "product": product_name,
                "metrics": metrics,
                "source_values": source_values,
            }),
            "created_at": datetime.utcnow(),
        })
        self._pending_raw_records += 1
        if self._pending_raw_records >= self.WRITE_BATCH_SIZE:
            self._flush_raw_batch()
        self.statistics["raw_records"] += 1
        return None

    def _flush_raw_batch(self):
        """Write RAW rows in bounded batches without retaining ORM instances."""
        if not self._raw_batch:
            return
        db.session.bulk_insert_mappings(IMSRawData, self._raw_batch)
        self._raw_batch.clear()
        self._pending_raw_records = 0

    def stage_raw_data(self, prepared_sheets, year, month, week_number=None):
        for sheet in prepared_sheets:
            if sheet.get("mode") == "normalized":
                continue
            dataframe = sheet["dataframe"]
            representative_columns = sheet.get("representative_columns") or [sheet["representative_column"]]
            representative_column = representative_columns[0]
            brick_column = sheet.get("brick_column")
            region_column = sheet.get("region_column")
            province_column = sheet.get("province_column")
            manager_column = sheet.get("manager_column")
            auto_create_representatives = bool(sheet.get("auto_create_representatives"))

            for dataframe_index, (_, row) in enumerate(dataframe.iterrows()):
                try:
                    representative_values = []
                    primary_representative = self.clean_text(row[representative_column])
                    if primary_representative:
                        representative_values.append(primary_representative)
                    else:
                        for candidate_column in representative_columns[1:]:
                            value = self.clean_text(row[candidate_column])
                            if value:
                                representative_values.append(value)
                                break
                    representative_values = list(dict.fromkeys(representative_values))
                    if not representative_values:
                        self.statistics["skipped_records"] += 1
                        self._log_skipped_row(
                            reason="missing_representative",
                            sheet_name=sheet["sheet_name"],
                            source_row=dataframe_index + sheet["header_row"] + 2,
                        )
                        continue

                    territory_value = self.clean_text(row[brick_column]) if brick_column else None
                    region_value = self.clean_text(row[region_column]) if region_column else None
                    province_value = self.clean_text(row[province_column]) if province_column else None
                    parsed_region, parsed_city = self._region_context(region_value, province_value)
                    manager_value = self.clean_text(row[manager_column]) if manager_column else None
                    source_row = dataframe_index + sheet["header_row"] + 2

                    for representative_name in representative_values:
                        if self._is_vacancy_representative(representative_name):
                            representative_id = self._ensure_vacancy_representative(region_value, province_value, representative_name)
                            existed = True
                        elif not self._is_probable_representative_name(representative_name):
                            if self._is_aggregate_label(representative_name):
                                self.statistics["aggregate_rows_excluded"] += 1
                                self._log_excluded_row(
                                    reason="aggregate_representative",
                                    sheet_name=sheet["sheet_name"],
                                    source_row=source_row,
                                    representative=representative_name,
                                )
                            else:
                                self.statistics["skipped_records"] += 1
                                self.statistics["unresolved_representative_rows"] += 1
                                self._log_skipped_row(
                                    reason="invalid_representative",
                                    sheet_name=sheet["sheet_name"],
                                    source_row=source_row,
                                    representative=representative_name,
                                )
                            continue
                        else:
                            representative_id, existed = self._resolve_representative(
                                representative_name=representative_name,
                                allow_auto_create=bool(auto_create_representatives or parsed_region or province_value),
                                sheet_name=sheet["sheet_name"],
                                source_row=source_row,
                                territory=territory_value,
                                manager=manager_value,
                                region=parsed_region or region_value or None,
                                city=province_value or parsed_city,
                            )
                        if representative_id is None:
                            self.statistics["unresolved_representative_rows"] += 1
                            continue
                        if not existed:
                            self.warnings.append(
                                f"{sheet['sheet_name']} satır {source_row}: "
                                f"yeni temsilci oluşturuldu ({representative_name})."
                            )

                        for product_info in sheet["products"].values():
                            self.statistics["source_metric_records"] += 1
                            metrics = {
                                "unit": 0.0,
                                "tl": 0.0,
                                "market_share": 0.0,
                                "value_share": 0.0,
                                "growth": 0.0,
                            }
                            source_values = {}
                            has_metric_value = False
                            for column in product_info["columns"]:
                                value = row.iloc[column["index"]]
                                parsed_value, metric_state = self.parse_metric_value(value)
                                if metric_state == "invalid":
                                    self.statistics["skipped_records"] += 1
                                    self.statistics["invalid_metric_records"] += 1
                                    self._log_skipped_row(
                                        reason="invalid_numeric_value",
                                        sheet_name=sheet["sheet_name"],
                                        source_row=source_row,
                                        representative=representative_name,
                                        product=product_info["product_name"],
                                        field=column["header"],
                                        value=self._value_for_json(value),
                                    )
                                    metrics = None
                                    break
                                if metric_state == "valid":
                                    has_metric_value = True
                                    metrics[column["metric"]] += parsed_value
                                source_values[column["header"]] = self._value_for_json(value)
                            if metrics is None:
                                continue

                            if not has_metric_value:
                                self.statistics["skipped_records"] += 1
                                self.statistics["blank_metric_records"] += 1
                                self._log_skipped_row(
                                    reason="empty_metrics",
                                    sheet_name=sheet["sheet_name"],
                                    source_row=source_row,
                                    representative=representative_name,
                                    product=product_info["product_name"],
                                )
                                continue

                            self.create_raw_record(
                                year=year,
                                month=month,
                                week_number=week_number,
                                sheet_name=sheet["sheet_name"],
                                sheet_type=sheet["sheet_type"],
                                source_row=source_row,
                                representative_name=representative_name,
                                representative_id=representative_id,
                                product=product_info["product_name"],
                                product_id=product_info["product_id"],
                                metrics=metrics,
                                source_values=source_values,
                                manager=manager_value,
                                territory=territory_value,
                                brick=territory_value,
                            )
                            self.statistics["stored_source_records"] += 1
                            if not any(metrics.values()):
                                self.statistics["zero_metric_records"] += 1
                    self.statistics["processed_rows"] += 1
                except Exception as exc:
                    self.statistics["rows_error"] += 1
                    self._log_skipped_row(
                        reason="row_processing_error",
                        sheet_name=sheet["sheet_name"],
                        source_row=dataframe_index + sheet["header_row"] + 2,
                        error=str(exc),
                    )
                    continue
            self.statistics["processed_sheets"] += 1

        self._flush_raw_batch()

    def transform_raw_to_facts(self, year, month, week_number=None):
        """UPSERT IMS facts: update existing rows for the same week/period, insert new ones."""
        raw_records = IMSRawData.query.filter_by(upload_id=self.upload.id, year=year, month=month).all()
        
        # Always populate existing_map to protect against intra-file duplicates even in monthly loads
        query = IMSFact.query.filter_by(year=year, month=month).filter(IMSFact.report_type.isnot(None))
        if week_number is not None:
            query = query.filter_by(week_number=week_number)
        else:
            query = query.filter_by(week_number=None)
            
        existing_facts = query.all()
        existing_map = {
            (fact.representative_id, fact.product_id, fact.report_type): fact
            for fact in existing_facts
        }
        
        # A workbook contains one RAW row for every brick.  Facts are scoped
        # to representative/product/report type, therefore writing each RAW
        # row directly caused the final brick encountered to overwrite all
        # previous bricks.  Aggregate first, retaining the latest raw id only
        # as the traceability pointer.
        aggregates = {}
        for raw in raw_records:
            if raw.representative_id is None or raw.product_id is None:
                self.statistics["skipped_records"] += 1
                continue
            key = (raw.representative_id, raw.product_id, raw.sheet_type)
            aggregate = aggregates.setdefault(key, {
                "raw": raw,
                "unit": 0.0,
                "tl": 0.0,
                "market_share_total": 0.0,
                "value_share_total": 0.0,
                "growth_total": 0.0,
                "count": 0,
            })
            aggregate["raw"] = raw
            aggregate["unit"] += raw.unit or 0.0
            aggregate["tl"] += raw.tl or 0.0
            aggregate["market_share_total"] += raw.market_share or 0.0
            aggregate["value_share_total"] += raw.value_share or 0.0
            aggregate["growth_total"] += raw.growth or 0.0
            aggregate["count"] += 1

        for key, aggregate in aggregates.items():
            raw = aggregate["raw"]
            existing = existing_map.get(key)
            count = aggregate["count"]
            unit = aggregate["unit"]
            tl = aggregate["tl"]
            market_share = aggregate["market_share_total"] / count
            value_share = aggregate["value_share_total"] / count
            growth = aggregate["growth_total"] / count

            if existing:
                existing.upload_id = self.upload.id
                existing.raw_data_id = raw.id
                existing.unit = unit
                existing.tl = tl
                existing.market_share = market_share
                existing.value_share = value_share
                existing.growth = growth
                existing.metrics_json = raw.raw_json
                self.statistics["facts_updated"] += 1
            else:
                fact = IMSFact(
                    upload_id=self.upload.id,
                    raw_data_id=raw.id,
                    representative_id=raw.representative_id,
                    product_id=raw.product_id,
                    year=raw.year,
                    month=raw.month,
                    week_number=week_number,
                    quarter=raw.quarter,
                    report_type=raw.sheet_type,
                    unit=unit,
                    tl=tl,
                    market_share=market_share,
                    value_share=value_share,
                    growth=growth,
                    metrics_json=raw.raw_json,
                )
                db.session.add(fact)
                existing_map[key] = fact
                self.statistics["facts_inserted"] += 1

            self.statistics["fact_records"] += 1
        db.session.flush()

    def rebuild_summary(self, year, month):
        IMSSummary.query.filter_by(year=year, month=month).delete(synchronize_session=False)

        rows = (
            db.session.query(
                IMSFact.representative_id,
                IMSFact.product_id,
                func.sum(IMSFact.unit).label("unit"),
                func.sum(IMSFact.tl).label("tl"),
                func.avg(IMSFact.market_share).label("market_share"),
                func.avg(IMSFact.value_share).label("value_share"),
                func.avg(IMSFact.growth).label("growth"),
            )
            .filter(IMSFact.year == year, IMSFact.month == month)
            .group_by(IMSFact.representative_id, IMSFact.product_id)
            .all()
        )

        quarter = self.quarter_for(month)
        targets = Target.query.filter_by(year=year, month=month).all()
        target_map = {(target.representative_id, target.product_id): target for target in targets}
        
        summaries_to_insert = []
        for row in rows:
            target = target_map.get((row.representative_id, row.product_id))
            target_unit = target.unit_target if target else 0.0
            target_tl = target.tl_target if target else 0.0
            realization_base = target_tl or target_unit
            realization_actual = row.tl if target_tl else row.unit
            realization_percent = (
                round(realization_actual * 100 / realization_base, 2) if realization_base else 0.0
            )
            summaries_to_insert.append(
                IMSSummary(
                    upload_id=self.upload.id,
                    representative_id=row.representative_id,
                    product_id=row.product_id,
                    year=year,
                    month=month,
                    quarter=quarter,
                    unit=row.unit or 0.0,
                    tl=row.tl or 0.0,
                    market_share=row.market_share or 0.0,
                    value_share=row.value_share or 0.0,
                    growth=row.growth or 0.0,
                    target_unit=target_unit,
                    target_tl=target_tl,
                    realization_percent=realization_percent,
                )
            )
            
        if summaries_to_insert:
            db.session.bulk_save_objects(summaries_to_insert)

        self.statistics["summary_records"] = len(rows)
        db.session.flush()

    def apply_balance_summary(self, year, month):
        """Load BAKİYE targets and only use its actuals when TTS is absent.

        ``TTS HAFTALIK ÇIKIŞLARI`` is the canonical representative-level
        actual-sales source.  BAKİYE can contain a different reconciliation
        scope, so allowing it to overwrite TTS creates product-level drift.
        Its target columns remain the authoritative target source; its actual
        columns are retained solely as a compatibility fallback for workbooks
        which do not include a TTS weekly sales report.
        """
        sheet_name = next((name for name in self.workbook if "BAKIYE" in AliasService.normalize(name)), None)
        if not sheet_name:
            return
        frame = self.workbook[sheet_name]
        header_row = next((i for i in range(min(12, len(frame))) if "HEDEF" in " ".join(AliasService.normalize(v) for v in frame.iloc[i]) and "CIKIS" in " ".join(AliasService.normalize(v) for v in frame.iloc[i])), None)
        if header_row is None or header_row + 1 >= len(frame):
            self.warnings.append(f"{sheet_name}: hedef/çıkış başlığı bulunamadı.")
            return
        sections, current = {}, ""
        for column in range(frame.shape[1]):
            label = self.clean_text(frame.iloc[header_row, column])
            normalized_label = AliasService.normalize(label)
            if "HEDEF" in normalized_label:
                current = "target_tl"
            elif "CIKIS" in normalized_label:
                current = "actual_tl"
            elif "BAKIYE" in normalized_label:
                current = "balance_unit" if "KUTU" in normalized_label else "balance_tl"
            sections[column] = current
        has_weekly_sales = any(
            "HAFTALIK" in AliasService.normalize(name) and "CIKIS" in AliasService.normalize(name)
            for name in self.workbook
        )
        targets = {(t.representative_id, t.product_id): t for t in Target.query.filter_by(year=year, month=month).all()}
        summaries = {(s.representative_id, s.product_id): s for s in IMSSummary.query.filter_by(year=year, month=month).all()}
        products_by_id = {product.id: product for product in Product.query.all()}
        for _, row in frame.iloc[header_row + 1:].iterrows():
            rep_name = self.clean_text(row.iloc[1])
            # The first column is the Excel hierarchy label (e.g.
            # ``101 ISTANBUL``).  Preserve its code for ordering and its city
            # name for display, without replacing the representative's brick.
            location = self.clean_text(row.iloc[0])
            location_match = re.match(r"^(\d{3})\s+(.+)$", location)
            if self._is_vacancy_representative(rep_name):
                rep_id = self._ensure_vacancy_representative(location, vacancy_name=rep_name)
                representative = Representative.query.get(rep_id)
            else:
                if not self._is_probable_representative_name(rep_name):
                    continue
                rep_match = self.resolve_representative_match(rep_name)
                if not rep_match["matched"]:
                    continue
                rep_id = rep_match["object"].id
                representative = rep_match["object"]
            if location_match:
                representative.region = location_match.group(1)
                representative.city = location_match.group(2).strip()
            values = {}
            for column in range(frame.shape[1]):
                product_match = self.resolve_product_match(self.clean_text(frame.iloc[header_row, column]))
                if not product_match["matched"]:
                    continue
                product_id = product_match["object"].id
                metric = self.safe_float(row.iloc[column])
                section = sections[column]
                if section == "target_tl":
                    values.setdefault(product_id, {})["target"] = metric
                elif section == "actual_tl":
                    values.setdefault(product_id, {})["actual"] = metric
                elif section == "balance_tl":
                    values.setdefault(product_id, {})["balance_tl"] = metric
                elif section == "balance_unit":
                    values.setdefault(product_id, {})["balance_unit"] = metric
            for product_id, item in values.items():
                if "target" not in item and "actual" not in item:
                    continue
                target = targets.get((rep_id, product_id))
                if target is None:
                    target = Target(year=year, month=month, quarter=self.quarter_for(month), representative_id=rep_id, product_id=product_id)
                    db.session.add(target); targets[(rep_id, product_id)] = target
                target.tl_target = item.get("target", target.tl_target or 0.0)
                balance_tl = item.get("balance_tl")
                balance_unit = item.get("balance_unit")
                if balance_tl is not None and balance_unit not in (None, 0):
                    # BAKİYE's MF'siz kutu block carries the representative /
                    # product-specific net unit factor.  It is the source
                    # used by the workbook's own remaining-box calculation.
                    net_unit_price = balance_tl / balance_unit
                    if net_unit_price > 0:
                        target.unit_target = float(round(target.tl_target / net_unit_price))
                elif not target.unit_target:
                    target.unit_target = TargetBoxCalculationService.unit_target(
                        target.tl_target,
                        products_by_id.get(product_id).unit_price if products_by_id.get(product_id) else 0,
                    )
                summary = summaries.get((rep_id, product_id))
                if summary is not None:
                    summary.target_tl = target.tl_target
                    summary.target_unit = target.unit_target
                if not has_weekly_sales and "actual" in item:
                    target.tl_realization = item["actual"]
                    target.realization_percent = round(target.tl_realization * 100 / target.tl_target, 2) if target.tl_target else 0.0
                    if summary is not None:
                        summary.tl = target.tl_realization
                        summary.realization_percent = target.realization_percent
        db.session.flush()

    def apply_weekly_sales_summary(self, year, month):
        """Apply TTS weekly TL/kutu values as the period actuals.

        The sheet has two adjacent product blocks (TL and KUTU) and does not
        require a fixed month name.  We discover the product header row and
        forward-fill its block labels, then update the same representative /
        product records used by the performance, target and prime screens.
        """
        sheet_name = next(
            (
                name for name in self.workbook
                if "HAFTALIK" in AliasService.normalize(name)
                and "CIKIS" in AliasService.normalize(name)
            ),
            None,
        )
        if not sheet_name:
            return {"rows": 0, "matched_representatives": 0, "updated_values": 0}

        frame = self.workbook[sheet_name]
        header_row = next(
            (
                index for index in range(min(12, len(frame)))
                if "TRAVAZOL" in " ".join(AliasService.normalize(value) for value in frame.iloc[index])
                and "MONUROL" in " ".join(AliasService.normalize(value) for value in frame.iloc[index])
            ),
            None,
        )
        if header_row is None or header_row == 0:
            self.warnings.append(f"{sheet_name}: TTS ürün başlığı bulunamadı.")
            return {"rows": 0, "matched_representatives": 0, "updated_values": 0}

        # A TTS sheet may also append a single-week block after the cumulative
        # period block.  The representative screen is MTD, so retain the first
        # TL and first KUTU ÇIKIŞI blocks (for example ``1-14 HAZİRAN``) only.
        sections, current_section, selected_metrics = {}, "", set()
        for column in range(frame.shape[1]):
            label = AliasService.normalize(self.clean_text(frame.iloc[header_row - 1, column]))
            if "TL" in label and "CIKIS" in label:
                current_section = "tl" if "tl" not in selected_metrics else ""
                selected_metrics.add("tl")
            elif ("KUTU" in label or "UNIT" in label) and "CIKIS" in label:
                current_section = "unit" if "unit" not in selected_metrics else ""
                selected_metrics.add("unit")
            sections[column] = current_section

        product_columns = {}
        for column in range(frame.shape[1]):
            if sections.get(column) not in {"tl", "unit"}:
                continue
            product_match = self.resolve_product_match(self.clean_text(frame.iloc[header_row, column]))
            if product_match["matched"]:
                product_columns[column] = (product_match["object"].id, sections[column])

        targets = {(item.representative_id, item.product_id): item for item in Target.query.filter_by(year=year, month=month).all()}
        summaries = {(item.representative_id, item.product_id): item for item in IMSSummary.query.filter_by(year=year, month=month).all()}
        matched_representatives = updated_values = rows = 0
        for _, row in frame.iloc[header_row + 1:].iterrows():
            rep_name = self.clean_text(row.iloc[1])
            if self._is_vacancy_representative(rep_name):
                rep_id = self._ensure_vacancy_representative(self.clean_text(row.iloc[0]), vacancy_name=rep_name)
            elif self._is_probable_representative_name(rep_name):
                rep_match = self.resolve_representative_match(rep_name)
                if not rep_match["matched"]:
                    continue
                rep_id = rep_match["object"].id
            else:
                continue

            rows += 1
            matched_representatives += 1
            values = {}
            for column, (product_id, metric) in product_columns.items():
                values.setdefault(product_id, {})[metric] = self.safe_float(row.iloc[column])
            for product_id, metrics in values.items():
                summary = summaries.get((rep_id, product_id))
                target = targets.get((rep_id, product_id))
                if summary is None:
                    continue
                if "tl" in metrics:
                    summary.tl = metrics["tl"]
                    if target is not None:
                        target.tl_realization = metrics["tl"]
                # Preserve the explicit cumulative KUTU ÇIKIŞI source value.
                # Never synthesize box actuals from TL / target ratios.
                if "unit" in metrics:
                    summary.unit = metrics["unit"]
                    if target is not None:
                        target.unit_realization = metrics["unit"]
                if target is not None:
                    target.realization_percent = round(summary.tl * 100 / target.tl_target, 2) if target.tl_target else 0.0
                    summary.target_tl = target.tl_target
                    summary.target_unit = target.unit_target
                    summary.realization_percent = target.realization_percent
                updated_values += 1
        db.session.flush()
        return {
            "rows": rows,
            "matched_representatives": matched_representatives,
            "updated_values": updated_values,
        }

    def persist_national_dashboard_metrics(self, year, month):
        """Persist source-authoritative National and region subtotal KPI rows.

        Person rows remain untouched for representative reporting. Company and
        region KPI totals use the workbook's own subtotal rows so a workbook
        whose person allocations do not reconcile cannot inflate executive or
        region totals.
        """
        if not self.upload or not self.workbook:
            return

        def upsert(sheet_name, sheet_type, product_id, unit, tl, metadata, representative="NATIONAL", territory=None):
            record = IMSRawData.query.filter_by(
                upload_id=self.upload.id,
                sheet_type=sheet_type,
                product_id=product_id,
                representative=representative,
                territory=territory,
            ).first()
            values = dict(
                year=year, month=month, quarter=self.quarter_for(month),
                week_number=self.upload.week_number, sheet_name=sheet_name,
                sheet_type=sheet_type, source_row=0, product_id=product_id,
                representative=representative, territory=territory,
                unit=float(unit or 0), tl=float(tl or 0),
                raw_json=json.dumps(metadata, ensure_ascii=False),
            )
            if record:
                for key, value in values.items():
                    setattr(record, key, value)
            else:
                db.session.add(IMSRawData(upload_id=self.upload.id, **values))

        def is_region_subtotal(row):
            if len(row) < 2:
                return False
            territory = self.clean_text(row.iloc[0])
            representative = self.clean_text(row.iloc[1])
            if not territory or not representative:
                return False
            return (
                AliasService.normalize(territory) == AliasService.normalize(representative)
                and bool(re.match(r"^\d{3}\b", AliasService.normalize(territory)))
            )

        balance_name = next((name for name in self.workbook if "BAKIYE" in AliasService.normalize(name)), None)
        if balance_name:
            frame = self.workbook[balance_name]
            header_row = next((
                i for i in range(min(12, len(frame)))
                if "HEDEF" in " ".join(AliasService.normalize(v) for v in frame.iloc[i])
                and "CIKIS" in " ".join(AliasService.normalize(v) for v in frame.iloc[i])
            ), None)
            if header_row is not None:
                sections, current = {}, ""
                for column in range(frame.shape[1]):
                    label = AliasService.normalize(self.clean_text(frame.iloc[header_row, column]))
                    if any(token in label for token in ("HEDEF", "CIKIS", "BAKIYE")):
                        current = label
                    sections[column] = current

                def balance_values(row):
                    metric_values = {}
                    for column, section in sections.items():
                        product_match = self.resolve_product_match(self.clean_text(frame.iloc[header_row, column]))
                        if not product_match["matched"]:
                            continue
                        product_id = product_match["object"].id
                        values = metric_values.setdefault(product_id, {"target_tl": 0.0, "actual_tl": 0.0})
                        if "HEDEF" in section:
                            values["target_tl"] = self.safe_float(row.iloc[column])
                        elif "CIKIS" in section:
                            values["actual_tl"] = self.safe_float(row.iloc[column])
                    return metric_values

                for row_index in range(header_row + 1, len(frame)):
                    row = frame.iloc[row_index]
                    territory = self.clean_text(row.iloc[0]) if frame.shape[1] > 0 else ""
                    representative = self.clean_text(row.iloc[1]) if frame.shape[1] > 1 else ""
                    normalized_rep = AliasService.normalize(representative)
                    if normalized_rep == "NATIONAL":
                        for product_id, values in balance_values(row).items():
                            upsert(balance_name, "dashboard_balance_national", product_id,
                                   values["target_tl"], values["actual_tl"], values)
                    elif is_region_subtotal(row):
                        for product_id, values in balance_values(row).items():
                            upsert(balance_name, "dashboard_balance_region", product_id,
                                   values["target_tl"], values["actual_tl"], values,
                                   representative=representative, territory=territory)

        weekly_name = next((name for name in self.workbook if "HAFTALIK" in AliasService.normalize(name) and "CIKIS" in AliasService.normalize(name)), None)
        if weekly_name:
            frame = self.workbook[weekly_name]
            if len(frame) >= 3:
                sections, current = {}, ""
                for column in range(frame.shape[1]):
                    label = AliasService.normalize(self.clean_text(frame.iloc[0, column]))
                    if label:
                        current = label
                    sections[column] = current
                selected_tl = next((s for s in sections.values() if "CIKIS" in s and "TL" in s), "")
                selected_unit = next((s for s in sections.values() if "CIKIS" in s and "KUTU" in s), "")

                def weekly_values(row):
                    values = {}
                    for column, section in sections.items():
                        if section not in {selected_tl, selected_unit}:
                            continue
                        product_match = self.resolve_product_match(self.clean_text(frame.iloc[1, column]))
                        if not product_match["matched"]:
                            continue
                        bucket = values.setdefault(product_match["object"].id, {"actual_tl": 0.0, "actual_unit": 0.0})
                        if section == selected_tl:
                            bucket["actual_tl"] = self.safe_float(row.iloc[column])
                        elif section == selected_unit:
                            bucket["actual_unit"] = self.safe_float(row.iloc[column])
                    return values

                for row_index in range(2, len(frame)):
                    row = frame.iloc[row_index]
                    territory = self.clean_text(row.iloc[0]) if frame.shape[1] > 0 else ""
                    representative = self.clean_text(row.iloc[1]) if frame.shape[1] > 1 else ""
                    normalized_rep = AliasService.normalize(representative)
                    if normalized_rep == "NATIONAL":
                        for product_id, values in weekly_values(row).items():
                            upsert(weekly_name, "dashboard_weekly_units", product_id,
                                   values["actual_unit"], values["actual_tl"], values)
                    elif is_region_subtotal(row):
                        for product_id, values in weekly_values(row).items():
                            upsert(weekly_name, "dashboard_weekly_region", product_id,
                                   values["actual_unit"], values["actual_tl"], values,
                                   representative=representative, territory=territory)
        from app.services.official_aggregate_service import persist_official_aggregates
        persist_official_aggregates(self, year, month)
        db.session.flush()

    def clear_week(self, year, week_number):
        """Remove all IMS data for a specific week (used only as a fallback)."""
        IMSFact.query.filter_by(year=year, week_number=week_number).delete(synchronize_session=False)
        IMSRawData.query.filter_by(year=year, week_number=week_number).delete(synchronize_session=False)
        db.session.flush()

    def clear_month(self, year, month):
        """Remove all IMS data for a calendar month (destructive; kept for backward compat)."""
        IMSFact.query.filter_by(year=year, month=month).delete(synchronize_session=False)
        IMSRawData.query.filter_by(year=year, month=month).delete(synchronize_session=False)
        IMSSummary.query.filter_by(year=year, month=month).delete(synchronize_session=False)
        db.session.flush()

    def _upsert_auto_brick_assignment(self, representative_id, year, month, brick, territory=None, city=None):
        """Add one membership without replacing a manually maintained member row."""
        assignment = RepresentativeBrickAssignment.query.filter_by(
            representative_id=representative_id, year=year, month=month, brick=brick
        ).first()
        if assignment is None:
            db.session.add(RepresentativeBrickAssignment(
                representative_id=representative_id, year=year, month=month,
                quarter=self.quarter_for(month), brick=brick, territory=territory,
                city=city, source="AUTO",
            ))
            return
        if assignment.source != "MANUAL":
            if territory and not assignment.territory:
                assignment.territory = territory
            if city and not assignment.city:
                assignment.city = city

    def sync_brick_assignments(self, year, month, prepared_sheets=None):
        """Create AUTO memberships from the upload, retaining valid shared bricks."""
        rows = (
            db.session.query(IMSRawData.representative_id, IMSRawData.brick, IMSRawData.territory, IMSRawData.province)
            .filter_by(upload_id=self.upload.id, year=year, month=month)
            .filter(IMSRawData.representative_id.isnot(None), IMSRawData.brick.isnot(None))
            .distinct()
            .all()
        )
        for representative_id, brick, territory, city in rows:
            self._upsert_auto_brick_assignment(representative_id, year, month, brick, territory, city)

        # 1. and 2. TTS columns sometimes list two people working the same
        # brick.  Sales remain attributed to the primary row to avoid double
        # counting, while both distinct master IDs are stored as members.
        for sheet in prepared_sheets or []:
            if sheet.get("sheet_type") != "brick_sales" or not sheet.get("brick_column"):
                continue
            rep_columns = sheet.get("representative_columns") or [sheet["representative_column"]]
            for _, row in sheet["dataframe"].iterrows():
                brick = self.clean_text(row[sheet["brick_column"]])
                if not brick:
                    continue
                for column in rep_columns:
                    rep_name = self.clean_text(row[column])
                    if not self._is_probable_representative_name(rep_name):
                        continue
                    match = self.resolve_representative_match(rep_name)
                    if match["matched"]:
                        self._upsert_auto_brick_assignment(match["object"].id, year, month, brick)

    def backfill_brick_assignments_from_workbook(self, year, month):
        """Populate assignments from a legacy upload without duplicating RAW data."""
        if not self.workbook:
            self.load_workbook()
        for source in self.analyze_workbook():
            sheet = self.prepare_sheet(source)
            if not sheet or sheet.get("sheet_type") != "brick_sales" or not sheet.get("brick_column"):
                continue
            for _, row in sheet["dataframe"].iterrows():
                brick = self.clean_text(row[sheet["brick_column"]])
                if not brick:
                    continue
                for column in sheet.get("representative_columns") or [sheet["representative_column"]]:
                    rep_name = self.clean_text(row[column])
                    if not self._is_probable_representative_name(rep_name):
                        continue
                    match = AliasService.find_representative(rep_name)
                    if match["matched"]:
                        self._upsert_auto_brick_assignment(match["object"].id, year, month, brick)

    def process_workbook(self, year, month, week_number=None):
        sheets = self.analyze_workbook()
        prepared_sheets = [sheet for sheet in (self.prepare_sheet(item) for item in sheets) if sheet]
        self.bootstrap_vacancy_representatives_from_balance()
        normalized_sheets = [sheet for sheet in prepared_sheets if sheet.get("mode") == "normalized"]
        wide_sheets = [sheet for sheet in prepared_sheets if sheet.get("mode") != "normalized"]
        if normalized_sheets:
            self.statistics["processed_sheets"] += len(normalized_sheets)
            normalized_rows = self.merge_normalized_sheets(normalized_sheets)
            self.stage_normalized_raw_data(normalized_rows, year, month, week_number=week_number)
        self.stage_raw_data(wide_sheets, year, month, week_number=week_number)
        self._flush_raw_batch()
        self.sync_brick_assignments(year, month, prepared_sheets=wide_sheets)
        TargetImportService(
            file_path=self.file_path,
            upload_id=self.upload.id,
            workbook=self.workbook,
        ).run(
            year=year,
            month=month,
        )
        self._remove_legacy_general_vacancy_facts(year, month)
        self.transform_raw_to_facts(year, month, week_number=week_number)
        self.rebuild_summary(year, month)
        self.apply_balance_summary(year, month)
        self.apply_weekly_sales_summary(year, month)
        self.persist_national_dashboard_metrics(year, month)

        available_sheets = (self.workbook or {}).keys()
        if CompetitionImportService.has_competition_sheets(available_sheets):
            competition_result = CompetitionImportService(
                file_path=self.file_path,
                upload_id=self.upload.id,
                year=year,
                month=month,
                week_number=week_number,
            ).run()
            competition_summary = competition_result.get("summary", {})
            self.statistics["competition_records"] = competition_summary.get("total_inserted", 0)
            self.statistics["competition_duplicates"] = competition_summary.get("total_duplicates", 0)
            self.statistics["competition_invalid"] = competition_summary.get("total_invalid", 0)
            self.statistics["competition_source_records"] = competition_summary.get("numeric_cells", 0)
        else:
            self.warnings.append("Rekabet etiketi taşıyan bir sayfa bulunamadığı için rekabet importu atlandı.")
        self._finalize_source_reconciliation()

    def write_audit_log(self, year, month, week_number, success):
        """Write an ImportAuditLog record for this import run."""
        log = ImportAuditLog(
            upload_id=self.upload.id,
            year=year,
            month=month,
            week_number=week_number,
            uploaded_by=self.uploaded_by,
            rows_inserted=self.statistics.get("facts_inserted", 0),
            rows_updated=self.statistics.get("facts_updated", 0),
            rows_skipped=self.statistics.get("skipped_records", 0),
            rows_unmatched=(
                self.statistics.get("unmatched_representatives", 0)
                + self.statistics.get("unmatched_products", 0)
                + self.statistics.get("unmatched_regions", 0)
                + self.statistics.get("unmatched_provinces", 0)
            ),
            rows_error=self.statistics.get("rows_error", 0) + len(self.errors),
            queued_for_manual=self.statistics.get("queued_for_manual", 0),
            processing_time=round(time.monotonic() - self.started, 2),
            status="COMPLETED" if success else "FAILED",
            notes=("\n".join(self.warnings) or None),
        )
        db.session.add(log)

    def finish(self, success=True, year=None, month=None, week_number=None):
        self.upload.processing_time = round(time.monotonic() - self.started, 2)
        self.upload.sheet_count = self.statistics["sheet_count"]
        self.upload.raw_record_count = self.statistics["raw_records"]
        self.upload.fact_record_count = self.statistics["fact_records"]
        self.upload.summary_record_count = self.statistics["summary_records"]
        self.upload.source_record_count = self.statistics["source_metric_records"]
        self.upload.stored_source_record_count = self.statistics["stored_source_records"]
        self.upload.zero_metric_count = self.statistics["zero_metric_records"]
        self.upload.blank_metric_count = self.statistics["blank_metric_records"]
        self.upload.invalid_metric_count = self.statistics["invalid_metric_records"]
        self.upload.excluded_aggregate_count = self.statistics["aggregate_rows_excluded"]
        self.upload.reconciliation_status = self.statistics["reconciliation_status"]
        self.upload.warning_message = "\n".join(self.warnings) or None
        self.upload.error_message = "\n".join(self.errors) or None
        self.upload.status = "COMPLETED" if success else "FAILED"
        self.upload.completed_at = datetime.utcnow()
        if year and month:
            self.write_audit_log(year, month, week_number, success)

    def report(self):
        return {
            "success": not self.errors,
            "upload_id": self.upload.id if self.upload else None,
            "statistics": self.statistics,
            "warnings": self.warnings,
            "errors": self.errors,
            "unknown_products": sorted(set(self.unknown_products)),
            "unknown_columns": sorted(set(self.unknown_columns)),
            "skipped_logs": self.skipped_logs,
            "excluded_logs": self.excluded_logs,
            "parser_decisions": self.parser_decisions,
            "processing_time": round(time.monotonic() - self.started, 2),
        }

    def validate(self):
        if not os.path.isfile(self.file_path):
            raise FileNotFoundError(f"IMS dosyası bulunamadı: {self.file_path}")
        if not self.file_path.lower().endswith((".xlsx", ".xls")):
            raise ValueError("Yalnızca .xlsx ve .xls dosyaları içe aktarılabilir.")
        return True

    def _persist_failure(self, year, month, week_number=None):
        failure_upload = IMSUpload(
            file_name=os.path.basename(self.file_path),
            year=year,
            month=month,
            week_number=week_number,
            quarter=self.quarter_for(month),
            uploaded_by=self.uploaded_by,
            status="FAILED",
            processing_time=round(time.monotonic() - self.started, 2),
            error_message="\n".join(self.errors),
            warning_message="\n".join(self.warnings) or None,
            completed_at=datetime.utcnow(),
            source_record_count=self.statistics["source_metric_records"],
            stored_source_record_count=self.statistics["stored_source_records"],
            zero_metric_count=self.statistics["zero_metric_records"],
            blank_metric_count=self.statistics["blank_metric_records"],
            invalid_metric_count=self.statistics["invalid_metric_records"],
            excluded_aggregate_count=self.statistics["aggregate_rows_excluded"],
            reconciliation_status=self.statistics.get("reconciliation_status", "FAILED"),
        )
        db.session.add(failure_upload)
        db.session.commit()
        self.upload = failure_upload

    def run(self, year, month, clear_before_import=False, week_number=None):
        """Run the full ETL pipeline.

        week_number is extracted from the file name when not provided explicitly.
        When a week_number is available the pipeline performs an idempotent UPSERT
        instead of a destructive clear+insert.  clear_before_import is retained for
        backward compatibility but defaults to False.
        """
        if week_number is None:
            week_number = self.extract_week_number(self.file_path)

        try:
            self.validate()
            AliasService.warmup()
            self.load_workbook()
            detected_month = self.detect_workbook_month()
            if detected_month and detected_month != month:
                self.warnings.append(
                    f"Form ayı ({month}) Excel üst bilgisinden algılanan ayla ({detected_month}) değiştirildi."
                )
                month = detected_month
            self.create_upload(year, month, week_number=week_number)
            workbook_rows_read = sum(len(dataframe.index) for dataframe in self.workbook.values())
            self._log_stage_metrics("workbook_rows_read", workbook_rows_read=workbook_rows_read)
            if clear_before_import and week_number is None:
                self.clear_month(year, month)
            self.process_workbook(year, month, week_number=week_number)
            skipped_rows = dict(Counter(item.get("reason", "unknown") for item in self.skipped_logs))
            self._log_stage_metrics("parsed_rows", parsed_rows=self.statistics.get("processed_rows", 0))
            self._log_stage_metrics(
                "detected_representatives",
                detected_representatives=(
                    self.statistics.get("matched_representatives", 0)
                    + self.statistics.get("unmatched_representatives", 0)
                ),
            )
            self._log_stage_metrics("skipped_rows", skipped_rows=skipped_rows)
            self._log_stage_metrics("staged_raw_rows", staged_raw_rows=self.statistics.get("raw_records", 0))
            self._log_stage_metrics(
                "created_raw_records",
                created_raw_records=self.statistics.get("raw_records", 0),
            )
            self._log_stage_metrics(
                "created_facts",
                created_facts=self.statistics.get("fact_records", 0),
                facts_inserted=self.statistics.get("facts_inserted", 0),
                facts_updated=self.statistics.get("facts_updated", 0),
            )
            self._log_stage_metrics(
                "created_summaries",
                created_summaries=self.statistics.get("summary_records", 0),
            )
            self.finish(success=True, year=year, month=month, week_number=week_number)
            db.session.commit()
        except (OSError, ValueError, SQLAlchemyError, Exception) as exc:
            db.session.rollback()
            logger.error(f"IMS Import Pipeline Failed: {exc}", exc_info=True)
            self.errors.append(str(exc))
            try:
                self._persist_failure(year, month, week_number=week_number)
            except SQLAlchemyError as persistence_error:
                db.session.rollback()
                logger.error(f"Failed to persist IMS upload failure: {persistence_error}", exc_info=True)
                self.errors.append(f"Hata kaydı yazılamadı: {persistence_error}")
        return self.report()

    def import_file(self, year, month):
        """Backward-compatible alias for callers of the previous service."""
        return self.run(year, month, clear_before_import=False)

    @classmethod
    def health(cls):
        return {"service": "IMSImportService", "version": "2.1.0", "status": "READY"}

    @classmethod
    def supported_reports(cls):
        return list(cls.REPORT_SHEETS.values())
