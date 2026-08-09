import copy
import json
import time
import threading
import hashlib
from datetime import date, datetime
from zoneinfo import ZoneInfo
from pathlib import Path

from flask import current_app
from openpyxl import Workbook
from app.models import IMSSummary, PrimeRule, Product, Setting, Target
from app.services.ai_analytics_service import AIAnalyticsService

_AI_SERVICE_INSTANCE = None
_AI_SERVICE_LOCK = threading.Lock()

def get_ai_service():
    global _AI_SERVICE_INSTANCE
    if _AI_SERVICE_INSTANCE is None:
        with _AI_SERVICE_LOCK:
            if _AI_SERVICE_INSTANCE is None:
                _AI_SERVICE_INSTANCE = AIAnalyticsService()
    return _AI_SERVICE_INSTANCE


class _CachedModel:
    def __init__(self, orm_obj):
        for k, v in orm_obj.__dict__.items():
            if not k.startswith('_'):
                setattr(self, k, v)


CACHE_TTL = 300


class TTLDataCache:
    def __init__(self, ttl=CACHE_TTL):
        self.ttl = ttl
        self._cache = {}
        self._lock = threading.RLock()

    def get(self, key):
        with self._lock:
            entry = self._cache.get(key)
            if not entry:
                return None
            value, expires_at = entry
            if time.time() > expires_at:
                self._cache.pop(key, None)
                return None
            return copy.deepcopy(value)

    def set(self, key, value):
        with self._lock:
            if len(self._cache) > 1000:
                now = time.time()
                expired = [k for k, v in self._cache.items() if v[1] <= now]
                for k in expired:
                    self._cache.pop(k, None)
            self._cache[key] = (copy.deepcopy(value), time.time() + self.ttl)

    def invalidate(self):
        with self._lock:
            self._cache.clear()


_GLOBAL_CACHE = TTLDataCache()
_RESULT_CACHE = TTLDataCache(ttl=CACHE_TTL)
_FILE_LOCK = threading.RLock()


def _cache_clear():
    """Compatibility hook used by the legacy wrapper and regression tests."""
    _GLOBAL_CACHE.invalidate()
    _RESULT_CACHE.invalidate()


class PrimeEngine:
    DEFAULT_SETTINGS = {
        "MAIN_PRIME": 50000.0,
        "CIRO_PRIME": 20000.0,
        "PRIME_STEP": 5.0,
        "STEP_AMOUNT": 2500.0,
        "MAX_PRIME_PERCENT": 140.0,
        "MIN_PRIME_PERCENT": 100.0,
        "TOTAL_PERCENT_REQUIRED": 100.0,
        "PRIME_PRODUCT_COUNT": 4.0,
        "REQUIRED_90_COUNT": 3.0,
        "REQUIRED_75_COUNT": 1.0,
        "TARGET_75": 75.0,
        "TARGET_90": 90.0,
        "ALLOW_CIRO_WITHOUT_PRODUCT": 1.0,
        "RECOVERY_EFFECT_RATE": 2.0,
        "QUARTER_EFFECT_RATE": 10.0,
        "PRODUCT_COEFFICIENT_DEFAULT": 1.0,
        "PRODUCT_BONUS_RATE": 1.0,
        "BONUS_RATE": 5.0,
        "PENALTY_RATE": 3.0,
        "PENALTY_PER_FAILED_PRODUCT": 1500.0,
        "WHAT_IF_WORST_FACTOR": 0.85,
        "WHAT_IF_EXPECTED_FACTOR": 1.10,
        "WHAT_IF_BEST_FACTOR": 1.25,
        "SLIDER_MAX_PERCENT": 150.0,
    }

    def __init__(self, representative_id, year, month, overrides=None, today=None, use_cache=True):
        self.rep_id = representative_id
        self.year = year
        self.month = month
        self.quarter = ((month - 1) // 3) + 1
        self.today = today or date.today()
        self.use_cache = use_cache
        self.overrides = overrides or {}
        
        self._calc_cache = {}
        
        self.settings = self.load_settings()
        self.products = self.load_products()
        self.product_map = {product.id: product for product in self.products}
        self.rules = self.load_rules()
        self.targets_by_period = self.load_targets()
        self.summaries_by_period = self.load_summaries()

    @classmethod
    def invalidate_global_cache(cls):
        _GLOBAL_CACHE.invalidate()
        _RESULT_CACHE.invalidate()

    def load_settings(self):
        key = "settings_all"
        cached = _GLOBAL_CACHE.get(key)
        if cached is not None:
            return cached
        settings = {item.setting_key: item.setting_value for item in Setting.query.all()}
        _GLOBAL_CACHE.set(key, settings)
        return settings

    def load_products(self):
        key = "products_all"
        cached = _GLOBAL_CACHE.get(key)
        if cached is not None:
            return cached
        products = [_CachedModel(p) for p in Product.query.filter_by(is_active=True).order_by(Product.display_order.asc(), Product.id.asc()).all()]
        _GLOBAL_CACHE.set(key, products)
        return products

    def load_rules(self):
        key = f"rules_all_{self.today.isoformat()}"
        cached = _GLOBAL_CACHE.get(key)
        if cached is not None:
            return cached
        query = PrimeRule.query.filter_by(active=True)
        rules = {}
        for rule in query.all():
            if rule.valid_from and rule.valid_from > self.today:
                continue
            if rule.valid_to and rule.valid_to < self.today:
                continue
            current = rules.get(rule.product_id)
            if current is None or ((rule.valid_from or date.min) >= (current.valid_from or date.min)):
                rules[rule.product_id] = _CachedModel(rule)
        _GLOBAL_CACHE.set(key, rules)
        return rules

    def load_targets(self):
        key = f"targets_{self.rep_id}_{self.year}"
        cached = _GLOBAL_CACHE.get(key)
        if cached is not None:
            return cached
        rows = (
            Target.query.filter_by(representative_id=self.rep_id, year=self.year)
            .filter(Target.month >= 1, Target.month <= 12)
            .all()
        )
        data = {(row.product_id, row.month): _CachedModel(row) for row in rows}
        _GLOBAL_CACHE.set(key, data)
        return data

    def load_summaries(self):
        key = f"summaries_{self.rep_id}_{self.year}"
        cached = _GLOBAL_CACHE.get(key)
        if cached is not None:
            return cached
        rows = (
            IMSSummary.query.filter_by(representative_id=self.rep_id, year=self.year)
            .filter(IMSSummary.month >= 1, IMSSummary.month <= 12)
            .all()
        )
        data = {(row.product_id, row.month): _CachedModel(row) for row in rows}
        _GLOBAL_CACHE.set(key, data)
        return data

    def get_setting(self, key, default=None, cast=float):
        if default is None:
            default = self.DEFAULT_SETTINGS.get(key, 0)
        value = self.settings.get(key, default)
        try:
            return cast(value)
        except Exception:
            try:
                return cast(default)
            except Exception:
                return default

    def get_product_coefficient(self, product):
        code_key = f"PRODUCT_COEFFICIENT_{(product.product_code or '').upper()}"
        id_key = f"PRODUCT_COEFFICIENT_ID_{product.id}"
        default = self.get_setting("PRODUCT_COEFFICIENT_DEFAULT", 1.0)
        return self.get_setting(code_key, self.get_setting(id_key, default))

    def get_prime_rule(self, product):
        return self.rules.get(product.id)

    def month_record(self, product_id, month):
        target = self.targets_by_period.get((product_id, month))
        summary = self.summaries_by_period.get((product_id, month))
        return {
            "target_unit": float(target.unit_target if target else 0),
            "target_tl": float(target.tl_target if target else 0),
            "actual_unit": float(summary.unit if summary else 0),
            "actual_tl": float(summary.tl if summary else 0),
            "bonus_amount": float(summary.bonus_amount if summary else 0),
        }

    def apply_override(self, product, month, record):
        override = (
            self.overrides.get(product.id)
            or self.overrides.get(str(product.id))
        )
        if not override or month != self.month:
            return {**record, "simulation": False, "slider_percent": None}

        actual_unit = record["actual_unit"]
        actual_tl = record["actual_tl"]
        mode = override.get("mode", "delta")
        slider_percent = override.get("slider_percent")
        target_percent = override.get("target_percent")

        if slider_percent is not None:
            slider_cap = self.get_setting("SLIDER_MAX_PERCENT", 150.0)
            slider_percent = max(0.0, min(float(slider_percent), slider_cap))
            actual_unit = actual_unit * (slider_percent / 100.0)
            actual_tl = actual_tl * (slider_percent / 100.0)

        if target_percent is not None:
            target_percent = max(0.0, float(target_percent))
            actual_unit = record["target_unit"] * (target_percent / 100.0)
            actual_tl = record["target_tl"] * (target_percent / 100.0)

        if mode == "replace":
            if override.get("unit") is not None:
                actual_unit = float(override.get("unit"))
            if override.get("tl") is not None:
                actual_tl = float(override.get("tl"))
        else:
            actual_unit += float(override.get("unit_delta", override.get("unit", 0)) or 0)
            actual_tl += float(override.get("tl_delta", override.get("tl", 0)) or 0)

        return {
            **record,
            "actual_unit": round(max(0.0, actual_unit), 2),
            "actual_tl": round(max(0.0, actual_tl), 2),
            "simulation": True,
            "slider_percent": slider_percent,
        }

    def calculate_product(self, product, month=None):
        month = month or self.month
        cache_key = (product.id, month)
        if cache_key in self._calc_cache:
            return self._calc_cache[cache_key]

        base = self.month_record(product.id, month)
        adjusted = self.apply_override(product, month, base)
        target_tl = adjusted["target_tl"]
        percent = round((adjusted["actual_tl"] / target_tl * 100.0), 2) if target_tl > 0 else 0.0
        gap_tl = round(max(0.0, target_tl - adjusted["actual_tl"]), 2)
        gap_percent = round((gap_tl / target_tl * 100.0), 2) if target_tl > 0 else 0.0
        coeff = self.get_product_coefficient(product)
        weighted_actual = round(adjusted["actual_tl"] * coeff, 2)
        weighted_target = round(target_tl * coeff, 2)
        rule = self.get_prime_rule(product)
        required_percent = float(rule.required_percent) if rule else float(getattr(product, 'required_percent', 0) or 0)
        include_in_total = bool(rule.include_in_total_tl) if rule else bool(getattr(product, 'include_total_tl', False))
        include_in_prime = bool(rule.include_in_prime) if rule else bool(getattr(product, 'is_prime_product', False))

        result = {
            "product_id": product.id,
            "product_code": product.product_code,
            "product_name": product.product_name,
            "month": month,
            "target_unit": round(base["target_unit"], 2),
            "target_tl": round(target_tl, 2),
            "actual_unit": round(adjusted["actual_unit"], 2),
            "actual_tl": round(adjusted["actual_tl"], 2),
            "percent": percent,
            "gap_tl": gap_tl,
            "gap_percent": gap_percent,
            "bonus_amount": round(base["bonus_amount"], 2),
            "required_percent": required_percent,
            "include_in_total_tl": include_in_total,
            "include_in_prime": include_in_prime,
            "coefficient": coeff,
            "weighted_actual_tl": weighted_actual,
            "weighted_target_tl": weighted_target,
            "simulation": adjusted["simulation"],
            "slider_percent": adjusted["slider_percent"],
            "passed": percent >= required_percent if include_in_prime else True,
        }
        self._calc_cache[cache_key] = result
        return result

    def calculate_monthly_products(self, month=None):
        month = month or self.month
        return [self.calculate_product(product, month=month) for product in self.products]

    def summarize_products(self, products):
        total_target = sum(item["target_tl"] for item in products if item["include_in_total_tl"])
        total_actual = sum(item["actual_tl"] for item in products if item["include_in_total_tl"])
        total_percent = round((total_actual / total_target * 100.0), 2) if total_target > 0 else 0.0
        required_total = self.get_setting("TOTAL_PERCENT_REQUIRED", 100.0)
        entitlement = self.evaluate_monthly_entitlement(products)
        passed = entitlement["passed_products"]
        failed = entitlement["failed_products"]
        return {
            "total_target": round(total_target, 2),
            "total_realization": round(total_actual, 2),
            "total_tl_percent": total_percent,
            "passed_products": passed,
            "failed_products": failed,
            "product_success": entitlement["product_success"],
            "total_success": total_percent >= required_total,
            "prime_eligible": bool(entitlement["product_success"] and total_percent >= required_total),
            "entitlement": entitlement,
        }

    def evaluate_monthly_entitlement(self, products):
        """Apply the flexible four-product monthly entitlement rule.

        The 75% allowance belongs to monthly performance, not to a fixed
        product: all configured prime products must be at least 75%, and at
        least three of the four must be at least 90%.
        """
        prime_products = [item for item in products if item["include_in_prime"]]
        configured_count = max(1, int(self.get_setting("PRIME_PRODUCT_COUNT", 4.0)))
        minimum_percent = self.get_setting("TARGET_75", 75.0)
        standard_percent = self.get_setting("TARGET_90", 90.0)
        allowed_below_standard = max(0, int(self.get_setting("REQUIRED_75_COUNT", 1.0)))
        required_standard = max(0, int(self.get_setting("REQUIRED_90_COUNT", configured_count - allowed_below_standard)))

        below_minimum = [item for item in prime_products if item["percent"] < minimum_percent]
        below_standard = [item for item in prime_products if minimum_percent <= item["percent"] < standard_percent]
        standard_or_above = [item for item in prime_products if item["percent"] >= standard_percent]
        configuration_complete = len(prime_products) == configured_count
        product_success = (
            configuration_complete
            and not below_minimum
            and len(below_standard) <= allowed_below_standard
            and len(standard_or_above) >= required_standard
        )

        blocked_reasons = []
        if not configuration_complete:
            blocked_reasons.append("Dört ana ürün yapılandırması eksik.")
        if below_minimum:
            blocked_reasons.append("%75 altındaki ürünler: " + ", ".join(item["product_name"] for item in below_minimum))
        if len(below_standard) > allowed_below_standard:
            blocked_reasons.append("%90 altındaki ürün sayısı izin verilen sınırı aşıyor.")
        if len(standard_or_above) < required_standard:
            blocked_reasons.append("En az üç ana ürünün %90 ve üzeri olması gerekiyor.")

        return {
            "required_product_count": configured_count,
            "actual_product_count": len(prime_products),
            "minimum_percent": minimum_percent,
            "standard_percent": standard_percent,
            "allowed_below_standard": allowed_below_standard,
            "required_standard_count": required_standard,
            "configuration_complete": configuration_complete,
            "passed_products": standard_or_above + below_standard,
            "failed_products": below_minimum,
            "below_minimum_products": below_minimum,
            "below_standard_products": below_standard,
            "standard_or_above_products": standard_or_above,
            "product_success": product_success,
            "blocked_reasons": blocked_reasons,
        }

    def calculate_main_prime(self, total_percent):
        minimum = self.get_setting("MIN_PRIME_PERCENT", 100.0)
        maximum = self.get_setting("MAX_PRIME_PERCENT", 140.0)
        base = self.get_setting("MAIN_PRIME", 50000.0)
        step = max(1.0, self.get_setting("PRIME_STEP", 5.0))
        step_amount = self.get_setting("STEP_AMOUNT", 2500.0)
        
        if total_percent < minimum:
            return 0.0
            
        capped = min(total_percent, maximum)
        level = int(round(capped - minimum, 4) // step)
        return round(base + (level * step_amount), 2)

    def calculate_ciro_prime(self, total_percent, product_success):
        minimum = self.get_setting("TOTAL_PERCENT_REQUIRED", 100.0)
        if total_percent < minimum or not product_success:
            return 0.0
        return round(self.get_setting("CIRO_PRIME", 20000.0), 2)

    def calculate_bonus(self, summary, products):
        db_bonus = sum(item["bonus_amount"] for item in products)
        bonus_rate = self.get_setting("BONUS_RATE", 5.0)
        threshold = self.get_setting("TOTAL_PERCENT_REQUIRED", 100.0)
        over_target = max(0.0, summary["total_tl_percent"] - threshold)
        config_bonus = summary["total_realization"] * (bonus_rate / 100.0) * (over_target / 100.0)
        return round(db_bonus + config_bonus, 2)

    def calculate_penalty(self, summary):
        penalty_rate = self.get_setting("PENALTY_RATE", 3.0)
        per_failed = self.get_setting("PENALTY_PER_FAILED_PRODUCT", 1500.0)
        shortfall_percent = max(0.0, self.get_setting("TOTAL_PERCENT_REQUIRED", 100.0) - summary["total_tl_percent"])
        penalty = (summary["total_target"] * (penalty_rate / 100.0) * (shortfall_percent / 100.0))
        penalty += len(summary["failed_products"]) * per_failed
        return round(penalty, 2)

    def calculate_recovery_component(self, products, baseline_products=None):
        baseline_products = baseline_products or products
        baseline_gap = sum(item["gap_tl"] for item in baseline_products if item["include_in_prime"])
        simulated_gap = sum(item["gap_tl"] for item in products if item["include_in_prime"])
        recovered = max(0.0, baseline_gap - simulated_gap)
        return round(recovered * (self.get_setting("RECOVERY_EFFECT_RATE", 2.0) / 100.0), 2)

    def calculate_quarter_component(self, quarter_percent):
        threshold = self.get_setting("TOTAL_PERCENT_REQUIRED", 100.0)
        if quarter_percent <= threshold:
            return 0.0
        return round((quarter_percent - threshold) * self.get_setting("QUARTER_EFFECT_RATE", 10.0), 2)

    def calculate_product_component(self, products):
        rate = self.get_setting("PRODUCT_BONUS_RATE", 1.0) / 100.0
        component = 0.0
        for item in products:
            if not item["include_in_prime"]:
                continue
            upside = max(0.0, item["weighted_actual_tl"] - item["weighted_target_tl"])
            component += upside * rate
        return round(component, 2)

    def months_in_quarter(self, quarter=None):
        quarter = quarter or self.quarter
        start = ((quarter - 1) * 3) + 1
        return [start, start + 1, start + 2]

    def build_quarter_analysis(self):
        products = []
        total_target = 0.0
        total_actual = 0.0
        for product in self.products:
            monthly = [self.calculate_product(product, month=month) for month in self.months_in_quarter()]
            target_tl = sum(item["target_tl"] for item in monthly)
            actual_tl = sum(item["actual_tl"] for item in monthly)
            percent = round((actual_tl / target_tl * 100.0), 2) if target_tl > 0 else 0.0
            remaining_tl = round(max(0.0, target_tl - actual_tl), 2)
            status = "Tamamlandı" if remaining_tl <= 0.0 else ("Takip" if percent >= 85 else "Riskli")
            product_entry = {
                "product_id": product.id,
                "product": product.product_name,
                "percent": percent,
                "target_tl": round(target_tl, 2),
                "actual_tl": round(actual_tl, 2),
                "remaining_tl": remaining_tl,
                "status": status,
                "monthly": [
                    {
                        "month": item["month"],
                        "target_tl": item["target_tl"],
                        "actual_tl": item["actual_tl"],
                        "percent": item["percent"],
                    }
                    for item in monthly
                ],
            }
            total_target += target_tl
            total_actual += actual_tl
            products.append(product_entry)

        total_percent = round((total_actual / total_target * 100.0), 2) if total_target > 0 else 0.0
        completed = len([item for item in products if item["remaining_tl"] <= 0.0])
        return {
            "quarter": self.quarter,
            "months": self.months_in_quarter(),
            "products": products,
            "completed_products": completed,
            "failed_products": len(products) - completed,
            "target_tl": round(total_target, 2),
            "realization_tl": round(total_actual, 2),
            "total_percent": total_percent,
        }

    def build_recovery_analysis(self, products):
        recovery_rows = []
        for item in products:
            remaining_tl = item["gap_tl"]
            remaining_box = max(0.0, item["target_unit"] - item["actual_unit"])
            risk_score = 100 if item["target_tl"] <= 0 else max(0, 100 - int(item["gap_percent"]))
            status = "Tamamlandı"
            if remaining_tl > 0:
                if risk_score >= 80:
                    status = "Güvenli"
                elif risk_score >= 60:
                    status = "Takip"
                elif risk_score >= 40:
                    status = "Riskli"
                else:
                    status = "Kritik"
            recovery_rows.append(
                {
                    "product_id": item["product_id"],
                    "product": item["product_name"],
                    "product_name": item["product_name"],
                    "remaining_box": round(remaining_box, 2),
                    "remaining_tl": remaining_tl,
                    "risk_score": risk_score,
                    "status": status,
                    "description": "Hedef kapandı." if remaining_tl <= 0.0 else f"₺{remaining_tl:,.0f} açık bulunuyor.",
                    "can_close": remaining_tl <= 0.0,
                }
            )
        recovery_rows.sort(key=lambda item: (item["remaining_tl"] > 0, item["remaining_tl"]), reverse=True)
        return recovery_rows

    def build_breakdown(self, monthly_products, baseline_products, summary, quarter):
        eligible = summary["prime_eligible"]
        main_prime = self.calculate_main_prime(summary["total_tl_percent"]) if eligible else 0.0
        ciro_prime = self.calculate_ciro_prime(summary["total_tl_percent"], eligible)
        quarter_effect = self.calculate_quarter_component(quarter["total_percent"]) if eligible else 0.0
        product_effect = self.calculate_product_component(monthly_products) if eligible else 0.0
        recovery_prime = self.calculate_recovery_component(monthly_products, baseline_products=baseline_products) if eligible else 0.0
        bonus = self.calculate_bonus(summary, monthly_products) if eligible else 0.0
        penalty = self.calculate_penalty(summary) if eligible else 0.0
        extra_prime = round(ciro_prime + quarter_effect + product_effect, 2)
        total = round(main_prime + extra_prime + recovery_prime + bonus - penalty, 2)
        return {
            "main_prime": round(main_prime, 2),
            "ciro_prime": round(ciro_prime, 2),
            "quarter_effect": round(quarter_effect, 2),
            "product_effect": round(product_effect, 2),
            "extra_prime": extra_prime,
            "recovery": round(recovery_prime, 2),
            "bonus": round(bonus, 2),
            "penalty": round(penalty, 2),
            "total": total,
        }

    def build_insights(self, products, breakdown):
        sorted_by_delta = sorted(products, key=lambda item: item["actual_tl"] - item["target_tl"], reverse=True)
        best = sorted_by_delta[0] if sorted_by_delta else None
        worst = sorted_by_delta[-1] if sorted_by_delta else None
        target_total = sum(item["target_tl"] for item in products if item["include_in_total_tl"])
        missing_total = sum(item["gap_tl"] for item in products if item["include_in_total_tl"])
        loss_percent = round((missing_total / target_total * 100.0), 2) if target_total > 0 else 0.0
        return {
            "most_profitable_product": {
                "product": best["product_name"],
                "impact_tl": round(best["actual_tl"] - best["target_tl"], 2),
                "percent": best["percent"],
            } if best else None,
            "most_harmful_product": {
                "product": worst["product_name"],
                "impact_tl": round(worst["actual_tl"] - worst["target_tl"], 2),
                "percent": worst["percent"],
            } if worst else None,
            "missing_product_impact_percent": loss_percent,
            "missing_product_impact_tl": round(missing_total, 2),
            "top_breakdown_driver": max(breakdown, key=lambda key: breakdown[key]) if breakdown else None,
        }

    def build_comparison(self, current_total, forecast_total, max_total):
        return {
            "labels": ["Gerçekleşen", "Beklenen", "Maksimum"],
            "values": [round(current_total, 2), round(forecast_total, 2), round(max_total, 2)],
        }

    def build_what_if_scenarios(self, base_products, baseline_products=None):
        baseline_products = baseline_products or base_products
        scenarios = []
        factors = [
            ("worst", self.get_setting("WHAT_IF_WORST_FACTOR", 0.85)),
            ("current", 1.0),
            ("expected", self.get_setting("WHAT_IF_EXPECTED_FACTOR", 1.10)),
            ("best", self.get_setting("WHAT_IF_BEST_FACTOR", 1.25)),
        ]
        for label, factor in factors:
            scaled_products = []
            for item in base_products:
                scaled = dict(item)
                if label != "current":
                    scaled["actual_unit"] = round(item["actual_unit"] * factor, 2)
                    scaled["actual_tl"] = round(item["actual_tl"] * factor, 2)
                    scaled["percent"] = round((scaled["actual_tl"] / scaled["target_tl"] * 100.0), 2) if scaled["target_tl"] > 0 else 0.0
                    scaled["gap_tl"] = round(max(0.0, scaled["target_tl"] - scaled["actual_tl"]), 2)
                    scaled["gap_percent"] = round((scaled["gap_tl"] / scaled["target_tl"] * 100.0), 2) if scaled["target_tl"] > 0 else 0.0
                    scaled["passed"] = scaled["percent"] >= scaled["required_percent"] if scaled["include_in_prime"] else True
                scaled_products.append(scaled)
            summary = self.summarize_products(scaled_products)
            quarter_percent = round(summary["total_tl_percent"] * factor, 2) if label != "current" else summary["total_tl_percent"]
            quarter = {"total_percent": quarter_percent}
            breakdown = self.build_breakdown(scaled_products, baseline_products, summary, quarter)
            scenarios.append(
                {
                    "key": label,
                    "label": {
                        "worst": "En Kötü",
                        "current": "Mevcut",
                        "expected": "Beklenen",
                        "best": "En İyi",
                    }[label],
                    "factor": round(factor, 2),
                    "total_percent": summary["total_tl_percent"],
                    "total_prime": breakdown["total"],
                    "main_prime": breakdown["main_prime"],
                    "bonus": breakdown["bonus"],
                    "penalty": breakdown["penalty"],
                }
            )
        return scenarios

    def build_trends(self):
        monthly_rows = []
        periods = sorted({month for _, month in self.targets_by_period.keys()} | {month for _, month in self.summaries_by_period.keys()})
        if not periods:
            periods = [self.month]
        for month in periods:
            products = self.calculate_monthly_products(month=month)
            summary = self.summarize_products(products)
            monthly_rows.append(
                {
                    "year": self.year,
                    "month": month,
                    "label": f"{self.year}-{month:02d}",
                    "target_tl": summary["total_target"],
                    "actual_tl": summary["total_realization"],
                    "prime": round(
                        self.calculate_main_prime(summary["total_tl_percent"]) + self.calculate_ciro_prime(summary["total_tl_percent"], summary["prime_eligible"])
                        if summary["prime_eligible"] else 0.0,
                        2,
                    ),
                    "percent": summary["total_tl_percent"],
                }
            )
        quarter_rows = []
        for quarter in range(1, 5):
            months = self.months_in_quarter(quarter)
            target = 0.0
            actual = 0.0
            for product in self.products:
                for month in months:
                    info = self.calculate_product(product, month=month)
                    if info["include_in_total_tl"]:
                        target += info["target_tl"]
                        actual += info["actual_tl"]
            percent = round((actual / target * 100.0), 2) if target > 0 else 0.0
            quarter_rows.append(
                {
                    "year": self.year,
                    "quarter": quarter,
                    "label": f"Q{quarter} {self.year}",
                    "prime": round(self.calculate_main_prime(percent), 2),
                    "percent": percent,
                }
            )
        yearly_target = sum(row["target_tl"] for row in monthly_rows)
        yearly_actual = sum(row["prime"] for row in monthly_rows)
        return {
            "monthly": monthly_rows,
            "quarterly": quarter_rows,
            "yearly": [
                {
                    "year": self.year,
                    "label": str(self.year),
                    "prime": round(yearly_actual, 2),
                    "target_tl": round(yearly_target, 2),
                }
            ],
        }

    def build_forecast(self, summary, breakdown):
        ai_service = get_ai_service()
        next_month = ai_service.predict_next_month()
        current_target = max(summary["total_target"], 1.0)
        forecast_percent = round((next_month.get("predicted_tl", 0.0) / current_target * 100.0), 2)
        expected_prime = self.calculate_main_prime(forecast_percent) + self.calculate_ciro_prime(forecast_percent, summary["product_success"])
        return {
            "predicted_tl": round(next_month.get("predicted_tl", 0.0), 2),
            "trend_direction": next_month.get("trend_direction", "stable"),
            "confidence": next_month.get("confidence", 0),
            "forecast_percent": forecast_percent,
            "expected_prime": round(expected_prime, 2),
            "current_total_prime": breakdown["total"],
        }

    def build_ai_messages(self, insights, forecast, recovery):
        messages = []
        if insights.get("most_profitable_product"):
            messages.append(f"En kârlı ürün: {insights['most_profitable_product']['product']}.")
        if insights.get("most_harmful_product"):
            messages.append(f"En zararlı ürün: {insights['most_harmful_product']['product']}.")
        if forecast.get("expected_prime", 0) > 0:
            messages.append(f"AI tahmini ay sonu prim: ₺{forecast['expected_prime']:,.0f}.")
        open_recovery = [item for item in recovery if item["remaining_tl"] > 0]
        if open_recovery:
            top = open_recovery[0]
            messages.append(f"En kritik recovery alanı: {top['product']} için ₺{top['remaining_tl']:,.0f} açık.")
        return messages

    def build_result(self, save_history=True):
        current_products = self.calculate_monthly_products(month=self.month)
        baseline_products = current_products
        
        if self.overrides:
            original_overrides = self.overrides
            self.overrides = {}
            self._calc_cache = {k: v for k, v in self._calc_cache.items() if k[1] != self.month}
            baseline_products = self.calculate_monthly_products(month=self.month)
            self.overrides = original_overrides
            self._calc_cache = {k: v for k, v in self._calc_cache.items() if k[1] != self.month}
            for prod in current_products:
                self._calc_cache[(prod["product_id"], self.month)] = prod

        summary = self.summarize_products(current_products)
        quarter = self.build_quarter_analysis()
        recovery = self.build_recovery_analysis(current_products)
        breakdown = self.build_breakdown(current_products, baseline_products, summary, quarter)
        insights = self.build_insights(current_products, breakdown)
        what_if = self.build_what_if_scenarios(current_products, baseline_products=baseline_products)
        forecast = self.build_forecast(summary, breakdown)
        comparison = self.build_comparison(breakdown["total"], forecast["expected_prime"], max((item["total_prime"] for item in what_if), default=breakdown["total"]))
        trends = self.build_trends()
        history_entry = self.save_history(breakdown, summary, insights, what_if) if save_history else None
        
        result = {
            "representative_id": self.rep_id,
            "year": self.year,
            "month": self.month,
            "quarter": self.quarter,
            "products": current_products,
            "product_results": {item["product_id"]: item for item in current_products},
            "total_target": summary["total_target"],
            "total_realization": summary["total_realization"],
            "total_tl_percent": summary["total_tl_percent"],
            "main_prime": breakdown["main_prime"],
            "ciro_prime": breakdown["ciro_prime"],
            "total_prime": breakdown["total"],
            "product_success": summary["product_success"],
            "prime_eligible": summary["prime_eligible"],
            "entitlement": summary["entitlement"],
            "success": breakdown["total"] > 0,
            "status": "Ana Prim" if breakdown["main_prime"] > 0 else ("Ciro Primi" if breakdown["ciro_prime"] > 0 else "Beklemede"),
            "message": "Tüm prim koşulları sağlandı." if summary["product_success"] and summary["total_success"] else "Bazı prim koşulları eksik.",
            "simulation": bool(self.overrides),
            "breakdown": breakdown,
            "quarter_analysis": quarter,
            "recovery_analysis": recovery,
            "what_if_analysis": what_if,
            "insights": insights,
            "comparison_graph": comparison,
            "trend_graphs": trends,
            "ai_forecast": forecast,
            "history_entry": history_entry,
            "cache_ttl_seconds": CACHE_TTL,
        }
        result["ai_messages"] = self.build_ai_messages(insights, forecast, recovery)
        return result

    def cache_key(self):
        fingerprint = {
            "rep": self.rep_id,
            "year": self.year,
            "month": self.month,
            "overrides": self.overrides,
            "settings": self.settings,
        }
        raw = json.dumps(fingerprint, sort_keys=True, default=str)
        return hashlib.md5(raw.encode('utf-8')).hexdigest()

    def history_file(self):
        folder = Path(current_app.config["REPORT_FOLDER"]) / "prime_simulation_history"
        folder.mkdir(parents=True, exist_ok=True)
        return folder / f"rep_{self.rep_id}_{self.year}_{self.month}.json"

    def save_history(self, breakdown, summary, insights, what_if):
        payload = {
            "scenario_id": f"{self.rep_id}-{self.year}-{self.month}-{int(time.time() * 1000)}",
            "created_at": datetime.now(ZoneInfo("Europe/Istanbul")).isoformat(),
            "representative_id": self.rep_id,
            "year": self.year,
            "month": self.month,
            "overrides": self.overrides,
            "summary": {
                "total_prime": breakdown["total"],
                "total_percent": summary["total_tl_percent"],
                "main_prime": breakdown["main_prime"],
            },
            "insights": insights,
            "what_if": what_if,
        }
        history_file = self.history_file()
        with _FILE_LOCK:
            entries = []
            if history_file.exists():
                try:
                    entries = json.loads(history_file.read_text(encoding="utf-8"))
                except Exception:
                    entries = []
            entries.append(payload)
            entries = entries[-25:]
            temp_path = history_file.with_suffix('.tmp')
            temp_path.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")
            temp_path.replace(history_file)
        return payload

    def load_history(self):
        history_file = self.history_file()
        with _FILE_LOCK:
            if not history_file.exists():
                return []
            try:
                entries = json.loads(history_file.read_text(encoding="utf-8"))
            except Exception:
                return []
            return list(reversed(entries))

    def export_pdf(self, result, report_type="prime_report"):
        folder = Path(current_app.config["REPORT_FOLDER"]) / "prime_exports"
        folder.mkdir(parents=True, exist_ok=True)
        safe_report_type = "".join(
            ch if ch.isalnum() or ch in ("-", "_") else "_"
            for ch in str(report_type or "prime_report")
        ).strip("_") or "prime_report"
        file_path = folder / f"{safe_report_type}_{self.rep_id}_{self.year}_{self.month}.pdf"
        lines = [
            f"Prime Report - {safe_report_type}",
            f"Representative: {self.rep_id}",
            f"Period: {self.year}-{self.month:02d}",
            f"Total Prime: {result['breakdown']['total']:.2f}",
            f"Main Prime: {result['breakdown']['main_prime']:.2f}",
            f"Recovery: {result['breakdown']['recovery']:.2f}",
            f"Bonus: {result['breakdown']['bonus']:.2f}",
            f"Penalty: {result['breakdown']['penalty']:.2f}",
            f"AI Forecast: {result['ai_forecast']['expected_prime']:.2f}",
        ]
        
        pdf_bytes = self._build_pdf(lines)
        temp_path = file_path.with_suffix('.tmp')
        with _FILE_LOCK:
            temp_path.write_bytes(pdf_bytes)
            temp_path.replace(file_path)
            
        return {"path": str(file_path), "name": file_path.name, "type": safe_report_type}

    def export_excel(self, result):
        folder = Path(current_app.config["REPORT_FOLDER"]) / "prime_exports"
        folder.mkdir(parents=True, exist_ok=True)
        file_path = folder / f"prime_report_{self.rep_id}_{self.year}_{self.month}.xlsx"
        
        workbook = Workbook()
        try:
            summary_sheet = workbook.active
            summary_sheet.title = "Summary"
            summary_sheet.append(["Metric", "Value"])
            summary_sheet.append(["Total Prime", result["breakdown"]["total"]])
            summary_sheet.append(["Main Prime", result["breakdown"]["main_prime"]])
            summary_sheet.append(["Extra Prime", result["breakdown"]["extra_prime"]])
            summary_sheet.append(["Recovery", result["breakdown"]["recovery"]])
            summary_sheet.append(["Bonus", result["breakdown"]["bonus"]])
            summary_sheet.append(["Penalty", result["breakdown"]["penalty"]])

            products_sheet = workbook.create_sheet("Products")
            products_sheet.append(["Product", "Target TL", "Actual TL", "Percent", "Gap TL", "Simulation"])
            for item in result["products"]:
                products_sheet.append([
                    item["product_name"],
                    item["target_tl"],
                    item["actual_tl"],
                    item["percent"],
                    item["gap_tl"],
                    "Yes" if item["simulation"] else "No",
                ])

            scenario_sheet = workbook.create_sheet("WhatIf")
            scenario_sheet.append(["Scenario", "Factor", "Total Percent", "Total Prime"])
            for item in result["what_if_analysis"]:
                scenario_sheet.append([item["label"], item["factor"], item["total_percent"], item["total_prime"]])

            temp_path = file_path.with_suffix('.tmp')
            with _FILE_LOCK:
                workbook.save(temp_path)
                temp_path.replace(file_path)
        finally:
            workbook.close()
            
        return {"path": str(file_path), "name": file_path.name, "type": "excel"}

    def _build_pdf(self, lines):
        escaped_lines = []
        for line in lines:
            safe = str(line).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
            escaped_lines.append(f"({safe}) Tj")
        content = "BT /F1 11 Tf 50 780 Td 14 TL " + " T* ".join(escaped_lines) + " ET"
        objects = [
            "1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n",
            "2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n",
            "3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj\n",
            "4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n",
            f"5 0 obj << /Length {len(content.encode('utf-8'))} >> stream\n{content}\nendstream endobj\n",
        ]
        pdf = "%PDF-1.4\n"
        offsets = []
        for obj in objects:
            offsets.append(len(pdf.encode("utf-8")))
            pdf += obj
        xref_position = len(pdf.encode("utf-8"))
        pdf += f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n"
        for offset in offsets:
            pdf += f"{offset:010d} 00000 n \n"
        pdf += f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_position}\n%%EOF"
        return pdf.encode("utf-8")

    def calculate(self, save_history=True):
        key = self.cache_key()
        if self.use_cache:
            cached = _RESULT_CACHE.get(key)
            if cached is not None:
                cached["cache"] = {"hit": True, "ttl_seconds": CACHE_TTL}
                return cached
        result = self.build_result(save_history=save_history)
        result["cache"] = {"hit": False, "ttl_seconds": CACHE_TTL}
        if self.use_cache:
            _RESULT_CACHE.set(key, result)
        return result
