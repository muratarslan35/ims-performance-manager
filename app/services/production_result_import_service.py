"""Atomic import for approved 1./2. production realization workbooks."""

from dataclasses import dataclass, field
from datetime import datetime
import math
from pathlib import Path

from openpyxl import load_workbook

from app.extensions import db
from app.models import (
    Product,
    ProductAlias,
    ProductionRepresentativeTotal,
    ProductionNationalProductResult,
    ProductionNationalTotal,
    ProductionRegionProductResult,
    ProductionRegionTotal,
    ProductionResult,
    Representative,
    RepresentativeAlias,
    Target,
)
from app.services.alias_service import AliasService


class ProductionWorkbookValidationError(ValueError):
    """Validation failure that must leave all production overrides unapplied."""


@dataclass
class ProductionImportReport:
    rows_seen: int = 0
    matched_rows: int = 0
    product_results: list = field(default_factory=list)
    representative_totals: list = field(default_factory=list)
    region_totals: list = field(default_factory=list)
    region_product_results: list = field(default_factory=list)
    national_product_results: list = field(default_factory=list)
    national_total: dict | None = None
    target_mismatch_count: int = 0

    @property
    def matched_result_count(self):
        return len(self.product_results)


class ProductionResultImportService:
    """Read production workbooks by semantic content, not a fixed sheet layout."""

    PERCENT_TOLERANCE = 0.05
    HEADER_SCAN_ROWS = 20

    TARGET_MARKERS = {
        "TL": {"TL HEDEF", "HEDEF TL", "TL TARGET", "TARGET TL"},
        "KUTU": {"KUTU HEDEF", "HEDEF KUTU", "KUTU TARGET", "TARGET KUTU", "BOX TARGET"},
    }
    ACTUAL_MARKERS = {
        "TL": {"TL CIKIS", "CIKIS TL", "TL SATIS", "SATIS TL", "TL GERCEKLESEN", "TL ACTUAL"},
        "KUTU": {"KUTU CIKIS", "CIKIS KUTU", "KUTU SATIS", "SATIS KUTU", "KUTU GERCEKLESEN", "KUTU ACTUAL", "BOX ACTUAL"},
    }
    PERCENT_MARKERS = {"REA", "REA %", "REALIZASYON", "REALIZASYON %", "REALIZATION", "REALIZATION %"}

    def __init__(self, file_path, year, month, production_stage=None):
        self.file_path = Path(file_path)
        self.year = int(year)
        self.month = int(month)
        self.production_stage = int(production_stage) if production_stage in (1, 2, "1", "2") else None
        self._products = {}
        self._representatives = {}
        self._vacancies = {}
        self._vacancy_names = set()
        self._layout_cache = {}

    @staticmethod
    def _number(value):
        """Parse Excel/Turkish/English numeric values without discarding real zero/negatives."""
        if value is None or isinstance(value, bool):
            return None
        if isinstance(value, (int, float)):
            result = float(value)
            return result if math.isfinite(result) else None

        text = str(value).strip().replace("\u00a0", "").replace(" ", "")
        if not text or text in {"-", "—", "–"}:
            return None

        negative_parentheses = text.startswith("(") and text.endswith(")")
        if negative_parentheses:
            text = text[1:-1]
        for token in ("%", "₺", "TL", "TRY"):
            text = text.replace(token, "").replace(token.lower(), "")
        if not text:
            return None

        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(".", "").replace(",", ".")
        elif text.count(".") > 1:
            parts = text.split(".")
            if all(part.isdigit() and len(part) == 3 for part in parts[1:]):
                text = "".join(parts)

        try:
            result = float(text)
        except ValueError:
            return None
        if negative_parentheses:
            result = -result
        return result if math.isfinite(result) else None

    @staticmethod
    def _is_region(label):
        return len(label) > 4 and label[:3].isdigit() and label[3] == " "

    @staticmethod
    def _metric(metric):
        normalized = AliasService.normalize(metric)
        if normalized in {"KUTU", "BOX", "UNIT", "ADET"}:
            return "KUTU"
        return "TL"

    def _load_master_maps(self):
        normalize = AliasService.normalize
        for product in Product.query.all():
            for label in (product.product_name, product.product_code, product.ims_name):
                if label:
                    self._products[normalize(label)] = product
        for alias in ProductAlias.query.all():
            self._products[normalize(alias.alias_name)] = alias.product

        # Departed real people remain excluded from current production matching.
        # Stable vacancy identities are retained as exact vacancy candidates even
        # when temporarily inactive because the position may reappear later.
        for representative in Representative.query.all():
            is_vacancy = str(representative.rep_code or "").upper().startswith("UNASSIGNED")
            if representative.active:
                for label in (representative.rep_name, representative.rep_code, representative.ims_code):
                    if label:
                        self._representatives[normalize(label)] = representative
            if is_vacancy:
                region_key = normalize(representative.region).split(" ", 1)[0]
                self._vacancies.setdefault(region_key, []).append(representative)
                for label in (representative.rep_name, representative.rep_code, representative.ims_code):
                    if label:
                        self._vacancy_names.add(normalize(label))
        for alias in RepresentativeAlias.query.all():
            if alias.representative and alias.representative.active:
                self._representatives[normalize(alias.alias_name)] = alias.representative

    @staticmethod
    def _find_marker(values, markers):
        for index, value in enumerate(values, start=1):
            if value in markers:
                return index
        return None

    def _product_block(self, values, total_column):
        """Walk left from a semantic total marker and discover the product block."""
        columns, product_ids = [], []
        column = total_column - 1
        while column >= 1:
            product = self._products.get(values[column - 1])
            if product is None:
                break
            columns.append(column)
            product_ids.append(product.id)
            column -= 1
        columns.reverse()
        product_ids.reverse()
        if not columns or len(product_ids) != len(set(product_ids)):
            return [], []
        return columns, product_ids

    def _identity_columns(self, sheet, header_row, first_metric_column):
        """Infer representative/region columns from actual row content."""
        candidates = list(range(1, max(first_metric_column, 1)))
        if not candidates:
            return None, None
        name_scores = {column: 0 for column in candidates}
        region_scores = {column: 0 for column in candidates}
        last_row = min(sheet.max_row, header_row + 60)
        for row_number in range(header_row + 1, last_row + 1):
            for column in candidates:
                label = AliasService.normalize(sheet.cell(row_number, column).value)
                if not label:
                    continue
                if label in self._representatives or label in self._vacancy_names:
                    name_scores[column] += 8
                elif label == "NATIONAL":
                    name_scores[column] += 6
                elif self._is_region(label):
                    name_scores[column] += 2
                    region_scores[column] += 6
        name_column = max(candidates, key=lambda item: name_scores[item])
        region_candidates = [item for item in candidates if item != name_column]
        region_column = max(region_candidates, key=lambda item: region_scores[item]) if region_candidates else None

        if name_scores[name_column] == 0:
            name_column = first_metric_column - 1 if first_metric_column > 1 else None
        if region_column is None or region_scores.get(region_column, 0) == 0:
            region_column = name_column - 1 if name_column and name_column > 1 else None
        return name_column, region_column

    def _layout(self, sheet, metric):
        metric = self._metric(metric)
        cache_key = (sheet.title, metric)
        if cache_key in self._layout_cache:
            return self._layout_cache[cache_key]

        for row_number in range(1, min(sheet.max_row, self.HEADER_SCAN_ROWS) + 1):
            values = [AliasService.normalize(sheet.cell(row_number, column).value) for column in range(1, sheet.max_column + 1)]
            target_total = self._find_marker(values, self.TARGET_MARKERS[metric])
            actual_total = self._find_marker(values, self.ACTUAL_MARKERS[metric])
            percent_total = self._find_marker(values, self.PERCENT_MARKERS)
            if target_total is None or actual_total is None or percent_total is None:
                continue

            target_columns, target_ids = self._product_block(values, target_total)
            actual_columns, actual_ids = self._product_block(values, actual_total)
            percent_columns, percent_ids = self._product_block(values, percent_total)
            if not target_ids or set(target_ids) != set(actual_ids) or set(target_ids) != set(percent_ids):
                continue

            actual_by_product = dict(zip(actual_ids, actual_columns))
            percent_by_product = dict(zip(percent_ids, percent_columns))
            actual_columns = [actual_by_product[product_id] for product_id in target_ids]
            percent_columns = [percent_by_product[product_id] for product_id in target_ids]

            first_metric_column = min(target_columns + actual_columns + percent_columns)
            name_column, region_column = self._identity_columns(sheet, row_number, first_metric_column)
            if name_column is None or region_column is None:
                continue

            stage_one = next((i + 1 for i, value in enumerate(values) if value == "1 URETIM"), None)
            stage_two = next((i + 1 for i, value in enumerate(values) if value == "2 URETIM"), None)
            if metric == "TL" and self.production_stage == 1 and stage_one is not None:
                total_percent_column = stage_one
            elif metric == "TL" and self.production_stage == 2 and stage_two is not None:
                total_percent_column = stage_two
            elif metric == "TL" and stage_two is not None:
                total_percent_column = stage_two
            elif metric == "TL" and stage_one is not None:
                total_percent_column = stage_one
            else:
                total_percent_column = percent_total

            layout = {
                "header_row": row_number,
                "name_column": name_column,
                "region_column": region_column,
                "actual_columns": actual_columns,
                "target_columns": target_columns,
                "percent_columns": percent_columns,
                "product_ids": target_ids,
                "target_total_column": target_total,
                "total_actual_column": actual_total,
                "total_percent_column": total_percent_column,
            }
            self._layout_cache[cache_key] = layout
            return layout
        raise ProductionWorkbookValidationError(f"{sheet.title} içinde {metric} üretim düzeni semantik olarak bulunamadı.")

    def _find_sheet(self, workbook, metric):
        """Discover a metric sheet by its content; sheet names are only a tie-breaker."""
        metric = self._metric(metric)
        candidates = []
        for sheet in workbook.worksheets:
            try:
                self._layout(sheet, metric)
            except ProductionWorkbookValidationError:
                continue
            title = AliasService.normalize(sheet.title)
            score = 0
            if metric in title:
                score += 4
            if "REALIZ" in title or "URETIM" in title or "SONUC" in title:
                score += 2
            candidates.append((score, sheet))
        if not candidates:
            raise ProductionWorkbookValidationError(f"{metric} üretim sayfası içerikten bulunamadı.")
        candidates.sort(key=lambda item: item[0], reverse=True)
        if len(candidates) > 1 and candidates[0][0] == candidates[1][0]:
            raise ProductionWorkbookValidationError(f"{metric} için birden fazla eşdeğer üretim sayfası bulundu.")
        return candidates[0][1]

    def _match_representative(self, raw_name, raw_region):
        name, region = AliasService.normalize(raw_name), AliasService.normalize(raw_region)
        if not name or name == "NATIONAL" or self._is_region(name):
            return None
        representative = self._representatives.get(name)
        if representative is not None:
            return representative
        region_key = region.split(" ", 1)[0]
        candidates = [
            item for item in self._vacancies.get(region_key, [])
            if AliasService.normalize(item.rep_name).endswith(name)
            or AliasService.normalize(item.rep_name) == name
        ]
        return candidates[0] if len(candidates) == 1 else None

    @staticmethod
    def _derived_percent(actual, target):
        if target == 0:
            return 0.0 if actual == 0 else None
        return actual * 100.0 / target

    def _read_metric_values(self, sheet, row_number, layout, context):
        targets, values, percentages = [], [], []
        for target_column, actual_column, percent_column in zip(
            layout["target_columns"], layout["actual_columns"], layout["percent_columns"]
        ):
            target = self._number(sheet.cell(row_number, target_column).value)
            actual = self._number(sheet.cell(row_number, actual_column).value)
            percent = self._number(sheet.cell(row_number, percent_column).value)
            if target is None or target < 0 or actual is None:
                raise ProductionWorkbookValidationError(
                    f"{sheet.title}!{row_number} {context} hedef/çıkış değeri geçersiz."
                )
            if percent is None:
                percent = self._derived_percent(actual, target)
            if percent is None:
                raise ProductionWorkbookValidationError(
                    f"{sheet.title}!{row_number} {context} realizasyonu güvenli biçimde türetilemedi."
                )
            targets.append(target)
            values.append(actual)
            percentages.append(percent)

        total_target = self._number(sheet.cell(row_number, layout["target_total_column"]).value)
        total_actual = self._number(sheet.cell(row_number, layout["total_actual_column"]).value)
        total_percent = self._number(sheet.cell(row_number, layout["total_percent_column"]).value)
        if total_target is None:
            total_target = sum(targets)
        if total_actual is None:
            total_actual = sum(values)
        if total_target < 0:
            raise ProductionWorkbookValidationError(f"{sheet.title}!{row_number} {context} toplam hedefi negatif olamaz.")
        if total_percent is None:
            total_percent = self._derived_percent(total_actual, total_target)
        if total_percent is None:
            raise ProductionWorkbookValidationError(
                f"{sheet.title}!{row_number} {context} nihai realizasyonu güvenli biçimde türetilemedi."
            )
        return {
            "row_number": row_number,
            "targets": targets,
            "values": values,
            "percentages": percentages,
            "total_target": total_target,
            "total_actual": total_actual,
            "total_percent": total_percent,
            "product_ids": layout["product_ids"],
        }

    def _read_sheet(self, sheet, metric):
        layout = self._layout(sheet, metric)
        rows, unresolved, duplicates = {}, [], []
        for row_number in range(layout["header_row"] + 1, sheet.max_row + 1):
            raw_name = sheet.cell(row_number, layout["name_column"]).value
            raw_region = sheet.cell(row_number, layout["region_column"]).value
            label = AliasService.normalize(raw_name)
            if not label or label == "NATIONAL" or self._is_region(label):
                continue
            representative = self._match_representative(raw_name, raw_region)
            if representative is None:
                unresolved.append(str(raw_name).strip())
                continue
            if representative.id in rows:
                duplicates.append(representative.rep_name)
                continue
            rows[representative.id] = self._read_metric_values(
                sheet, row_number, layout, "temsilci"
            )
        if unresolved:
            raise ProductionWorkbookValidationError(
                "Eşleşmeyen temsilci/boş kadro satırları: " + ", ".join(sorted(set(unresolved))[:10])
            )
        if duplicates:
            raise ProductionWorkbookValidationError(
                "Bir temsilci birden fazla kez bulundu: " + ", ".join(sorted(set(duplicates))[:10])
            )
        return rows

    def _read_national(self, sheet, metric):
        """Read the workbook's NATIONAL row; never infer it from rep rows."""
        layout = self._layout(sheet, metric)
        for row_number in range(layout["header_row"] + 1, sheet.max_row + 1):
            if AliasService.normalize(sheet.cell(row_number, layout["name_column"]).value) == "NATIONAL":
                return self._read_metric_values(sheet, row_number, layout, "NATIONAL")
        raise ProductionWorkbookValidationError(f"{sheet.title} NATIONAL satırı bulunamadı.")

    def _read_regions(self, sheet, metric):
        """Read workbook-owned region subtotal rows; never infer them from reps."""
        layout = self._layout(sheet, metric)
        rows = {}
        for row_number in range(layout["header_row"] + 1, sheet.max_row + 1):
            label = AliasService.normalize(sheet.cell(row_number, layout["name_column"]).value)
            if not self._is_region(label):
                continue
            rows[label.split(" ", 1)[0]] = self._read_metric_values(
                sheet, row_number, layout, "bölge"
            )
        return rows

    @staticmethod
    def _product_position(row, product_id):
        try:
            return row["product_ids"].index(product_id)
        except ValueError as exc:
            raise ProductionWorkbookValidationError(
                f"Ürün {product_id} TL/kutu blokları arasında eşleşmedi."
            ) from exc

    def parse(self):
        if not self.file_path.exists():
            raise ProductionWorkbookValidationError("Üretim Excel dosyası bulunamadı.")
        try:
            workbook = load_workbook(self.file_path, read_only=False, data_only=True)
        except Exception as exc:
            raise ProductionWorkbookValidationError("Excel dosyası okunamadı.") from exc

        self._load_master_maps()
        tl_sheet = self._find_sheet(workbook, "TL")
        unit_sheet = self._find_sheet(workbook, "KUTU")
        tl_rows = self._read_sheet(tl_sheet, "TL")
        unit_rows = self._read_sheet(unit_sheet, "KUTU")
        region_tl = self._read_regions(tl_sheet, "TL")
        region_unit = self._read_regions(unit_sheet, "KUTU")
        national_tl = self._read_national(tl_sheet, "TL")
        national_unit = self._read_national(unit_sheet, "KUTU")

        if set(tl_rows) != set(unit_rows):
            raise ProductionWorkbookValidationError("TL ve kutu sonuçlarında temsilci kapsamı eşit değil.")
        if set(region_tl) != set(region_unit):
            raise ProductionWorkbookValidationError("TL ve kutu sonuçlarında bölge kapsamı eşit değil.")
        if set(national_tl["product_ids"]) != set(national_unit["product_ids"]):
            raise ProductionWorkbookValidationError("TL ve kutu NATIONAL ürün kapsamı eşit değil.")
        for representative_id in tl_rows:
            if set(tl_rows[representative_id]["product_ids"]) != set(unit_rows[representative_id]["product_ids"]):
                raise ProductionWorkbookValidationError("TL ve kutu temsilci ürün kapsamı eşit değil.")
        for region_code in region_tl:
            if set(region_tl[region_code]["product_ids"]) != set(region_unit[region_code]["product_ids"]):
                raise ProductionWorkbookValidationError("TL ve kutu bölge ürün kapsamı eşit değil.")

        targets = {
            (row.representative_id, row.product_id): row
            for row in Target.query.filter_by(year=self.year, month=self.month).all()
        }
        source_keys = {
            (representative_id, product_id)
            for representative_id, row in tl_rows.items()
            for product_id in row["product_ids"]
        }
        if source_keys != set(targets):
            raise ProductionWorkbookValidationError(
                f"Üretim dosyası kapsamı dönem hedefleriyle eşit değil "
                f"(eksik={len(set(targets)-source_keys)}, fazlalık={len(source_keys-set(targets))})."
            )

        report = ProductionImportReport(rows_seen=len(tl_rows), matched_rows=len(tl_rows))
        for tl_index, product_id in enumerate(national_tl["product_ids"]):
            unit_index = self._product_position(national_unit, product_id)
            report.national_product_results.append({
                "product_id": product_id,
                "actual_tl": national_tl["values"][tl_index],
                "actual_unit": national_unit["values"][unit_index],
                "realization_percent": national_tl["percentages"][tl_index],
                "unit_realization_percent": national_unit["percentages"][unit_index],
                "source_sheet": tl_sheet.title,
                "source_row": national_tl["row_number"],
            })
        report.national_total = {
            "target_tl": national_tl["total_target"],
            "target_unit": national_unit["total_target"],
            "actual_tl": national_tl["total_actual"],
            "actual_unit": national_unit["total_actual"],
            "realization_percent": national_tl["total_percent"],
            "unit_realization_percent": national_unit["total_percent"],
            "source_sheet": tl_sheet.title,
            "source_row": national_tl["row_number"],
        }

        for region_code, tl_row in region_tl.items():
            unit_row = region_unit[region_code]
            report.region_totals.append({
                "region_code": region_code,
                "target_tl": tl_row["total_target"],
                "target_unit": unit_row["total_target"],
                "actual_tl": tl_row["total_actual"],
                "actual_unit": unit_row["total_actual"],
                "realization_percent": tl_row["total_percent"],
                "unit_realization_percent": unit_row["total_percent"],
                "source_sheet": tl_sheet.title,
                "source_row": tl_row["row_number"],
            })
            for tl_index, product_id in enumerate(tl_row["product_ids"]):
                unit_index = self._product_position(unit_row, product_id)
                report.region_product_results.append({
                    "region_code": region_code,
                    "product_id": product_id,
                    "target_tl": tl_row["targets"][tl_index],
                    "target_unit": unit_row["targets"][unit_index],
                    "actual_tl": tl_row["values"][tl_index],
                    "actual_unit": unit_row["values"][unit_index],
                    "realization_percent": tl_row["percentages"][tl_index],
                    "unit_realization_percent": unit_row["percentages"][unit_index],
                    "source_sheet": tl_sheet.title,
                    "source_row": tl_row["row_number"],
                })

        for representative_id, tl_row in tl_rows.items():
            unit_row = unit_rows[representative_id]
            for tl_index, product_id in enumerate(tl_row["product_ids"]):
                unit_index = self._product_position(unit_row, product_id)
                database_target = targets[(representative_id, product_id)]
                actual_tl = tl_row["values"][tl_index]
                actual_unit = unit_row["values"][unit_index]
                percent = tl_row["percentages"][tl_index]
                unit_percent = unit_row["percentages"][unit_index]
                source_target_tl = tl_row["targets"][tl_index]
                source_target_unit = unit_row["targets"][unit_index]

                expected = self._derived_percent(actual_tl, source_target_tl)
                if expected is None or abs(percent - expected) > self.PERCENT_TOLERANCE:
                    raise ProductionWorkbookValidationError(
                        f"{tl_sheet.title}!{tl_row['row_number']} TL realizasyonu doğrulanamadı."
                    )
                if (
                    abs(source_target_tl - float(database_target.tl_target or 0)) > self.PERCENT_TOLERANCE
                    or abs(source_target_unit - float(database_target.unit_target or 0)) > self.PERCENT_TOLERANCE
                ):
                    report.target_mismatch_count += 1
                expected_unit_percent = self._derived_percent(actual_unit, source_target_unit)
                if expected_unit_percent is None or abs(unit_percent - expected_unit_percent) > self.PERCENT_TOLERANCE:
                    raise ProductionWorkbookValidationError(
                        f"{unit_sheet.title}!{unit_row['row_number']} kutu realizasyonu doğrulanamadı."
                    )
                report.product_results.append({
                    "representative_id": representative_id,
                    "product_id": product_id,
                    "realization_percent": percent,
                    "unit_realization_percent": unit_percent,
                    "target_tl": source_target_tl,
                    "target_unit": source_target_unit,
                    "actual_tl": actual_tl,
                    "actual_unit": actual_unit,
                    "source_sheet": tl_sheet.title,
                    "source_row": tl_row["row_number"],
                })
            report.representative_totals.append({
                "representative_id": representative_id,
                "realization_percent": tl_row["total_percent"],
                "unit_realization_percent": unit_row["total_percent"],
                "target_tl": tl_row["total_target"],
                "target_unit": unit_row["total_target"],
                "actual_tl": tl_row["total_actual"],
                "actual_unit": unit_row["total_actual"],
                "source_sheet": tl_sheet.title,
                "source_row": tl_row["row_number"],
            })
        return report

    @staticmethod
    def apply(upload, report):
        for row in report.product_results:
            db.session.add(ProductionResult(upload_id=upload.id, **row))
        for row in report.representative_totals:
            db.session.add(ProductionRepresentativeTotal(upload_id=upload.id, **row))
        for row in report.national_product_results:
            db.session.add(ProductionNationalProductResult(upload_id=upload.id, **row))
        db.session.add(ProductionNationalTotal(upload_id=upload.id, **report.national_total))
        ProductionResultImportService.apply_region_totals(upload, report)
        upload.row_count, upload.matched_row_count = report.rows_seen, report.matched_result_count
        upload.status = upload.STATUS_APPLIED
        upload.validated_at = datetime.utcnow()
        upload.applied_at = datetime.utcnow()
        upload.warning_message = (
            f"{report.matched_rows} temsilci ve {report.matched_result_count} ürün satırı doğrulandı. "
            "TL/kutu hedef ve sonuçları üretim dosyasındaki nihai değerleriyle ayrı korunur; "
            "öncelik 2. üretim → 1. üretim → IMS'tir."
            + (
                f" {report.target_mismatch_count} IMS hedef farkı, kaynak veriyi değiştirmeden "
                "üretim sonucu kapsamında kaydedildi."
                if report.target_mismatch_count else ""
            )
        )

    @staticmethod
    def apply_region_totals(upload, report):
        """Upsert official region rows; safe to call for historical backfill."""
        ProductionRegionProductResult.query.filter_by(upload_id=upload.id).delete(synchronize_session=False)
        ProductionRegionTotal.query.filter_by(upload_id=upload.id).delete(synchronize_session=False)
        for row in report.region_product_results:
            db.session.add(ProductionRegionProductResult(upload_id=upload.id, **row))
        for row in report.region_totals:
            db.session.add(ProductionRegionTotal(upload_id=upload.id, **row))
