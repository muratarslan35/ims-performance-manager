"""Read workbook-authoritative regional MF'siz KUTU BAKİYE values.

The existing ``dashboard_balance_region`` compatibility row predates box-balance
persistence and stores TL target/actual in its generic ``unit``/``tl`` columns.
For current archived IMS workbooks, regional remaining boxes therefore need to
be read from the archived BAKİYE sheet itself. This helper is read-only, cached
per immutable upload archive, and limited to the regional unit-balance section.
"""
from __future__ import annotations

import math
import os
from threading import Lock

from flask import current_app
from openpyxl import load_workbook

from app.models import Product
from app.services.alias_service import AliasService

_CACHE = {}
_CACHE_LOCK = Lock()


def _key(value):
    return "".join(ch for ch in AliasService.normalize(value) if ch.isalnum())


def _number_or_none(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return None if math.isnan(number) else number
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except (TypeError, ValueError):
        return None


def _product_aliases():
    result = {}
    for product in Product.query.all():
        aliases = {
            _key(product.product_name),
            _key(product.product_code),
            _key(product.ims_name),
        } - {""}
        result[int(product.id)] = aliases
    return result


def _resolve_product_id(value, aliases):
    value_key = _key(value)
    if not value_key:
        return None
    exact = [pid for pid, keys in aliases.items() if value_key in keys]
    if len(exact) == 1:
        return exact[0]
    contains = [
        pid for pid, keys in aliases.items()
        if any(key in value_key or value_key in key for key in keys)
    ]
    return contains[0] if len(contains) == 1 else None


def _parse_archive(upload_id):
    archive = os.path.join(current_app.config["UPLOAD_FOLDER"], "ims_archive", f"upload-{int(upload_id)}.xlsx")
    if not os.path.exists(archive):
        return {}
    mtime = os.path.getmtime(archive)
    cache_key = (int(upload_id), mtime)
    with _CACHE_LOCK:
        cached = _CACHE.get(cache_key)
    if cached is not None:
        return cached

    aliases = _product_aliases()
    workbook = load_workbook(archive, read_only=True, data_only=True)
    try:
        sheet = next((ws for ws in workbook.worksheets if "BAKIYE" in AliasService.normalize(ws.title)), None)
        if sheet is None:
            parsed = {}
        else:
            header_limit = min(12, sheet.max_row)
            unit_start = None
            for row_index in range(1, header_limit + 1):
                for column in range(1, sheet.max_column + 1):
                    label = AliasService.normalize(sheet.cell(row_index, column).value)
                    if "BAKIYE" in label and any(token in label for token in ("KUTU", "UNIT", "ADET")):
                        unit_start = column
                        break
                if unit_start is not None:
                    break

            product_columns = {}
            if unit_start is not None:
                for column in range(unit_start, sheet.max_column + 1):
                    resolved = None
                    for row_index in range(1, header_limit + 1):
                        resolved = _resolve_product_id(sheet.cell(row_index, column).value, aliases)
                        if resolved is not None:
                            break
                    if resolved is not None:
                        product_columns[column] = resolved

            parsed = {}
            if product_columns:
                for row_index in range(header_limit + 1, sheet.max_row + 1):
                    left = sheet.cell(row_index, 1).value
                    right = sheet.cell(row_index, 2).value
                    left_key, right_key = _key(left), _key(right)
                    if not left_key or left_key != right_key:
                        continue
                    region_code = "".join(ch for ch in str(left or "").strip() if ch.isdigit())[:3]
                    if not region_code:
                        continue
                    for column, product_id in product_columns.items():
                        value = _number_or_none(sheet.cell(row_index, column).value)
                        if value is not None:
                            parsed[(region_code, int(product_id))] = value
    finally:
        workbook.close()

    with _CACHE_LOCK:
        _CACHE.clear()
        _CACHE[cache_key] = parsed
    return parsed


def region_balance_units(upload_id, region_key):
    """Return remaining box balance by product for one sales region."""
    region_code = "".join(ch for ch in str(region_key or "") if ch.isdigit())[:3]
    if not upload_id or not region_code:
        return {}
    parsed = _parse_archive(int(upload_id))
    return {
        product_id: value
        for (code, product_id), value in parsed.items()
        if code == region_code
    }
