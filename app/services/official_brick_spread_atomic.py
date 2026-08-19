"""Integrate an official brick-spread master into the IMS publish transaction.

Presence is discovered from content, never from a required sheet title. Workbooks
that do not contain this capability remain valid; when the capability is present,
it is persisted inside the same IMS transaction and any ambiguity/unresolved row
blocks publication. The legacy route call is idempotent.
"""
from __future__ import annotations

from openpyxl import load_workbook

from app.models import IMSRawData
from app.services.alias_service import AliasService


def _discover_spread_sheet(cls, workbook):
    candidates = []
    for worksheet in workbook.worksheets:
        score = 0
        scanned = []
        for row in worksheet.iter_rows(min_row=1, max_row=min(30, worksheet.max_row or 0), values_only=True):
            normalized = [AliasService.normalize(value) for value in row if value is not None]
            scanned.extend(value for value in normalized if value)
        text = " | ".join(scanned)
        if "BRICK SAYISI" in text:
            score += 8
        if "BOLGE" in text or "REGION" in text:
            score += 2
        if "TOPLAM" in text or "TOTAL" in text:
            score += 1
        if "BRICK SAYISI" in text and len(set(scanned)) >= 6:
            score += 2
        if score >= 10:
            candidates.append((score, worksheet.title))
    if not candidates:
        return None
    candidates.sort(key=lambda item: (-item[0], item[1]))
    best_score = candidates[0][0]
    best = [name for score, name in candidates if score == best_score]
    if len(best) != 1:
        from app.services.official_brick_spread_service import OfficialBrickSpreadError
        raise OfficialBrickSpreadError(
            "Resmi brick yayılım masterı birden fazla sheet ile aynı güven skorunda eşleşti: "
            + ", ".join(best)
        )
    return best[0]


def _file_has_spread(cls, file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return cls._sheet_name(workbook) is not None
    finally:
        workbook.close()


def install_official_brick_spread_atomic():
    from app.services.ims_import_service import IMSImportService
    from app.services.official_brick_spread_service import OfficialBrickSpreadService

    if getattr(IMSImportService, "_official_brick_spread_atomic_installed", False):
        return

    OfficialBrickSpreadService._sheet_name = classmethod(_discover_spread_sheet)
    original_persist = OfficialBrickSpreadService.persist

    def persist_idempotent(cls, *, file_path, upload_id, year=None, month=None, week_number=None):
        existing = IMSRawData.query.filter_by(upload_id=int(upload_id), sheet_type=cls.SHEET_TYPE).all()
        if existing:
            representatives = len({row.representative_id for row in existing if row.representative_id is not None})
            return {
                "sheet_name": existing[0].sheet_name,
                "representatives": representatives,
                "product_columns": max(0, int(len(existing) / max(1, representatives)) - 1),
                "records": len(existing),
                "aggregate_rows_ignored": 0,
                "already_persisted": True,
            }
        # Capability-aware behavior: absence is not an error. If content says
        # the master exists, the original strict parser owns all validation.
        if not _file_has_spread(cls, file_path):
            return {
                "sheet_name": None,
                "representatives": 0,
                "product_columns": 0,
                "records": 0,
                "aggregate_rows_ignored": 0,
                "already_persisted": False,
                "present": False,
            }
        result = original_persist(
            file_path=file_path,
            upload_id=upload_id,
            year=year,
            month=month,
            week_number=week_number,
        )
        result["present"] = True
        return result

    OfficialBrickSpreadService.persist = classmethod(persist_idempotent)
    original_process = IMSImportService.process_workbook

    def process_with_official_spread(self, year, month, week_number=None):
        result = original_process(self, year, month, week_number=week_number)
        spread = OfficialBrickSpreadService.persist(
            file_path=self.file_path,
            upload_id=self.upload.id,
            year=year,
            month=month,
            week_number=week_number,
        )
        self.statistics["official_brick_spread_records"] = spread["records"]
        self.statistics["official_brick_spread_representatives"] = spread["representatives"]
        self.statistics["official_brick_spread_product_columns"] = spread["product_columns"]
        self.statistics["official_brick_spread_present"] = int(bool(spread.get("present", spread["records"])))
        self.statistics["official_brick_spread_atomic"] = 1
        return result

    IMSImportService.process_workbook = process_with_official_spread
    IMSImportService._official_brick_spread_atomic_installed = True
