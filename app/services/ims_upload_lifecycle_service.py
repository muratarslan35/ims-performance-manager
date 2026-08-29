"""Safe lifecycle management for imported IMS workbooks.

The service deliberately separates three concerns:

* hide/show only controls the IMS history UI and never changes calculations;
* duplicate detection first uses source SHA-256 and, when an archived source is
  available, also compares workbook cell data while ignoring formatting/metadata;
* physical deletion is allowed only when the current period state can be restored.

Weekly raw/fact/competition history is upload-scoped, while Target, IMSSummary and
RepresentativeBrickAssignment are period-scoped current-state tables. Before a
new import is published we persist an exact rollback snapshot of those three
period-scoped tables. Deleting the latest upload removes the newer upload-owned
rows first and restores the previous current-state snapshot as the final write.
"""
from __future__ import annotations

import hashlib
import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from flask import current_app
from sqlalchemy import DateTime

from app.extensions import db
from app.models import (
    IMSImportJob,
    IMSSummary,
    IMSUpload,
    RepresentativeBrickAssignment,
    Setting,
    Target,
)


class IMSUploadLifecycleService:
    HIDDEN_KEY_PREFIX = "IMS_UPLOAD_HIDDEN_"
    SNAPSHOT_VERSION = 2

    @classmethod
    def _archive_root(cls) -> Path:
        root = Path(current_app.config["UPLOAD_FOLDER"]) / "ims_archive"
        root.mkdir(parents=True, exist_ok=True)
        return root

    @classmethod
    def _snapshot_root(cls) -> Path:
        root = cls._archive_root() / "snapshots"
        root.mkdir(parents=True, exist_ok=True)
        return root

    @classmethod
    def pending_snapshot_path(cls, job_id: int) -> Path:
        return cls._snapshot_root() / f"job-{int(job_id)}.json"

    @classmethod
    def upload_snapshot_path(cls, upload_id: int) -> Path:
        return cls._snapshot_root() / f"upload-{int(upload_id)}.json"

    @classmethod
    def archived_source_path(cls, upload_id: int, suffix: str = ".xlsx") -> Path:
        safe_suffix = suffix.lower() if suffix.lower() in {".xlsx", ".xls"} else ".xlsx"
        return cls._archive_root() / f"upload-{int(upload_id)}{safe_suffix}"

    @classmethod
    def archived_source_for_upload(cls, upload_id: int) -> Path | None:
        for suffix in (".xlsx", ".xls"):
            candidate = cls.archived_source_path(upload_id, suffix)
            if candidate.is_file():
                return candidate
        return None

    @staticmethod
    def _semantic_cell(value):
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, Decimal):
            value = float(value)
        if isinstance(value, float):
            if value != value:
                return None
            if value == 0:
                return 0
            return format(value, ".15g")
        if isinstance(value, int):
            return str(value)
        if isinstance(value, str):
            return value.strip()
        return str(value).strip()

    @classmethod
    def semantic_workbook_hash(cls, file_path: Path) -> str | None:
        """Hash workbook sheet/cell values while ignoring workbook formatting."""
        path = Path(file_path)
        if path.suffix.lower() != ".xlsx" or not path.is_file():
            return None
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=False)
            digest = hashlib.sha256()
            try:
                for sheet in workbook.worksheets:
                    digest.update(b"S\0")
                    digest.update(str(sheet.title).strip().encode("utf-8"))
                    digest.update(b"\0")
                    last_nonempty_row = 0
                    buffered_rows = []
                    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        normalized = [cls._semantic_cell(value) for value in row]
                        while normalized and normalized[-1] is None:
                            normalized.pop()
                        if normalized:
                            last_nonempty_row = row_index
                        buffered_rows.append((row_index, normalized))
                    for row_index, normalized in buffered_rows:
                        if row_index > last_nonempty_row:
                            break
                        digest.update(b"R\0")
                        digest.update(str(row_index).encode("ascii"))
                        digest.update(b"\0")
                        digest.update(
                            json.dumps(normalized, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
                        )
                        digest.update(b"\0")
            finally:
                workbook.close()
            return digest.hexdigest()
        except Exception:
            current_app.logger.exception("ims_semantic_fingerprint_failed file=%s", path.name)
            return None

    @classmethod
    def same_semantic_workbook(cls, new_path: Path, existing_upload_id: int) -> bool | None:
        archived = cls.archived_source_for_upload(existing_upload_id)
        if archived is None:
            return None
        new_hash = cls.semantic_workbook_hash(new_path)
        old_hash = cls.semantic_workbook_hash(archived)
        if new_hash is None or old_hash is None:
            return None
        return new_hash == old_hash

    @staticmethod
    def _serialize_rows(rows: Iterable[object]) -> list[dict]:
        result = []
        for row in rows:
            values = {}
            for column in row.__table__.columns:
                value = getattr(row, column.name)
                if hasattr(value, "isoformat"):
                    value = value.isoformat()
                values[column.name] = value
            result.append(values)
        return result

    @classmethod
    def capture_period_snapshot(cls, *, job_id: int, year: int, month: int) -> Path:
        """Persist the exact current-state rows before a new IMS replaces them."""
        payload = {
            "version": cls.SNAPSHOT_VERSION,
            "year": int(year),
            "month": int(month),
            "targets": cls._serialize_rows(
                Target.query.filter_by(year=int(year), month=int(month)).all()
            ),
            "summaries": cls._serialize_rows(
                IMSSummary.query.filter_by(year=int(year), month=int(month)).all()
            ),
            "brick_assignments": cls._serialize_rows(
                RepresentativeBrickAssignment.query.filter_by(year=int(year), month=int(month)).all()
            ),
        }
        path = cls.pending_snapshot_path(job_id)
        path.write_text(json.dumps(payload, ensure_ascii=False, sort_keys=True), encoding="utf-8")
        return path

    @classmethod
    def discard_pending_snapshot(cls, job_id: int) -> None:
        cls.pending_snapshot_path(job_id).unlink(missing_ok=True)

    @classmethod
    def finalize_snapshot(cls, *, job_id: int, upload_id: int) -> Path | None:
        source = cls.pending_snapshot_path(job_id)
        if not source.exists():
            return None
        destination = cls.upload_snapshot_path(upload_id)
        source.replace(destination)
        return destination

    @classmethod
    def archive_successful_source(cls, *, staging_path: Path, upload_id: int) -> Path:
        destination = cls.archived_source_path(upload_id, staging_path.suffix)
        destination.write_bytes(staging_path.read_bytes())
        return destination

    @classmethod
    def exact_duplicate_job(cls, source_hash: str) -> IMSImportJob | None:
        return (
            IMSImportJob.query
            .join(IMSUpload, IMSImportJob.ims_upload_id == IMSUpload.id)
            .filter(
                IMSImportJob.source_hash == str(source_hash),
                IMSImportJob.status == IMSImportJob.STATUS_COMPLETED,
                IMSUpload.status == "COMPLETED",
            )
            .order_by(IMSImportJob.completed_at.desc(), IMSImportJob.id.desc())
            .first()
        )

    @classmethod
    def existing_week_job(cls, *, year: int, month: int, week_number: int | None) -> IMSImportJob | None:
        if week_number is None:
            return None
        return (
            IMSImportJob.query
            .join(IMSUpload, IMSImportJob.ims_upload_id == IMSUpload.id)
            .filter(
                IMSImportJob.year == int(year),
                IMSImportJob.month == int(month),
                IMSImportJob.status == IMSImportJob.STATUS_COMPLETED,
                IMSUpload.status == "COMPLETED",
                IMSUpload.week_number == int(week_number),
            )
            .order_by(IMSUpload.completed_at.desc(), IMSUpload.id.desc())
            .first()
        )

    @classmethod
    def hidden_setting_key(cls, upload_id: int) -> str:
        return f"{cls.HIDDEN_KEY_PREFIX}{int(upload_id)}"

    @classmethod
    def hidden_upload_ids(cls) -> set[int]:
        rows = Setting.query.filter(Setting.setting_key.like(f"{cls.HIDDEN_KEY_PREFIX}%")).all()
        hidden = set()
        for row in rows:
            try:
                hidden.add(int(row.setting_key[len(cls.HIDDEN_KEY_PREFIX):]))
            except (TypeError, ValueError):
                continue
        return hidden

    @classmethod
    def set_hidden(cls, upload_id: int, hidden: bool) -> None:
        key = cls.hidden_setting_key(upload_id)
        row = Setting.query.filter_by(setting_key=key).first()
        if hidden:
            if row is None:
                db.session.add(Setting(
                    setting_key=key,
                    setting_value="1",
                    category="IMS",
                    description="IMS geçmiş listesinde gizlenen yükleme",
                ))
        elif row is not None:
            db.session.delete(row)
        db.session.commit()

    @classmethod
    def _latest_completed_for_period(cls, upload: IMSUpload) -> IMSUpload | None:
        return (
            IMSUpload.query
            .filter_by(year=upload.year, month=upload.month, status="COMPLETED")
            .order_by(IMSUpload.week_number.desc(), IMSUpload.completed_at.desc(), IMSUpload.id.desc())
            .first()
        )

    @classmethod
    def can_delete(cls, upload: IMSUpload) -> tuple[bool, str]:
        if upload.status != "COMPLETED":
            return True, ""
        latest = cls._latest_completed_for_period(upload)
        if latest is None or latest.id != upload.id:
            return True, ""
        if cls.upload_snapshot_path(upload.id).exists():
            return True, ""
        return False, "Bu eski IMS için geri dönüş snapshot'ı yok; aktif dönem güvenle geri alınamaz."

    @staticmethod
    def _prepare_core_rows(model, rows: list[dict]) -> list[dict]:
        columns = {column.name: column for column in model.__table__.columns}
        prepared = []
        for values in rows:
            clean = {}
            for key, value in values.items():
                column = columns.get(key)
                if column is None:
                    continue
                if value is not None and isinstance(column.type, DateTime) and isinstance(value, str):
                    value = datetime.fromisoformat(value)
                clean[key] = value
            prepared.append(clean)
        return prepared

    @classmethod
    def _restore_rows_core(cls, model, rows: list[dict]) -> None:
        prepared = cls._prepare_core_rows(model, rows)
        if prepared:
            db.session.execute(model.__table__.insert(), prepared)

    @classmethod
    def _expunge_current_state_models(cls) -> None:
        current_state_models = (Target, IMSSummary, RepresentativeBrickAssignment)
        for obj in list(db.session.identity_map.values()):
            if isinstance(obj, current_state_models):
                db.session.expunge(obj)

    @classmethod
    def _restore_period_snapshot(cls, upload: IMSUpload) -> None:
        upload_id = int(upload.id)
        year = int(upload.year)
        month = int(upload.month)
        path = cls.upload_snapshot_path(upload_id)
        if not path.exists():
            raise RuntimeError("Silme için gerekli geri dönüş snapshot'ı bulunamadı.")
        payload = json.loads(path.read_text(encoding="utf-8"))
        if int(payload.get("version", 0)) != cls.SNAPSHOT_VERSION:
            raise RuntimeError("Geri dönüş snapshot sürümü desteklenmiyor.")
        if (int(payload.get("year", 0)), int(payload.get("month", 0))) != (year, month):
            raise RuntimeError("Geri dönüş snapshot dönemi IMS dönemiyle eşleşmiyor.")

        cls._expunge_current_state_models()

        target_table = Target.__table__
        summary_table = IMSSummary.__table__
        assignment_table = RepresentativeBrickAssignment.__table__
        db.session.execute(
            target_table.delete().where(target_table.c.year == year, target_table.c.month == month)
        )
        db.session.execute(
            summary_table.delete().where(summary_table.c.year == year, summary_table.c.month == month)
        )
        db.session.execute(
            assignment_table.delete().where(
                assignment_table.c.year == year,
                assignment_table.c.month == month,
            )
        )

        cls._restore_rows_core(Target, payload.get("targets") or [])
        cls._restore_rows_core(IMSSummary, payload.get("summaries") or [])
        cls._restore_rows_core(RepresentativeBrickAssignment, payload.get("brick_assignments") or [])

    @classmethod
    def _delete_direct_upload_children(cls, upload_id: int) -> None:
        """Delete every declared direct child of IMSUpload in dependency order.

        Runtime SQLite FK introspection can be incomplete depending on how an
        existing database was migrated. SQLAlchemy model metadata is the source
        contract used by the application itself, so use it to discover upload
        ownership. Reversing ``sorted_tables`` guarantees dependent rows such as
        IMSFact are deleted before IMSRawData.
        """
        deleted_tables = set()
        for table in reversed(db.metadata.sorted_tables):
            if table.name == IMSUpload.__tablename__:
                continue
            upload_columns = []
            for foreign_key in table.foreign_keys:
                if foreign_key.column.table.name != IMSUpload.__tablename__:
                    continue
                if foreign_key.column.name != "id":
                    continue
                upload_columns.append(foreign_key.parent)
            for column in upload_columns:
                db.session.execute(table.delete().where(column == int(upload_id)))
                deleted_tables.add(table.name)

        # Fail closed if core IMS ownership declarations unexpectedly disappear.
        required = {"ims_raw_data", "ims_facts", "ims_summary", "ims_competition_data"}
        missing = required.intersection(db.metadata.tables) - deleted_tables
        if missing:
            raise RuntimeError(
                "IMS silme sahiplik sözleşmesi eksik: " + ", ".join(sorted(missing))
            )

    @classmethod
    def delete_upload(cls, upload_id: int) -> dict:
        active = IMSImportJob.query.filter(
            IMSImportJob.status.in_((IMSImportJob.STATUS_QUEUED, IMSImportJob.STATUS_PROCESSING))
        ).first()
        if active is not None:
            raise RuntimeError("Aktif IMS importu varken silme yapılamaz.")

        upload = db.session.get(IMSUpload, int(upload_id))
        if upload is None:
            raise LookupError("IMS yüklemesi bulunamadı.")
        allowed, reason = cls.can_delete(upload)
        if not allowed:
            raise RuntimeError(reason)

        deleted_upload_id = int(upload.id)
        hidden_key = cls.hidden_setting_key(deleted_upload_id)
        latest = cls._latest_completed_for_period(upload) if upload.status == "COMPLETED" else None
        restored = bool(latest is not None and latest.id == deleted_upload_id)

        try:
            with db.session.no_autoflush:
                cls._delete_direct_upload_children(deleted_upload_id)
                db.session.execute(
                    Setting.__table__.delete().where(Setting.setting_key == hidden_key)
                )
                db.session.execute(
                    IMSUpload.__table__.delete().where(IMSUpload.id == deleted_upload_id)
                )
                if restored:
                    cls._restore_period_snapshot(upload)
            db.session.commit()
        except Exception:
            db.session.rollback()
            raise

        for suffix in (".xlsx", ".xls"):
            cls.archived_source_path(deleted_upload_id, suffix).unlink(missing_ok=True)
        cls.upload_snapshot_path(deleted_upload_id).unlink(missing_ok=True)
        return {"deleted_upload_id": deleted_upload_id, "restored_previous_period_state": restored}
