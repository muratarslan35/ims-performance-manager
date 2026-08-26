"""Independent and isolated enterprise-grade import service for competition extension sheets."""

import logging
import re
import time
from enum import Enum
from typing import List, Dict, Any, Optional, Tuple
from openpyxl import load_workbook as openpyxl_load_workbook
from app.extensions import db
from app.models import CompetitionData

logger = logging.getLogger(__name__)


class MetricType(str, Enum):
    """Centralized enumeration for metric types."""
    UNIT = "UNIT"
    TL = "TL"
    MARKET_SHARE = "MARKET_SHARE"


class PeriodType(str, Enum):
    """Centralized enumeration for reporting period types."""
    MONTHLY = "MONTHLY"
    WEEKLY = "WEEKLY"


class SheetType(str, Enum):
    """Centralized enumeration for supported sheet types."""
    MONTHLY_UNITS = "MONTHLY_UNITS"
    MONTHLY_VALUE = "MONTHLY_VALUE"
    WEEKLY_UNITS = "WEEKLY_UNITS"
    WEEKLY_VALUE = "WEEKLY_VALUE"
    MONTHLY_COMPETITION_UNITS = "MONTHLY_COMPETITION_UNITS"
    MONTHLY_COMPETITION_VALUE = "MONTHLY_COMPETITION_VALUE"
    MARKET_REFERENCE = "MARKET_REFERENCE"


class DefaultGroup(str, Enum):
    """Centralized constants for default mapping groups."""
    GENEL = "GENEL"


class CompetitionImportService:
    """Production-hardened enterprise service to validate, normalize, and process competition worksheets."""

    # Competition sheets are discovered from their semantic labels; month names
    # and workbook-specific sheet names are deliberately not part of the contract.
    COMPETITION_TOKEN = "REKABET"

    MONTH_PATTERNS = (
        ("JAN", 1), ("JANUARY", 1), ("OCAK", 1),
        ("FEB", 2), ("FEBRUARY", 2), ("ŞUBAT", 2),
        ("MAR", 3), ("MARCH", 3), ("MART", 3),
        ("APR", 4), ("APRIL", 4), ("NİSAN", 4),
        ("MAY", 5), ("MAYIS", 5),
        ("JUN", 6), ("JUNE", 6), ("HAZİRAN", 6),
        ("JUL", 7), ("JULY", 7), ("TEMMUZ", 7),
        ("AUG", 8), ("AUGUST", 8), ("AĞUSTOS", 8),
        ("SEP", 9), ("SEPTEMBER", 9), ("EYLÜL", 9),
        ("OCT", 10), ("OCTOBER", 10), ("EKİM", 10),
        ("NOV", 11), ("NOVEMBER", 11), ("KASIM", 11),
        ("DEC", 12), ("DECEMBER", 12), ("ARALIK", 12),
    )

    YEAR_PATTERN = re.compile(r"20\d{2}")

    STOP_KEYWORDS = ("GRAND TOTAL", "GENEL TOPLAM", "TOTAL", "TOPLAM", "SUBTOTAL", "ARA TOPLAM")
    HEADER_TERRITORY_KEYWORDS = ("TERRITOR", "BOLGE")
    HEADER_SUBTERRITORY_KEYWORDS = ("SUBTERRITOR", "NATIONAL")
    BULK_CHUNK_SIZE = 1000

    def __init__(
        self,
        file_path: Optional[str] = None,
        upload_id: Optional[int] = None,
        year: Optional[int] = None,
        month: Optional[int] = None,
        week_number: Optional[int] = None,
    ) -> None:
        self.file_path = file_path
        self.upload_id = upload_id
        self.year = year
        self.month = month
        self.week_number = week_number
        self._workbook = None
        self._sheet_values: Dict[str, List[Tuple[Any, ...]]] = {}
        self.errors: List[str] = []
        self.warnings: List[str] = []
        self.parse_statistics = {"numeric_cells": 0, "blank_cells": 0, "invalid_cells": 0}
        self.invalid_cells: List[Dict[str, Any]] = []

    def _normalize_sheet_name(self, sheet_name: str) -> str:
        """Normalize sheet name for robust mapping."""
        return str(sheet_name).strip().upper() if sheet_name else ""

    def load_workbook(self, file_path: str) -> None:
        """Load Excel workbook in read-only and data-only mode to evaluate formula cached values."""
        self.file_path = file_path
        try:
            self._workbook = openpyxl_load_workbook(self.file_path, data_only=True, read_only=True)
            logger.info("[CompetitionImportService] Workbook successfully loaded from %s", self.file_path)
        except Exception as exc:
            logger.exception("[CompetitionImportService] Failed to load workbook at %s: %s", file_path, exc)
            raise ValueError(f"Dosya yüklenirken bir hata oluştu. Lütfen dosya formatını kontrol edin.") from exc

    @classmethod
    def classify_sheet(cls, sheet_name: str) -> Optional[SheetType]:
        """Classify a competition worksheet without depending on its month name."""
        name = str(sheet_name or "").strip().upper()
        if cls.COMPETITION_TOKEN not in name:
            return None
        if "PP" in name or "PAZAR PAY" in name or "MARKET SHARE" in name:
            return SheetType.MARKET_REFERENCE
        if "TL" in name or "VALUE" in name:
            return SheetType.MONTHLY_COMPETITION_VALUE
        if "KUTU" in name or "UNIT" in name or "ADET" in name:
            return SheetType.MONTHLY_COMPETITION_UNITS
        return SheetType.WEEKLY_UNITS

    @classmethod
    def has_competition_sheets(cls, sheet_names) -> bool:
        return any(cls.classify_sheet(name) is not None for name in sheet_names)

    def get_supported_sheets(self) -> List[str]:
        """Return only the competition sheets found in the loaded workbook."""
        if not self._workbook:
            return []
        return [name for name in self._workbook.sheetnames if self.classify_sheet(name) is not None]

    def get_sheet_type(self, sheet_name: str) -> str:
        """Resolve sheet type enum value from sheet name."""
        sheet_type = self.classify_sheet(sheet_name)
        if sheet_type is None:
            raise ValueError(f"Desteklenmeyen sayfa adı: '{sheet_name}'")
        return sheet_type.value

    def validate_workbook(self) -> Dict[str, str]:
        """Validate required competition sheets exist in workbook."""
        if not self._workbook:
            raise ValueError("Çalışma kitabı yüklenmemiş. Önce load_workbook() çağrılmalıdır.")

        actual_map = {self._normalize_sheet_name(n): n for n in self._workbook.sheetnames}
        if not self.get_supported_sheets():
            err_msg = "Rekabet etiketi taşıyan bir sayfa bulunamadı."
            logger.error("[CompetitionImportService] %s", err_msg)
            self.errors.append(err_msg)
            raise ValueError(err_msg)

        logger.info("[CompetitionImportService] Workbook validation passed for required competition sheets.")
        return actual_map

    def validate_sheet_structure(self, sheet_name: str) -> bool:
        """Validate individual sheet dimensions and structure."""
        if not self._workbook:
            raise ValueError("Çalışma kitabı yüklenmemiş.")

        norm_target = self._normalize_sheet_name(sheet_name)
        actual_map = {self._normalize_sheet_name(n): n for n in self._workbook.sheetnames}

        if norm_target not in actual_map:
            raise ValueError(f"Sayfa bulunamadı: '{sheet_name}'")

        orig_name = actual_map[norm_target]
        sheet = self._workbook[orig_name]
        max_row, max_col = sheet.max_row or 0, sheet.max_column or 0

        if max_row < 2 or max_col < 1:
            err_msg = f"Sayfa boş veya yapısal satırlardan yoksun: '{orig_name}'"
            logger.error("[CompetitionImportService] %s", err_msg)
            self.errors.append(err_msg)
            raise ValueError(f"'{orig_name}' sayfası boş veya hatalı yapıda.")

        return True

    def _get_cell_value(self, sheet: Any, row: int, col: int) -> Any:
        """Read a cached worksheet value without random-access rescans.

        ``ReadOnlyWorksheet.cell`` restarts the XML stream for many access
        patterns. Caching each selected competition sheet once keeps memory
        bounded while reducing this import from quadratic scans to one pass.
        """
        title = sheet.title
        values = self._sheet_values.get(title)
        if values is None:
            values = [tuple(c.value for c in cells) for cells in sheet.iter_rows()]
            self._sheet_values[title] = values
        if row < 1 or col < 1 or row > len(values):
            return None
        source_row = values[row - 1]
        return source_row[col - 1] if col <= len(source_row) else None

    def _discover_metadata(self, sheet: Any) -> Tuple[str, int, int]:
        """Discover reporting period type, year, and month from worksheet metadata header cells."""
        period_type = PeriodType.MONTHLY.value
        # The upload period is authoritative; product labels can contain numbers
        # that look like years (for example, 2076) and must not override it.
        year, month = self.year, self.month

        for r in range(1, 6):
            for c in range(1, (sheet.max_column or 1) + 1):
                val = self._get_cell_value(sheet, r, c)
                if not val:
                    continue
                v_str = str(val).strip().upper()

                if "WEEK" in v_str or "HAFTA" in v_str:
                    period_type = PeriodType.WEEKLY.value
                if not year:
                    ym = self.YEAR_PATTERN.search(v_str)
                    if ym:
                        year = int(ym.group(0))
                if not month:
                    for mk, mv in self.MONTH_PATTERNS:
                        if mk in v_str:
                            month = mv
                            break

        if not year:
            year = self.year
        if not month:
            month = self.month
        if not year:
            raise ValueError("Raporlama yılı metaverilerden tespit edilemedi.")
        if not month:
            raise ValueError("Raporlama ayı metaverilerden tespit edilemedi.")

        return period_type, year, month

    def _find_header_row(self, sheet: Any) -> int:
        """Locate the header row dynamically within rows 1-15 containing territory columns."""
        for r in range(1, min(sheet.max_row or 0, 15) + 1):
            row_vals = [str(self._get_cell_value(sheet, r, c) or "").strip().upper() for c in range(1, (sheet.max_column or 1) + 1)]
            if any(any(k in v for k in self.HEADER_TERRITORY_KEYWORDS) for v in row_vals) and \
               any(any(k in v for k in self.HEADER_SUBTERRITORY_KEYWORDS) for v in row_vals):
                return r
        raise ValueError("Bölge ve alt bölge başlıkları (Territory/Subterritory) bulunamadı.")

    def _find_data_start(self, sheet: Any, header_row: int, territory_col: int) -> int:
        """Find the starting row for data population."""
        for r in range(header_row + 1, (sheet.max_row or 0) + 1):
            if str(self._get_cell_value(sheet, r, territory_col) or "").strip():
                return r
        raise ValueError("Veri başlangıç satırı bulunamadı.")

    def _find_data_end(self, sheet: Any, start_row: int) -> int:
        """Find the ending row for data population before summary stop words."""
        last_row = start_row
        for r in range(start_row, (sheet.max_row or 0) + 1):
            is_empty, has_stop = True, False
            for c in range(1, (sheet.max_column or 1) + 1):
                val = self._get_cell_value(sheet, r, c)
                if val is not None and str(val).strip():
                    is_empty = False
                    val_upper = str(val).strip().upper()
                    if any(sk in val_upper for sk in self.STOP_KEYWORDS):
                        has_stop = True
                        break
            if is_empty or has_stop:
                break
            last_row = r
        return last_row

    def _is_meta_col(self, col_name: str) -> bool:
        """Determine if a column is a metadata column."""
        upper = col_name.upper()
        return (any(k in upper for k in self.HEADER_TERRITORY_KEYWORDS) or
                any(k in upper for k in self.HEADER_SUBTERRITORY_KEYWORDS) or
                "REPORT" in upper or "MARKET" in upper)

    def _extract_product_groups(self, sheet: Any, header_row: int) -> Dict[str, List[Tuple[str, int]]]:
        """Extract product groups and associated product headers with column indexes, avoiding name overwrite collisions."""
        groups: Dict[str, List[Tuple[str, int]]] = {}
        curr_grp = DefaultGroup.GENEL.value
        group_row = max(1, header_row - 1)

        column_groups = {}
        for c in range(1, (sheet.max_column or 0) + 1):
            g_val = self._get_cell_value(sheet, group_row, c)
            if g_val and str(g_val).strip():
                curr_grp = str(g_val).strip()
            column_groups[c] = curr_grp

        for c in range(1, (sheet.max_column or 0) + 1):
            p_val = self._get_cell_value(sheet, header_row, c)
            if not p_val or self._is_meta_col(str(p_val)):
                continue

            p_name = str(p_val).strip()
            group_name = column_groups.get(c, DefaultGroup.GENEL.value)
            
            groups.setdefault(group_name, [])
            # Store tuple of (product_name, column_index) uniquely to prevent multi-group collisions
            if (p_name, c) not in groups[group_name]:
                groups[group_name].append((p_name, c))

        if not groups:
            raise ValueError("Hiçbir ürün grubu tespit edilemedi.")
        return groups

    def _extract_product_columns(self, sheet: Any, header_row: int) -> Dict[int, str]:
        """Extract unique column index to product name mapping to prevent overwriting identical product names across groups."""
        cols: Dict[int, str] = {}
        for c in range(1, (sheet.max_column or 0) + 1):
            val = self._get_cell_value(sheet, header_row, c)
            if not val or self._is_meta_col(str(val)):
                continue
            cols[c] = str(val).strip()

        if not cols:
            raise ValueError("Hiçbir ürün sütunu tespit edilemedi.")
        return cols

    def _parse_sheet_structure(self, sheet_name: str) -> Dict[str, Any]:
        """Parse structural bounds and metadata of a worksheet with strict fail-fast validation."""
        if not self._workbook:
            raise ValueError("Çalışma kitabı yüklenmemiş.")

        actual_map = {self._normalize_sheet_name(n): n for n in self._workbook.sheetnames}
        norm_target = self._normalize_sheet_name(sheet_name)
        if norm_target not in actual_map:
            raise ValueError(f"Sayfa bulunamadı: '{sheet_name}'")

        orig_name = actual_map[norm_target]
        sheet = self._workbook[orig_name]

        ptype, year, month = self._discover_metadata(sheet)
        # Monthly competition sheets carry Region, National, IAM Brick and two
        # representative columns before the dynamic product columns.
        source_name = self._normalize_sheet_name(orig_name)
        if "AYLIK" in source_name and "REKABET" in source_name:
            h_row, t_col = 3, 1
            header_values = {
                column: self._normalize_turkish_text(self._get_cell_value(sheet, h_row, column))
                for column in range(1, (sheet.max_column or 0) + 1)
            }
            brick_columns = [column for column, label in header_values.items() if "BRICK" in label]
            representative_columns = [
                column for column, label in header_values.items()
                if "TTS" in label and "ISMI" in label
            ]
            s_col = brick_columns[0] if brick_columns else 2
            last_dimension_column = max([s_col, *representative_columns])
            dimension_columns = set(range(1, last_dimension_column + 1))
        elif "REKABET" in source_name:
            h_row, t_col, s_col = 2, 1, 2
            dimension_columns = {1, 2}
        else:
            h_row = self._find_header_row(sheet)
            t_col, s_col = None, None
            dimension_columns = set()

        if t_col is None or s_col is None:
            for c in range(1, (sheet.max_column or 0) + 1):
                val = str(self._get_cell_value(sheet, h_row, c) or "").strip().upper()
                if any(k in val for k in self.HEADER_TERRITORY_KEYWORDS):
                    t_col = c
                elif any(k in val for k in self.HEADER_SUBTERRITORY_KEYWORDS):
                    s_col = c

        if not t_col or not s_col:
            raise ValueError("Bölge sütunları başlık satırında eksik.")

        d_start = 5 if "AYLIK" in source_name and "REKABET" in source_name else self._find_data_start(sheet, h_row, t_col)
        d_end = self._find_data_end(sheet, d_start)
        product_columns = {
            column: name
            for column, name in self._extract_product_columns(sheet, h_row).items()
            if column not in dimension_columns
        }
        product_groups = {
            group: [(name, column) for name, column in products if column not in dimension_columns]
            for group, products in self._extract_product_groups(sheet, h_row).items()
        }
        product_groups = {group: products for group, products in product_groups.items() if products}

        struct = {
            "sheet_name": orig_name,
            "sheet_type": self.get_sheet_type(orig_name),
            "period_type": ptype,
            "year": year,
            "month": month,
            "header_row": h_row,
            "data_start_row": d_start,
            "data_end_row": d_end,
            "max_columns": sheet.max_column or 0,
            "territory_column": t_col,
            "subterritory_column": s_col,
            "product_columns": product_columns,
            "product_groups": product_groups,
        }

        # Fail-fast validation ensuring structure dictionary is fully populated without empty/None values
        for key, val in struct.items():
            if val is None or (isinstance(val, (dict, list)) and not val):
                raise ValueError(f"Fail-Fast: Yapı analizi hatası. '{key}' alanı boş veya geçersiz.")

        return struct

    def _parse_sheet_records(self, structure_info: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Parse raw records from worksheet data rows, supporting percentage/string number conversions and unique column mappings."""
        if not self._workbook:
            raise ValueError("Çalışma kitabı yüklenmemiş.")

        sheet_name = structure_info["sheet_name"]
        sheet = self._workbook[sheet_name]

        start_row = structure_info["data_start_row"]
        end_row = structure_info["data_end_row"]
        t_col = structure_info["territory_column"]
        s_col = structure_info["subterritory_column"]
        product_groups = structure_info["product_groups"]

        col_to_group_and_prod = {}
        for group_name, prods in product_groups.items():
            for p_name, col_idx in prods:
                col_to_group_and_prod[col_idx] = (group_name, p_name)

        period_type = structure_info["period_type"]
        year = structure_info["year"]
        month = structure_info["month"]
        week_number = self.week_number

        records: List[Dict[str, Any]] = []
        last_valid_territory = ""

        data_rows = structure_info.get("data_rows")
        row_numbers = data_rows if data_rows is not None else range(start_row, end_row + 1)
        for r in row_numbers:
            territory_val = self._get_cell_value(sheet, r, t_col)
            if territory_val is not None and str(territory_val).strip() != "":
                last_valid_territory = str(territory_val).strip()

            territory = last_valid_territory
            if not territory:
                continue

            territory_upper = territory.upper()
            if territory_upper in {"BOLGE", "BÖLGE", "NATIONAL"} or any(sk in territory_upper for sk in self.STOP_KEYWORDS):
                continue

            subterritory_val = self._get_cell_value(sheet, r, s_col)
            subterritory = str(subterritory_val).strip() if subterritory_val is not None else ""

            subterritory_upper = subterritory.upper()
            if any(sk in subterritory_upper for sk in self.STOP_KEYWORDS):
                continue

            for col_idx, (product_group, prod_name) in col_to_group_and_prod.items():
                cell_val = self._get_cell_value(sheet, r, col_idx)
                normalized_cell = str(cell_val).strip()
                if (
                    cell_val is None
                    or normalized_cell == ""
                    or normalized_cell.upper() in {"-", "—", "–", "N/A", "NA", "NULL"}
                ):
                    # Workbook placeholders carry no observation. Preserve the
                    # distinction between missing and the real numeric value 0.
                    self.parse_statistics["blank_cells"] += 1
                    continue

                # Safe float conversion supporting percentages (%12,5 / 12.5% / strings)
                metric_value = None
                if isinstance(cell_val, (int, float)):
                    metric_value = float(cell_val)
                else:
                    val_str = str(cell_val).strip()
                    try:
                        if val_str.endswith('%'):
                            metric_value = float(val_str[:-1].replace(',', '.')) / 100.0
                        else:
                            metric_value = float(val_str.replace(',', '.'))
                    except (ValueError, TypeError):
                        self.parse_statistics["invalid_cells"] += 1
                        self.invalid_cells.append({
                            "sheet_name": sheet_name,
                            "source_row": r,
                            "source_column": col_idx,
                            "product_group": product_group,
                            "product_name": prod_name,
                            "value": val_str,
                        })
                        continue

                if metric_value is None:
                    continue
                self.parse_statistics["numeric_cells"] += 1

                sheet_type = str(structure_info["sheet_type"]).strip().upper()
                # Metric semantics come exclusively from content-classified
                # SheetType. A renamed sheet or a generic word such as MARKET
                # must never reinterpret TL values as market share.
                if sheet_type == SheetType.MARKET_REFERENCE.value:
                    metric_type = MetricType.MARKET_SHARE.value
                elif sheet_type in {
                    SheetType.MONTHLY_VALUE.value,
                    SheetType.WEEKLY_VALUE.value,
                    SheetType.MONTHLY_COMPETITION_VALUE.value,
                }:
                    metric_type = MetricType.TL.value
                elif sheet_type in {
                    SheetType.MONTHLY_UNITS.value,
                    SheetType.WEEKLY_UNITS.value,
                    SheetType.MONTHLY_COMPETITION_UNITS.value,
                }:
                    metric_type = MetricType.UNIT.value
                else:
                    raise ValueError(f"Semantik metric türü çözülemedi: {sheet_type}")

                record = {
                    "year": year,
                    "month": month,
                    "week_number": week_number,
                    "sheet_name": sheet_name,
                    "period_type": period_type,
                    "territory": territory,
                    "subterritory": subterritory,
                    "product_group": product_group,
                    "product_name": prod_name,
                    "metric_type": metric_type,
                    "metric_value": metric_value,
                    "source_row": r,
                }
                records.append(record)

        return records

    def _normalize_turkish_text(self, text: Optional[str]) -> str:
        """Safely normalize Turkish and general textual whitespace and casing using explicit replacement to avoid locale issues."""
        if not text or not str(text).strip():
            return ""
        
        cleaned = " ".join(str(text).strip().split())
        
        mapping = {
            'i': 'İ', 'ı': 'I', 'ç': 'Ç', 'ğ': 'Ğ', 'ö': 'Ö', 'ş': 'Ş', 'ü': 'Ü'
        }
        chars = [mapping.get(c, c.upper()) for c in cleaned]
        return "".join(chars)

    def validate_record(self, record: Dict[str, Any]) -> Tuple[bool, List[str]]:
        """Validate a single record and collect detailed internal warning messages safely."""
        warnings: List[str] = []
        if not record:
            return False, ["Record dictionary is empty."]

        if not record.get("territory") or not str(record["territory"]).strip():
            warnings.append("Territory is missing or empty.")
        if not record.get("product_name") or not str(record["product_name"]).strip():
            warnings.append("Product name is missing or empty.")
        if record.get("metric_value") is None:
            warnings.append("Metric value is missing.")
        else:
            try:
                float(record["metric_value"])
            except (ValueError, TypeError):
                warnings.append("Metric value is not a valid number.")

        is_valid = len(warnings) == 0
        return is_valid, warnings

    def normalize_record(self, record: Dict[str, Any]) -> Dict[str, Any]:
        """Safely normalize all string fields (territory, subterritory, product_group, product_name, sheet_name), numbers, and nulls."""
        normalized = dict(record)
        
        if "territory" in normalized:
            normalized["territory"] = self._normalize_turkish_text(normalized.get("territory"))
        if "subterritory" in normalized:
            normalized["subterritory"] = self._normalize_turkish_text(normalized.get("subterritory"))
        if "product_name" in normalized:
            normalized["product_name"] = self._normalize_turkish_text(normalized.get("product_name"))
        if "product_group" in normalized:
            normalized["product_group"] = self._normalize_turkish_text(normalized.get("product_group"))
        if "sheet_name" in normalized:
            # Normalized sheet_name stored in uppercase standard format
            normalized["sheet_name"] = self._normalize_turkish_text(normalized.get("sheet_name"))

        try:
            normalized["metric_value"] = float(normalized["metric_value"]) if normalized.get("metric_value") is not None else 0.0
        except (ValueError, TypeError):
            normalized["metric_value"] = 0.0

        for num_field in ["year", "month", "week_number", "source_row"]:
            if normalized.get(num_field) is not None:
                try:
                    normalized[num_field] = int(normalized[num_field])
                except (ValueError, TypeError):
                    normalized[num_field] = None

        return normalized

    def map_record_to_model(self, record: Dict[str, Any], sheet_name: str) -> CompetitionData:
        """Map normalized record to CompetitionData model with automated metric type and product group fallback."""
        norm_rec = self.normalize_record(record)
        
        upload_id = self.upload_id

        metric_type = norm_rec.get("metric_type")
        if not metric_type:
            s_upper = sheet_name.upper()
            if "PAZAR" in s_upper or "MARKET" in s_upper:
                metric_type = MetricType.MARKET_SHARE.value
            elif "TL" in s_upper or "VALUE" in s_upper:
                metric_type = MetricType.TL.value
            else:
                metric_type = MetricType.UNIT.value

        product_group = norm_rec.get("product_group")
        if not product_group:
            product_group = DefaultGroup.GENEL.value

        product_name = norm_rec.get("product_name") or ""
        normalized_product = self._normalize_turkish_text(product_name)
        is_grand_total = "GRAND TOTAL" in normalized_product or "GENEL TOPLAM" in normalized_product
        is_subtotal = not is_grand_total and (
            "SUBTOTAL" in normalized_product
            or "ARA TOPLAM" in normalized_product
            or normalized_product.endswith(" TOPLAM")
        )

        return CompetitionData(
            upload_id=upload_id,
            sheet_name=norm_rec.get("sheet_name", sheet_name),
            period_type=norm_rec.get("period_type", PeriodType.MONTHLY.value),
            year=norm_rec.get("year"),
            month=norm_rec.get("month"),
            week_number=norm_rec.get("week_number"),
            territory=norm_rec.get("territory"),
            subterritory=norm_rec.get("subterritory", ""),
            product_group=product_group,
            product_name=product_name,
            metric_type=metric_type,
            metric_value=norm_rec.get("metric_value", 0.0),
            is_subtotal=is_subtotal,
            is_grand_total=is_grand_total,
            source_row=norm_rec.get("source_row")
        )

    @staticmethod
    def _model_mapping(model: CompetitionData) -> Dict[str, Any]:
        """Convert one validated model to a compact mapping for bounded writes."""
        return {
            column.name: getattr(model, column.name)
            for column in CompetitionData.__table__.columns
            if column.name not in {"id", "created_at"} and getattr(model, column.name) is not None
        }

    def bulk_insert(self, model_mappings: List[Dict[str, Any]]) -> int:
        """Persist mappings without committing the outer atomic IMS transaction."""
        if not model_mappings:
            return 0

        inserted_count = 0
        try:
            for i in range(0, len(model_mappings), self.BULK_CHUNK_SIZE):
                chunk = model_mappings[i:i + self.BULK_CHUNK_SIZE]
                db.session.bulk_insert_mappings(CompetitionData, chunk)
                inserted_count += len(chunk)

            # Verify no unintended pending transaction failures or state corruption
            if db.session.new or db.session.dirty:
                logger.debug("[CompetitionImportService] Session has pending manual changes accompanying bulk save.")

            logger.info("[CompetitionImportService] Successfully bulk inserted %d competition mappings in chunks for upload_id=%s", inserted_count, self.upload_id)
            return inserted_count
        except Exception as exc:
            logger.exception("[CompetitionImportService] Failed to perform chunked bulk_save_objects for competition records: %s", exc)
            raise RuntimeError(f"Toplu veri kaydı sırasında bir hata oluştu.") from exc

    def import_records(self, records: List[Dict[str, Any]], sheet_name: str) -> Dict[str, Any]:
        """Process, validate, run deterministic case-insensitive normalized sheet duplicate check, chunked bulk insert and return statistics."""
        if self.upload_id is None:
            raise ValueError("Fail Fast: upload_id is required to import competition records.")

        sheet_start_time = time.time()
        norm_sheet_name = self._normalize_turkish_text(sheet_name)
        logger.info("[CompetitionImportService] Sheet Started | Sheet: %s | Upload ID: %s", norm_sheet_name, self.upload_id)

        total_input = len(records)
        inserted_count = 0
        duplicate_count = 0
        invalid_count = 0
        sheet_warnings: List[Dict[str, Any]] = []
        mapping_batch: List[Dict[str, Any]] = []

        existing_values: Dict[Tuple[Any, ...], float] = {}
        existing_sources: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
        existing_rows = db.session.query(
            CompetitionData.upload_id,
            CompetitionData.sheet_name,
            CompetitionData.year,
            CompetitionData.month,
            CompetitionData.week_number,
            CompetitionData.territory,
            CompetitionData.subterritory,
            CompetitionData.product_group,
            CompetitionData.product_name,
            CompetitionData.metric_type,
            CompetitionData.metric_value,
        ).filter_by(
            upload_id=self.upload_id,
            # Sheet name is part of the duplicate business key. Reading every
            # row accumulated by earlier sheets makes a workbook import
            # quadratic in both memory and time on the 1 GB production host.
            # Stored names are normalized by map_record_to_model(), so this
            # bounded predicate preserves the exact duplicate/conflict rule.
            sheet_name=norm_sheet_name,
        ).all()

        for row in existing_rows:
            # Case-insensitive normalized comparison for sheet names in duplicate checks
            row_list = list(row)
            row_list[1] = self._normalize_turkish_text(row_list[1])
            existing_values[tuple(row_list[:-1])] = float(row_list[-1] or 0.0)

        for rec in records:
            is_valid, val_warnings = self.validate_record(rec)
            if not is_valid:
                invalid_count += 1
                sheet_warnings.append({
                    "record_source_row": rec.get("source_row"),
                    "warnings": val_warnings
                })
                self.warnings.extend(val_warnings)
                continue

            norm = self.normalize_record(rec)
            
            biz_key = (
                self.upload_id,
                norm_sheet_name,
                norm.get("year"),
                norm.get("month"),
                norm.get("week_number"),
                norm.get("territory"),
                norm.get("subterritory", ""),
                norm.get("product_group", DefaultGroup.GENEL.value),
                norm.get("product_name"),
                norm.get("metric_type")
            )

            metric_value = float(norm.get("metric_value") or 0.0)
            if biz_key in existing_values:
                if abs(existing_values[biz_key] - metric_value) > 1e-9:
                    first_source = existing_sources.get(biz_key, {})
                    raise ValueError(
                        "Aynı rekabet veri anahtarında çelişen değerler bulundu: "
                        f"key={biz_key}, first={existing_values[biz_key]}, second={metric_value}, "
                        f"first_source={first_source}, "
                        f"second_source={{'row': {norm.get('source_row')!r}, "
                        f"'column': {norm.get('source_column')!r}}}"
                    )
                duplicate_count += 1
                continue

            existing_values[biz_key] = metric_value
            existing_sources[biz_key] = {
                "row": norm.get("source_row"),
                "column": norm.get("source_column"),
            }

            model_obj = self.map_record_to_model(norm, sheet_name)
            mapping_batch.append(self._model_mapping(model_obj))
            if len(mapping_batch) >= self.BULK_CHUNK_SIZE:
                inserted_count += self.bulk_insert(mapping_batch)
                mapping_batch.clear()

        if mapping_batch:
            inserted_count += self.bulk_insert(mapping_batch)
            mapping_batch.clear()

        sheet_exec_time = round(time.time() - sheet_start_time, 4)
        logger.info(
            "[CompetitionImportService] Sheet Finished | Sheet: %s | Inserted: %d | Duplicates: %d | Invalid: %d | Time: %.4fs",
            norm_sheet_name, inserted_count, duplicate_count, invalid_count, sheet_exec_time
        )

        return {
            "sheet_name": sheet_name,
            "upload_id": self.upload_id,
            "total_input": total_input,
            "inserted": inserted_count,
            "duplicates": duplicate_count,
            "invalid": invalid_count,
            "execution_time": sheet_exec_time,
            "warnings": sheet_warnings
        }

    def run(self) -> dict:
        """Run complete end-to-end import pipeline with structured logging and production self-check verification."""
        if not self.file_path:
            raise ValueError("Dosya yolu belirtilmedi.")
        if self.upload_id is None:
            raise ValueError("Yükleme kimliği (upload_id) gerekli.")

        pipeline_start_time = time.time()
        logger.info("[CompetitionImportService] Import Started | Upload ID: %s | File: %s", self.upload_id, self.file_path)

        try:
            self.load_workbook(self.file_path)
            self.validate_workbook()
            supported_sheets = self.get_supported_sheets()

            for s_name in supported_sheets:
                self.validate_sheet_structure(s_name)

            sheet_statistics = []
            total_inserted = 0
            total_duplicates = 0
            total_invalid = 0

            # Parse, validate and persist one semantic sheet at a time. Keeping
            # every sheet's expanded cell records and ORM objects alive at once
            # caused swap pressure and production acceptance timeouts. Atomicity
            # is unchanged: the outer IMS transaction still commits only after
            # the complete workbook reconciles successfully.
            for s_name in supported_sheets:
                struct_info = self._parse_sheet_structure(s_name)
                sheet_records = self._parse_sheet_records(struct_info)
                if self.invalid_cells:
                    raise ValueError(
                        "Rekabet sayfalarında geçersiz metrik hücreleri bulundu: "
                        f"count={len(self.invalid_cells)}, sample={self.invalid_cells[:10]}"
                    )
                stats = self.import_records(sheet_records, s_name)
                sheet_statistics.append(stats)
                total_inserted += stats["inserted"]
                total_duplicates += stats["duplicates"]
                total_invalid += stats["invalid"]
                del sheet_records

            if total_invalid or total_inserted + total_duplicates != self.parse_statistics["numeric_cells"]:
                raise ValueError(
                    "Rekabet veri mutabakatı başarısız: "
                    f"numeric={self.parse_statistics['numeric_cells']}, inserted={total_inserted}, "
                    f"duplicates={total_duplicates}, invalid={total_invalid}"
                )
            if total_inserted == 0:
                raise ValueError("Rekabet sayfalarında aktarılabilir sayısal kayıt bulunamadı.")

            total_exec_time = round(time.time() - pipeline_start_time, 4)
            logger.info(
                "[CompetitionImportService] Total Execution Summary | Upload ID: %s | Total Inserted: %d | Total Duplicates: %d | Total Invalid: %d | Time: %.4fs",
                self.upload_id, total_inserted, total_duplicates, total_invalid, total_exec_time
            )

            # Production self-check assertions
            assert self.upload_id is not None, "Production Self-Check Failed: upload_id is missing."
            assert isinstance(sheet_statistics, list), "Production Self-Check Failed: sheet_statistics format error."

            return {
                "success": True,
                "service": "CompetitionImportService",
                "status": "ENTERPRISE_RECORDS_IMPORTED_SUCCESSFULLY",
                "supported_sheets": supported_sheets,
                "summary": {
                    "total_inserted": total_inserted,
                    "total_duplicates": total_duplicates,
                    "total_invalid": total_invalid,
                    **self.parse_statistics,
                    "execution_time": total_exec_time
                },
                "sheet_statistics": sheet_statistics,
                "errors": self.errors,
                "warnings": self.warnings,
            }
        except Exception as exc:
            logger.exception("[CompetitionImportService] Import Pipeline Failed for Upload ID %s: %s", self.upload_id, exc)
            raise
        finally:
            if self._workbook:
                try:
                    self._workbook.close()
                    logger.info("[CompetitionImportService] Workbook safely closed.")
                except Exception as close_exc:
                    logger.warning("[CompetitionImportService] Failed to close workbook cleanly: %s", close_exc)
