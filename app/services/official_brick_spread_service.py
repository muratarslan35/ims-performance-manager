"""Authoritative import/query support for the ``Satış Brick Yayılımı`` sheet.

The workbook's spread counts are an official aggregate source.  They must not be
reconstructed from brick-sales facts because the workbook can intentionally
include a wider assignment/distribution scope than rows carrying a sale.

Values are stored in ``IMSRawData`` as a side-channel with ``product_id=NULL``.
That is deliberate: the normal RAW -> FACT transformer ignores rows without a
product id, so brick counts can never contaminate box/TL sales, realizations or
prime calculations.  The original workbook value remains fully traceable via
``raw_json`` and ``source_row``.
"""

from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.extensions import db
from app.models import IMSRawData, IMSUpload, Product
from app.services.alias_service import AliasService


class OfficialBrickSpreadError(ValueError):
    """Raised when an official spread sheet exists but cannot be reconciled."""


class OfficialBrickSpreadService:
    SHEET_TOKEN = "SATIS BRICK YAYILIMI"
    SHEET_TYPE = "official_brick_spread"
    TOTAL_PRODUCT_LABEL = "__TOTAL__"

    @staticmethod
    def _normalize(value: Any) -> str:
        return AliasService.normalize(value)

    @staticmethod
    def _number(value: Any) -> int:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return 0
        if isinstance(value, bool):
            raise OfficialBrickSpreadError("Brick yayılımında boolean metrik kullanılamaz.")
        if isinstance(value, (int, float)):
            number = float(value)
        else:
            text = str(value).strip().replace("\u00a0", "")
            if text in {"", "-"}:
                return 0
            text = re.sub(r"[^0-9,.-]", "", text)
            if text.count(",") and not text.count("."):
                text = text.replace(",", ".")
            elif text.count(",") and text.count("."):
                text = text.replace(",", "")
            try:
                number = float(text)
            except ValueError as exc:
                raise OfficialBrickSpreadError(f"Geçersiz brick yayılım değeri: {value!r}") from exc
        rounded = round(number)
        if abs(number - rounded) > 1e-9 or rounded < 0:
            raise OfficialBrickSpreadError(f"Brick yayılımı negatif veya tam sayı değil: {value!r}")
        return int(rounded)

    @classmethod
    def _sheet_name(cls, workbook) -> str | None:
        for name in workbook.sheetnames:
            if cls.SHEET_TOKEN in cls._normalize(name):
                return name
        return None

    @classmethod
    def _header_row(cls, worksheet) -> int:
        for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=min(20, worksheet.max_row), values_only=True), start=1):
            normalized = [cls._normalize(value) for value in row]
            if any("BRICK SAYISI" in value for value in normalized):
                return row_number
        raise OfficialBrickSpreadError("Satış Brick Yayılımı başlığı (Brick Sayısı) bulunamadı.")

    @classmethod
    def _product_columns(cls, worksheet, header_row: int) -> dict[int, Product]:
        header_values = [cell.value for cell in worksheet[header_row]]
        products = Product.query.filter_by(is_active=True).all()
        columns: dict[int, Product] = {}
        for column_index, value in enumerate(header_values, start=1):
            normalized_header = cls._normalize(value)
            if not normalized_header:
                continue
            candidates = []
            for product in products:
                labels = {
                    cls._normalize(product.product_name),
                    cls._normalize(product.product_code),
                    cls._normalize(getattr(product, "ims_name", "")),
                }
                labels.discard("")
                if any(label == normalized_header or label in normalized_header or normalized_header in label for label in labels):
                    candidates.append(product)
            unique = {product.id: product for product in candidates}
            if len(unique) == 1:
                columns[column_index] = next(iter(unique.values()))
            elif len(unique) > 1:
                names = ", ".join(sorted(product.product_name for product in unique.values()))
                raise OfficialBrickSpreadError(
                    f"'{value}' brick yayılım başlığı birden fazla ürüne eşleşiyor: {names}"
                )

        expected = {product.id for product in products}
        found = {product.id for product in columns.values()}
        missing = [product.product_name for product in products if product.id not in found]
        if expected and found != expected:
            raise OfficialBrickSpreadError(
                "Satış Brick Yayılımı ürün kapsamı eksik/fazla. Eksik ürünler: " + ", ".join(missing or ["—"])
            )
        return columns

    @classmethod
    def persist(
        cls,
        *,
        file_path: str | Path,
        upload_id: int,
        year: int | None = None,
        month: int | None = None,
        week_number: int | None = None,
    ) -> dict:
        """Persist every representative total/product spread count atomically.

        The caller owns the transaction.  This method only flushes.  Any
        unresolved representative, duplicate representative, malformed metric,
        or missing company-product column fails the integration rather than
        silently dropping a master value.
        """
        upload = db.session.get(IMSUpload, int(upload_id))
        if upload is None:
            raise OfficialBrickSpreadError(f"IMS upload bulunamadı: {upload_id}")

        year = int(year or upload.year)
        month = int(month or upload.month)
        week_number = week_number if week_number is not None else upload.week_number
        quarter = upload.quarter or f"Q{((month - 1) // 3) + 1}"

        workbook = load_workbook(str(file_path), read_only=True, data_only=True)
        try:
            sheet_name = cls._sheet_name(workbook)
            if not sheet_name:
                raise OfficialBrickSpreadError("Satış Brick Yayılımı master sayfası bulunamadı.")
            worksheet = workbook[sheet_name]
            header_row = cls._header_row(worksheet)
            product_columns = cls._product_columns(worksheet, header_row)

            # Idempotent for retries/backfills on the same upload.
            IMSRawData.query.filter_by(upload_id=upload.id, sheet_type=cls.SHEET_TYPE).delete(synchronize_session=False)

            seen_representatives: set[int] = set()
            inserted = 0
            matched_representatives = 0
            aggregate_rows = 0
            unresolved = []

            for source_row, values in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, max_row=worksheet.max_row, values_only=True),
                start=header_row + 1,
            ):
                if not values or not any(value is not None and str(value).strip() for value in values):
                    continue
                region = str(values[0]).strip() if len(values) > 0 and values[0] is not None else ""
                representative_name = str(values[1]).strip() if len(values) > 1 and values[1] is not None else ""
                normalized_rep = cls._normalize(representative_name)
                if not normalized_rep:
                    continue

                match = AliasService.find_representative(representative_name)
                if not match.get("matched"):
                    # NATIONAL and region subtotal rows are useful workbook
                    # checks but are not representative-level master records.
                    if normalized_rep == "NATIONAL" or re.match(r"^\d{3}\s+", normalized_rep):
                        aggregate_rows += 1
                        continue
                    metric_cells = [values[2] if len(values) > 2 else None]
                    metric_cells.extend(values[index - 1] if len(values) >= index else None for index in product_columns)
                    if any(cell not in (None, "", "-") for cell in metric_cells):
                        unresolved.append({"row": source_row, "representative": representative_name})
                    continue

                representative = match["object"]
                if representative.id in seen_representatives:
                    raise OfficialBrickSpreadError(
                        f"Satış Brick Yayılımı içinde temsilci tekrarı: {representative.rep_name} (satır {source_row})"
                    )
                seen_representatives.add(representative.id)
                matched_representatives += 1

                total_count = cls._number(values[2] if len(values) > 2 else 0)
                base_payload = {
                    "authoritative": True,
                    "source": "workbook_master",
                    "sheet": sheet_name,
                    "source_row": source_row,
                    "region": region,
                    "representative": representative_name,
                    "metric": "brick_count",
                }
                db.session.add(IMSRawData(
                    upload_id=upload.id,
                    year=year,
                    month=month,
                    week_number=week_number,
                    quarter=quarter,
                    sheet_name=sheet_name,
                    sheet_type=cls.SHEET_TYPE,
                    source_row=source_row,
                    representative_id=representative.id,
                    product_id=None,
                    representative=representative_name,
                    product=cls.TOTAL_PRODUCT_LABEL,
                    territory=region or representative.region,
                    unit=float(total_count),
                    tl=0.0,
                    market_share=0.0,
                    value_share=0.0,
                    growth=0.0,
                    raw_json=json.dumps({**base_payload, "scope": "representative_total", "brick_count": total_count}, ensure_ascii=False, sort_keys=True),
                ))
                inserted += 1

                for column_index, product in product_columns.items():
                    raw_value = values[column_index - 1] if len(values) >= column_index else None
                    product_count = cls._number(raw_value)
                    db.session.add(IMSRawData(
                        upload_id=upload.id,
                        year=year,
                        month=month,
                        week_number=week_number,
                        quarter=quarter,
                        sheet_name=sheet_name,
                        sheet_type=cls.SHEET_TYPE,
                        source_row=source_row,
                        representative_id=representative.id,
                        product_id=None,
                        representative=representative_name,
                        product=product.product_name,
                        territory=region or representative.region,
                        unit=float(product_count),
                        tl=0.0,
                        market_share=0.0,
                        value_share=0.0,
                        growth=0.0,
                        raw_json=json.dumps({
                            **base_payload,
                            "scope": "representative_product",
                            "product_id": product.id,
                            "product_name": product.product_name,
                            "brick_count": product_count,
                        }, ensure_ascii=False, sort_keys=True),
                    ))
                    inserted += 1

            if unresolved:
                sample = ", ".join(f"satır {item['row']}: {item['representative']}" for item in unresolved[:10])
                raise OfficialBrickSpreadError(
                    f"Satış Brick Yayılımı master satırlarında eşleşmeyen temsilci var ({len(unresolved)}): {sample}"
                )
            if matched_representatives == 0:
                raise OfficialBrickSpreadError("Satış Brick Yayılımı içinde hiçbir temsilci eşleştirilemedi.")

            db.session.flush()
            return {
                "sheet_name": sheet_name,
                "representatives": matched_representatives,
                "product_columns": len(product_columns),
                "records": inserted,
                "aggregate_rows_ignored": aggregate_rows,
            }
        finally:
            workbook.close()

    @classmethod
    def for_representative(cls, *, upload_id: int, representative_id: int) -> dict | None:
        """Return authoritative spread counts, or ``None`` for legacy uploads."""
        rows = IMSRawData.query.filter_by(
            upload_id=upload_id,
            representative_id=representative_id,
            sheet_type=cls.SHEET_TYPE,
        ).all()
        if not rows:
            return None
        result = {"total": None, "products": {}, "source": "official_workbook_master"}
        for row in rows:
            try:
                payload = json.loads(row.raw_json or "{}")
            except (TypeError, ValueError, json.JSONDecodeError):
                payload = {}
            count = int(round(row.unit or 0))
            if row.product == cls.TOTAL_PRODUCT_LABEL or payload.get("scope") == "representative_total":
                result["total"] = count
            else:
                result["products"][payload.get("product_name") or row.product] = count
        return result
