"""Independent Target Import Service for processing and upserting target records."""

import json
import logging
import math
import re
import time

import pandas as pd
from openpyxl import load_workbook as openpyxl_load_workbook

from app.extensions import db
from app.models import (
    Target,
)
from app.services.alias_service import AliasService

logger = logging.getLogger(__name__)


class TargetImportService:
    """Service to independently parse and upsert target records from workbooks."""

    TARGET_SHEET_KEYWORDS = {
        "HEDEF",
        "TARGET",
        "HEDEFLER",
        "TARGETS",
        "MONTHLY TARGET",
        "ANNUAL TARGET",
        "YILLIK HEDEF",
    }
    REPRESENTATIVE_HEADERS = {
        "TEMSILCI",
        "REPRESENTATIVE",
        "MUMESSIL",
        "REP",
        "ADI SOYADI",
        "SATIS TEMSILCISI",
        "TTS ISMI",
    }
    PRODUCT_GROUP_HEADERS = {"URUN GRUBU", "PRODUCT GROUP", "MARKA", "URUN", "PRODUCT"}
    PRODUCT_HEADER_NOISE_TOKENS = {
        "SUBTOTAL",
        "TOPLAM",
        "TOTAL",
        "PAZAR",
        "MARKET",
        "GRUP",
        "GROUP",
        "HEDEF",
        "CIKIS",
        "ÇIKIŞ",
        "NATIONAL",
        "GRAND",
    }
    REPRESENTATIVE_NOISE_TOKENS = {
        "SUBTOTAL",
        "TOPLAM",
        "TOTAL",
        "NATIONAL",
        "GRAND",
        "BOS",
        "KADRO",
        "BRICK",
    }
    STRICT_PRODUCT_MATCH_METHODS = {
        "MATCH_TABLE",
        "EXACT",
        "ALIAS_EXACT",
        "NORMALIZED",
        "ALIAS_NORMALIZED",
        "CONTAINS",
        "ALIAS_CONTAINS",
    }
    TOTAL_LABELS = {"NATIONAL", "TOPLAM", "GRAND TOTAL", "TOTAL", "GENEL TOPLAM"}
    NOISE_ROW_TOKENS = {
        "NOT",
        "NOTE",
        "ACIKLAMA",
        "AÇIKLAMA",
        "COMMENTS",
        "YORUM",
        "BILGI",
        "BİLGİ",
    }
    MAX_WARNINGS_PER_SHEET = 100
    def __init__(self, file_path: str, upload_id: int, workbook: dict | None = None):
        self.file_path = str(file_path)
        self.upload_id = upload_id
        self.started = time.monotonic()
        self.workbook = {}
        self._preloaded_workbook = workbook
        self.errors = []
        self.warnings = []
        self.statistics = {
            "target_records": 0,
            "targets_inserted": 0,
            "targets_updated": 0,
            "targets_skipped": 0,
            "targets_errors": 0,
            "processing_time": 0.0,
            "processed_sheets": 0,
            "target_sheets": 0,
            "parsed_rows": 0,
            "matched_products": 0,
            "matched_representatives": 0,
            "unmatched_products": 0,
            "unmatched_representatives": 0,
            "duplicate_headers": 0,
            "duplicate_rows": 0,
            "warnings_count": 0,
        }
        self.parser_decisions = []
        self._representative_match_cache = {}
        self._product_match_cache = {}
        self._sheet_raw_frames = {}
        self._sheet_warnings = {}

    @staticmethod
    def quarter_for(month: int) -> str:
        if month < 1 or month > 12:
            raise ValueError("Ay değeri 1 ile 12 arasında olmalıdır.")
        return f"Q{((month - 1) // 3) + 1}"

    @staticmethod
    def safe_float(value) -> float:
        if value is None or (isinstance(value, float) and (math.isnan(value) or math.isinf(value))):
            return 0.0
        if isinstance(value, bool):
            return float(value)
        if isinstance(value, (int, float)):
            return float(value)

        text = str(value).strip().replace("\u00a0", "")
        if not text or text.upper() in {"NAN", "NONE", "-"}:
            return 0.0
        text = re.sub(r"[^0-9,.-]", "", text)
        if text.count(",") and text.count("."):
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif text.count(","):
            text = text.replace(".", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return 0.0

    @staticmethod
    def clean_text(value) -> str:
        if value is None or (isinstance(value, float) and math.isnan(value)):
            return ""
        text = str(value).strip()
        return "" if AliasService.normalize(text) in {"", "NAN", "NONE"} else text

    def _log_warning(self, reason: str, sheet_name: str, source_row: int, **context) -> None:
        sheet_warnings = self._sheet_warnings.setdefault(sheet_name, [])
        if len(sheet_warnings) >= self.MAX_WARNINGS_PER_SHEET:
            if len(sheet_warnings) == self.MAX_WARNINGS_PER_SHEET:
                sheet_warnings.append(json.dumps({
                    "reason": "warnings_suppressed",
                    "sheet_name": sheet_name,
                    "source_row": source_row,
                    "message": f"Further warnings suppressed after {self.MAX_WARNINGS_PER_SHEET} rows.",
                }))
            return
        payload = {
            "reason": reason,
            "sheet_name": sheet_name,
            "source_row": source_row,
            "header": context.get("header"),
            "product": context.get("product"),
            "representative": context.get("representative"),
            **{k: v for k, v in context.items() if k not in ("header", "product", "representative")}
        }
        payload_json = json.dumps(payload, ensure_ascii=False, default=str)
        self.warnings.append(payload_json)
        sheet_warnings.append(payload_json)
        self.statistics["warnings_count"] += 1
        logger.warning("target_import_warning %s", payload_json)

    def load_workbook(self) -> dict:
        """Load workbook once, keeping reference without redundant dataframe copies."""
        if self._preloaded_workbook is not None:
            self.workbook = self._preloaded_workbook
            self._sheet_raw_frames = self._preloaded_workbook
            return self.workbook
        try:
            raw_workbook = pd.read_excel(self.file_path, sheet_name=None, header=None)
        except Exception as exc:
            logger.error(f"target_import_service: Failed to read excel workbook: {exc}", exc_info=True)
            raise

        hidden_by_sheet = {}
        workbook_obj = None
        try:
            workbook_obj = openpyxl_load_workbook(self.file_path, data_only=True, read_only=False)
            for worksheet in workbook_obj.worksheets:
                hidden_rows = {
                    row_number - 1
                    for row_number, dimensions in worksheet.row_dimensions.items()
                    if getattr(dimensions, "hidden", False)
                }
                hidden_by_sheet[str(worksheet.title)] = hidden_rows
        except Exception as exc:
            logger.warning(f"target_import_service: Could not read hidden rows dimensions. {exc}")
        finally:
            if workbook_obj is not None:
                workbook_obj.close()

        for sheet_name, dataframe in raw_workbook.items():
            s_name = str(sheet_name)
            self._sheet_raw_frames[s_name] = dataframe
            hidden_rows = hidden_by_sheet.get(s_name, set())
            
            if hidden_rows:
                prepared = dataframe.drop(index=[idx for idx in hidden_rows if idx in dataframe.index], errors="ignore")
            else:
                prepared = dataframe
            prepared = prepared.dropna(axis=0, how="all").dropna(axis=1, how="all")
            prepared.reset_index(drop=True, inplace=True)
            self.workbook[s_name] = prepared

        self.statistics["processed_sheets"] = len(self.workbook)
        return self.workbook

    def _find_header_row(self, dataframe: pd.DataFrame):
        max_rows = min(80, len(dataframe))
        best_index = None
        best_score = -1
        for index in range(max_rows):
            values = [AliasService.normalize(value) for value in dataframe.iloc[index].tolist()]
            row_text = " ".join(values)
            score = 0
            if any(candidate in row_text for candidate in self.REPRESENTATIVE_HEADERS):
                score += 3
            if any(candidate in row_text for candidate in self.PRODUCT_GROUP_HEADERS):
                score += 2
            if "TL" in row_text or "KUTU" in row_text or "HEDEF" in row_text or "TARGET" in row_text or "VALUES REPORT" in row_text or "UNITS REPORT" in row_text:
                score += 3
            if score > best_score:
                best_index = index
                best_score = score
        if best_score >= 3:
            return best_index
        return None

    def _detect_header_span(self, dataframe: pd.DataFrame, header_row: int) -> int:
        """Dynamically determine header row height/span upwards from header_row."""
        span = 1
        for idx in range(header_row - 1, max(-1, header_row - 5), -1):
            row_vals = [self.clean_text(v) for v in dataframe.iloc[idx].values if v is not None]
            non_empty = [v for v in row_vals if v and AliasService.normalize(v) not in {"NAN", "NONE"}]
            if non_empty:
                span += 1
            else:
                break
        return span

    def _build_headers(self, dataframe: pd.DataFrame, header_row: int) -> pd.DataFrame:
        headers = []
        used_headers = {}
        span = self._detect_header_span(dataframe, header_row)
        header_rows = [row_index for row_index in range(max(0, header_row - span + 1), header_row + 1)]

        for column_index in range(dataframe.shape[1]):
            parts = []
            for row_index in header_rows:
                token = AliasService.normalize(dataframe.iloc[row_index, column_index])
                if token and token not in parts:
                    parts.append(token)

            header = " ".join(parts) or f"COLUMN_{column_index + 1}"
            duplicate_count = used_headers.get(header, 0)
            if duplicate_count > 0:
                self.statistics["duplicate_headers"] += 1
            used_headers[header] = duplicate_count + 1
            headers.append(header if duplicate_count == 0 else f"{header}_{duplicate_count + 1}")

        result = dataframe.iloc[header_row + 1 :].copy()
        result.columns = headers
        result.reset_index(drop=True, inplace=True)
        return result

    def _extract_product_name(self, header: str) -> str:
        """Extract clean product name via tokenization and noise removal without hardcoding product names."""
        cleaned = str(header)
        for token in ["TL", "CIRO", "TUTAR", "VALUE", "KUTU", "BOX", "ADET", "UNIT", "PP", "PAY", "MARKET SHARE", "VALUE SHARE", "HEDEF", "TARGET", "VALUES REPORT", "UNITS REPORT", "VALUE REPORT", "UNIT REPORT", "GROWTH", "BUYUME"]:
            cleaned = re.sub(rf"\b{token}\b", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\s+", " ", cleaned).strip()
        return cleaned

    def _detect_target_metric(self, header: str) -> str:
        """Detect target metric type dynamically from multi-row header hierarchy."""
        norm = AliasService.normalize(header)
        if any(token in norm for token in ["TL", "CIRO", "TUTAR", "VALUE", "VALUES REPORT", "VALUE REPORT"]):
            return "tl"
        if any(token in norm for token in ["KUTU", "BOX", "ADET", "UNIT", "UNITS REPORT", "UNIT REPORT"]):
            return "unit"
        if any(token in norm for token in ["MARKET SHARE", "VALUE SHARE"]):
            return "value_share"
        if any(token in norm for token in ["PP", "PAY"]):
            return "market_share"
        if any(token in norm for token in ["GROWTH", "BUYUME"]):
            return "growth"
        return "unknown"

    def _is_target_sheet(self, sheet_name: str, dataframe: pd.DataFrame) -> bool:
        """Evaluate sheet name and multi-row content against scoring rules to determine if it is a target sheet."""
        score = 0
        normalized_name = AliasService.normalize(sheet_name)
        # A generic sales/realisation layout can contain a cell named HEDEF.
        # Unless it is the compact layout handled above, require the sheet's
        # own semantic label to identify it as a target source.
        if "HEDEF" not in normalized_name and "TARGET" not in normalized_name:
            return False
        
        for kw in self.TARGET_SHEET_KEYWORDS:
            if AliasService.normalize(kw) in normalized_name:
                score += 5

        max_rows = min(15, len(dataframe))
        sheet_text = ""
        for idx in range(max_rows):
            sheet_text += " " + " ".join(str(val) for val in dataframe.iloc[idx].values if val is not None)
        normalized_text = AliasService.normalize(sheet_text)

        if any(h in normalized_text for h in self.REPRESENTATIVE_HEADERS):
            score += 3
        if any(p in normalized_text for p in self.PRODUCT_GROUP_HEADERS):
            score += 2
        has_target_marker = "HEDEF" in normalized_text or "TARGET" in normalized_text
        if has_target_marker:
            score += 3
        if "TL" in normalized_text or "CIRO" in normalized_text or "VALUES REPORT" in normalized_text:
            score += 2
        if "KUTU" in normalized_text or "ADET" in normalized_text or "UNITS REPORT" in normalized_text:
            score += 2

        # Sales/realization sheets also contain representative, TL and unit
        # labels. They must not be parsed as targets without an explicit
        # target marker, otherwise the same workbook is needlessly processed
        # twice and conflicting target rows are produced.
        is_target = has_target_marker and score >= 6
        if not is_target and score > 0:
            self._log_warning("sheet_below_target_threshold", sheet_name, 0, score=score)
        return is_target

    def _is_compact_target_layout(self, dataframe: pd.DataFrame) -> bool:
        """Detect a target table from its cells, independent of sheet/month name."""
        if len(dataframe.index) < 3 or len(dataframe.columns) < 3:
            return False
        title = AliasService.normalize(" ".join(str(value) for value in dataframe.iloc[0].values if value is not None))
        product_matches = sum(
            AliasService.find_product(self.clean_text(value))["matched"]
            for value in dataframe.iloc[1].values[2:]
            if self.clean_text(value)
        )
        return ("HEDEF" in title or "TARGET" in title) and product_matches >= 2

    def _detect_representative_column(self, dataframe: pd.DataFrame):
        for column in dataframe.columns:
            normalized = AliasService.normalize(column)
            if normalized in self.REPRESENTATIVE_HEADERS or any(candidate in normalized for candidate in self.REPRESENTATIVE_HEADERS if len(candidate) > 3):
                return column
        return None

    def _detect_product_group_column(self, dataframe: pd.DataFrame):
        for column in dataframe.columns:
            normalized = AliasService.normalize(column)
            if any(token in normalized for token in self.PRODUCT_GROUP_HEADERS):
                return column
        return None

    def _detect_product_columns(self, sheet_name: str, dataframe: pd.DataFrame, representative_column: str, header_row: int) -> dict:
        """Dynamically detect product and metric columns using single source of truth col['metric']."""
        products = {}
        raw_df = self._sheet_raw_frames.get(sheet_name)
        span = self._detect_header_span(raw_df, header_row) if raw_df is not None else 1
        
        for column_index, header in enumerate(dataframe.columns):
            if header == representative_column:
                continue
            normalized_header = AliasService.normalize(str(header))
            if any(token in normalized_header for token in self.PRODUCT_HEADER_NOISE_TOKENS):
                continue
            
            hierarchical_tokens = [str(header)]
            if raw_df is not None and header_row is not None:
                for r_idx in range(max(0, header_row - span + 1), header_row + 1):
                    val = raw_df.iloc[r_idx, column_index]
                    cleaned_val = self.clean_text(val)
                    if cleaned_val and AliasService.normalize(cleaned_val) not in hierarchical_tokens:
                        hierarchical_tokens.insert(0, cleaned_val)

            combined_header = " ".join(hierarchical_tokens)
            extracted_name = self._extract_product_name(combined_header)
            
            match = self._resolve_product_match(extracted_name if extracted_name else str(header))
            if not match["matched"] or match["method"] not in self.STRICT_PRODUCT_MATCH_METHODS:
                for token in hierarchical_tokens:
                    sub_extracted = self._extract_product_name(token)
                    if sub_extracted:
                        match = self._resolve_product_match(sub_extracted)
                        if match["matched"] and match["method"] in self.STRICT_PRODUCT_MATCH_METHODS:
                            break
                if not match["matched"] or match["method"] not in self.STRICT_PRODUCT_MATCH_METHODS:
                    self._log_warning("product_not_matched", sheet_name, header_row or 0, header=str(header))
                    continue

            product = match["object"]
            product_info = products.setdefault(
                product.id,
                {"product": product, "columns": []},
            )
            metric_type = self._detect_target_metric(combined_header)
            if metric_type == "unknown":
                metric_type = self._detect_target_metric(str(header))
            if metric_type == "unknown":
                self._log_warning("metric_not_detected", sheet_name, header_row or 0, header=str(header))
                metric_type = "unit"

            product_info["columns"].append(
                {
                    "index": column_index,
                    "header": str(header),
                    "metric": metric_type,
                }
            )
        return products

    def _resolve_representative_match(self, representative_name: str) -> dict:
        normalized = AliasService.normalize(representative_name)
        if normalized not in self._representative_match_cache:
            res = AliasService.find_representative(representative_name)
            self._representative_match_cache[normalized] = res
            if res["matched"]:
                self.statistics["matched_representatives"] += 1
            else:
                self.statistics["unmatched_representatives"] += 1
        return self._representative_match_cache[normalized]

    def _resolve_product_match(self, product_group_name: str) -> dict:
        normalized = AliasService.normalize(product_group_name)
        if normalized not in self._product_match_cache:
            res = AliasService.find_product(product_group_name)
            self._product_match_cache[normalized] = res
            if res["matched"]:
                self.statistics["matched_products"] += 1
            else:
                self.statistics["unmatched_products"] += 1
        return self._product_match_cache[normalized]

    def _is_probable_representative_name(self, text: str) -> bool:
        normalized = AliasService.normalize(text)
        if not normalized or normalized in self.TOTAL_LABELS or normalized == "NATIONAL":
            return False
        if bool(re.search(r"\d", normalized)):
            return False
        tokens = normalized.split()
        if any(token in self.REPRESENTATIVE_NOISE_TOKENS for token in tokens):
            return False
        if len(tokens) < 2:
            return False
        return bool(re.search(r"[A-ZÇĞİÖŞÜ]", normalized))

    def _upsert_target(self, target_map: dict, pending_targets: list, representative_id: int, product_id: int, year: int, month: int, quarter: str, unit_target: float, tl_target: float) -> None:
        """Centralized helper method to perform insert or update on target records."""
        t_key = (representative_id, product_id)
        existing = target_map.get(t_key)

        if existing:
            existing.unit_target = unit_target
            existing.tl_target = tl_target
            self.statistics["targets_updated"] += 1
        else:
            new_target = Target(
                representative_id=representative_id,
                product_id=product_id,
                year=year,
                month=month,
                quarter=quarter,
                unit_target=unit_target,
                tl_target=tl_target,
            )
            pending_targets.append(new_target)
            target_map[t_key] = new_target
            self.statistics["targets_inserted"] += 1
        self.statistics["target_records"] += 1

    def _parse_wide_row(self, row, products, representative_id, year, month, quarter, target_map, pending_targets) -> bool:
        """Process a wide format target row utilizing single source of truth col['metric']."""
        matched_any = False
        for product_info in products.values():
            unit_target = 0.0
            tl_target = 0.0
            for col in product_info["columns"]:
                val, valid = self.safe_float(row.iloc[col["index"]]), True
                if not valid:
                    continue
                metric = col.get("metric", "unit")
                if metric == "tl":
                    tl_target += val
                else:
                    unit_target += val

            if unit_target == 0.0 and tl_target == 0.0:
                continue

            product_id = product_info["product"].id
            self._upsert_target(target_map, pending_targets, representative_id, product_id, year, month, quarter, unit_target, tl_target)
            matched_any = True
        return matched_any

    def _parse_normalized_row(self, row, dataframe, representative_column, product_group_column, representative_id, year, month, quarter, target_map, pending_targets, metric_map=None) -> bool:
        """Process a normalized format target row utilizing col['metric'] single source of truth without re-detecting metric."""
        prod_group_name = self.clean_text(row[product_group_column])
        if not prod_group_name:
            return False
        product_match = self._resolve_product_match(prod_group_name)
        if not product_match["matched"]:
            self.statistics["targets_skipped"] += 1
            self.statistics["unmatched_products"] += 1
            return False
        product_id = product_match["object"].id

        unit_target = 0.0
        tl_target = 0.0
        for col_name in dataframe.columns:
            if col_name == representative_column or col_name == product_group_column:
                continue
            val = self.safe_float(row[col_name])
            if val == 0.0:
                continue
            
            # SSOT: Retrieve pre-mapped metric from metric_map or default to unit
            m_type = metric_map.get(col_name, "unit") if metric_map else "unit"
            if m_type == "tl":
                tl_target += val
            else:
                unit_target += val

        if unit_target == 0.0 and tl_target == 0.0:
            return False

        self._upsert_target(target_map, pending_targets, representative_id, product_id, year, month, quarter, unit_target, tl_target)
        return True

    def _process_tts_target_sheet(self, sheet_name, raw_df, year, month, quarter, target_map, pending_targets):
        """Parse the workbook's compact ``HAZİRAN HEDEF`` layout.

        Its representative names are in column B and product names in row 2;
        there is no textual representative header, so generic header discovery
        cannot safely identify it.
        """
        self.statistics["target_sheets"] += 1
        if len(raw_df.index) < 3 or len(raw_df.columns) < 3:
            self._log_warning("target_layout_too_small", sheet_name, 0)
            return
        product_columns = []
        # The merged ``HAZİRAN HEDEF`` label begins one column before the
        # first product (column B), so seed the horizontal carry from there.
        active_section = AliasService.normalize(self.clean_text(raw_df.iloc[0, 1]))
        for column_index in range(2, len(raw_df.columns)):
            # Hedef, Çıkış and REAL% product blocks coexist on this compact
            # sheet.  Carry Excel's merged section label rightward and use
            # only the actual target block, never the later percentage data.
            section_label = self.clean_text(raw_df.iloc[0, column_index])
            if section_label:
                active_section = AliasService.normalize(section_label)
            if "HEDEF" not in active_section and "TARGET" not in active_section:
                continue
            product_name = self.clean_text(raw_df.iloc[1, column_index])
            if not product_name or AliasService.normalize(product_name) in {"TOPLAM", "TOTAL"}:
                continue
            match = self._resolve_product_match(product_name)
            if match["matched"]:
                product_columns.append((column_index, match["object"].id))
        for row_index in range(2, len(raw_df.index)):
            rep_name = self.clean_text(raw_df.iloc[row_index, 1])
            if not self._is_probable_representative_name(rep_name):
                self.statistics["targets_skipped"] += 1
                continue
            rep_match = self._resolve_representative_match(rep_name)
            if not rep_match["matched"]:
                self.statistics["targets_skipped"] += 1
                self._log_warning("representative_not_matched", sheet_name, row_index + 1, representative=rep_name)
                continue
            for column_index, product_id in product_columns:
                value = self.safe_float(raw_df.iloc[row_index, column_index])
                if value:
                    self._upsert_target(
                        target_map, pending_targets, rep_match["object"].id,
                        product_id, year, month, quarter, 0.0, value,
                    )
        self.parser_decisions.append({"sheet_name": sheet_name, "parse_mode": "tts_target_compact", "product_columns": len(product_columns)})

    def _process_sheet(self, sheet_name: str, raw_df: pd.DataFrame, year: int, month: int, quarter: str, target_map: dict, pending_targets: list) -> None:
        """Process a single worksheet for targets with O(1) sheet warnings and precise duplicate classification."""
        if self._is_compact_target_layout(raw_df):
            self._process_tts_target_sheet(sheet_name, raw_df, year, month, quarter, target_map, pending_targets)
            return

        if not self._is_target_sheet(sheet_name, raw_df):
            return

        self.statistics["target_sheets"] += 1
        header_row = self._find_header_row(raw_df)
        if header_row is None:
            self._log_warning("header_row_not_found", sheet_name, 0)
            return

        dataframe = self._build_headers(raw_df, header_row)
        representative_column = self._detect_representative_column(dataframe)
        if representative_column is None:
            self._log_warning("representative_column_not_found", sheet_name, header_row)
            return

        product_group_column = self._detect_product_group_column(dataframe)
        products = self._detect_product_columns(sheet_name, dataframe, representative_column, header_row)

        # SSOT metric map for normalized mode columns
        norm_metric_map = {}
        for col_name in dataframe.columns:
            if col_name != representative_column and col_name != product_group_column:
                norm_metric_map[col_name] = self._detect_target_metric(str(col_name))

        sheet_rep_prod_signatures = {}
        sheet_parsed_rows = 0

        for dataframe_index, (_, row) in enumerate(dataframe.iterrows()):
            try:
                source_row = dataframe_index + header_row + 2
                rep_name = self.clean_text(row[representative_column])
                if not rep_name or not self._is_probable_representative_name(rep_name):
                    self.statistics["targets_skipped"] += 1
                    continue

                rep_match = self._resolve_representative_match(rep_name)
                if not rep_match["matched"]:
                    self.statistics["targets_skipped"] += 1
                    self.statistics["unmatched_representatives"] += 1
                    self._log_warning("representative_not_matched", sheet_name, source_row, representative=rep_name)
                    continue
                representative_id = rep_match["object"].id

                if products:
                    for p_info in products.values():
                        p_id = p_info["product"].id
                        pair_key = (representative_id, p_id)

                        row_vals_signature = tuple(self.safe_float(row.iloc[c["index"]]) for c in p_info["columns"])
                        
                        if pair_key in sheet_rep_prod_signatures:
                            self.statistics["duplicate_rows"] += 1
                            prev_vals = sheet_rep_prod_signatures[pair_key]
                            if prev_vals == row_vals_signature:
                                self._log_warning("exact_duplicate_row", sheet_name, source_row, representative=rep_name, product=p_info["product"].product_name)
                            else:
                                self._log_warning("conflicting_duplicate_row", sheet_name, source_row, representative=rep_name, product=p_info["product"].product_name)
                        else:
                            sheet_rep_prod_signatures[pair_key] = row_vals_signature

                sheet_parsed_rows += 1
                self.statistics["parsed_rows"] += 1

                if products:
                    self._parse_wide_row(row, products, representative_id, year, month, quarter, target_map, pending_targets)
                elif product_group_column:
                    self._parse_normalized_row(row, dataframe, representative_column, product_group_column, representative_id, year, month, quarter, target_map, pending_targets, metric_map=norm_metric_map)
                else:
                    self._log_warning("no_product_columns_or_group", sheet_name, source_row)
            except Exception as row_exc:
                self.statistics["targets_errors"] += 1
                logger.warning(f"target_import_service row error: {row_exc}")
                self._log_warning("row_processing_error", sheet_name, dataframe_index, error=str(row_exc))
                continue

        sheet_warns = self._sheet_warnings.get(sheet_name, [])
        self.parser_decisions.append({
            "sheet_name": sheet_name,
            "header_row": header_row,
            "representative_column": str(representative_column),
            "product_columns": len(products),
            "metric_columns": sum(len(p["columns"]) for p in products.values()),
            "parse_mode": "wide" if products else ("normalized" if product_group_column else "unknown"),
            "found_products": len(products),
            "found_representatives": sheet_parsed_rows,
            "warnings": sheet_warns
        })

    def run(self, year: int, month: int) -> dict:
        """Execute the target import pipeline independently and return metrics dictionary."""
        AliasService.warmup()
        try:
            self.load_workbook()
        except Exception as exc:
            logger.error(f"TargetImportService failed to load workbook: {exc}", exc_info=True)
            self.errors.append(str(exc))
            self.statistics["targets_errors"] += 1
            return self.report()

        quarter = self.quarter_for(month)
        existing_targets = Target.query.filter_by(year=year, month=month).all()
        target_map = {(t.representative_id, t.product_id): t for t in existing_targets}
        pending_targets = []

        for sheet_name, raw_df in self.workbook.items():
            raw_original = self._sheet_raw_frames.get(sheet_name, raw_df)
            self._process_sheet(sheet_name, raw_original, year, month, quarter, target_map, pending_targets)

        if pending_targets:
            db.session.add_all(pending_targets)

        try:
            db.session.flush()
        except Exception as commit_exc:
            db.session.rollback()
            logger.error(f"TargetImportService database flush failed: {commit_exc}", exc_info=True)
            self.errors.append(str(commit_exc))
            self.statistics["targets_errors"] += 1

        self.statistics["processing_time"] = round(time.monotonic() - self.started, 2)
        return self.report()

    def report(self, debug: bool = False) -> dict:
        """Return statistics and optional debug collections (parser_decisions, warnings, errors)."""
        self.statistics["processing_time"] = round(time.monotonic() - self.started, 2)
        result = {
            "target_records": self.statistics["target_records"],
            "targets_inserted": self.statistics["targets_inserted"],
            "targets_updated": self.statistics["targets_updated"],
            "targets_skipped": self.statistics["targets_skipped"],
            "targets_errors": self.statistics["targets_errors"],
            "processing_time": self.statistics["processing_time"],
            "processed_sheets": self.statistics["processed_sheets"],
            "target_sheets": self.statistics["target_sheets"],
            "parsed_rows": self.statistics["parsed_rows"],
            "matched_products": self.statistics["matched_products"],
            "matched_representatives": self.statistics["matched_representatives"],
            "unmatched_products": self.statistics["unmatched_products"],
            "unmatched_representatives": self.statistics["unmatched_representatives"],
            "duplicate_headers": self.statistics["duplicate_headers"],
            "duplicate_rows": self.statistics["duplicate_rows"],
            "warnings_count": self.statistics["warnings_count"],
        }
        if debug:
            result["parser_decisions"] = self.parser_decisions
            result["warnings"] = self.warnings
            result["errors"] = self.errors
        return result
